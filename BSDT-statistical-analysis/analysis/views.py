import os
import matplotlib  as mpl
import uuid
import matplotlib.font_manager as fm
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import scipy.stats as stats
import seaborn as sns
from sklearn.preprocessing import OrdinalEncoder
from django.views.decorators.csrf import csrf_exempt
import json
mpl.use('Agg') 
from itertools import combinations
import networkx as nx
import statsmodels.api as sm
import statsmodels.formula.api as smf
from django.conf import settings
from django.http import JsonResponse
from django.shortcuts import render
from googletrans import Translator
from PIL import Image, ImageDraw, ImageFont
from scipy.stats import chi2_contingency
from sklearn.linear_model import LinearRegression
from sklearn.metrics import mean_squared_error, r2_score
from sklearn.preprocessing import OrdinalEncoder
from .forms import AnalysisForm
from scipy.stats import shapiro, anderson
from matplotlib.patches import Patch
from scipy.stats import mannwhitneyu, rankdata

def infer_type(s, thresh=10):
    if pd.api.types.is_numeric_dtype(s):
        return 'categorical' if s.nunique() < thresh else 'numeric'
    else:
        return 'categorical'
    
try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


def save_uploaded_file(user_id: str, django_file) -> dict:
    filename = django_file.name
    user_folder = os.path.join(settings.MEDIA_ROOT, f"ID_{user_id}_uploads", "temporary_uploads")
    os.makedirs(user_folder, exist_ok=True)

    save_path = os.path.join(user_folder, filename)
    with open(save_path, 'wb+') as dest:
        for chunk in django_file.chunks():
            dest.write(chunk)

    file_url = os.path.join(
        settings.MEDIA_URL, f"ID_{user_id}_uploads", "temporary_uploads", filename
    ).replace('\\', '/')

    return {
        "save_path": save_path,
        "filename": filename,
        "file_url": file_url,
        "user_folder": user_folder,
    }


def _get_active_sheet_name_xlsx(path: str) -> str | None:
    if load_workbook is None:
        return None
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        return wb.active.title if wb and wb.active else None
    except Exception:
        return None


def infer_columns_from_file(save_path: str, filename: str, active_sheet_name: str | None = None):
    ext = os.path.splitext(filename)[1].lower()
    cols: list[str] = []
    sheet_names: list[str] = []
    used_sheet: str | None = None

    if ext == '.csv':
        df = pd.read_csv(save_path, nrows=200)
        cols = list(df.columns)
        return cols, sheet_names, used_sheet

    try:
        xls = pd.ExcelFile(save_path)
        sheet_names = xls.sheet_names or []
    except Exception as e:
        try:
            df = pd.read_excel(save_path, nrows=200)
            cols = list(df.columns)
            return cols, sheet_names, used_sheet
        except Exception as inner:
            raise RuntimeError(f"Failed to read Excel file: {inner}") from e

    target_sheet = None

    if active_sheet_name and active_sheet_name in sheet_names:
        target_sheet = active_sheet_name

    if target_sheet is None and ext in ('.xlsx', '.xlsm', '.xltx', '.xltm'):
        active_name = _get_active_sheet_name_xlsx(save_path)
        if active_name and active_name in sheet_names:
            target_sheet = active_name

    if target_sheet is None and sheet_names:
        target_sheet = sheet_names[0]

    used_sheet = target_sheet

    df = pd.read_excel(save_path, sheet_name=target_sheet, nrows=200)
    cols = list(df.columns)

    return cols, sheet_names, used_sheet


@csrf_exempt
def upload_file(request):
    
    if not request.FILES.get('file'):
        return JsonResponse({'success': False, 'error': 'No file uploaded'})

    user_id = request.POST.get('userID')
    if not user_id:
        return JsonResponse({'success': False, 'error': 'User ID not provided'})

    try:
        save_info = save_uploaded_file(user_id, request.FILES['file'])
        return JsonResponse({
            'success': True,
            'user_id': user_id,
            **{
                'save_path': save_info['save_path'],
                'filename': save_info['filename'],
                'fileURL': save_info['file_url'],
                'user_folder': save_info['user_folder'],
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e), 'user_id': user_id})


@csrf_exempt
def get_columns(request):
    user_id = request.POST.get('userID')
    filename = request.POST.get('filename')
    active_sheet_name = request.POST.get('activeSheet')  
    file_url = request.POST.get('Fileurl')

    if not file_url:
        return JsonResponse({'success': False, 'error': 'File URL not provided'})

    user_folder = os.path.join(settings.MEDIA_ROOT, file_url.replace("/media/", ""))
    save_path = user_folder 

    if not user_id:
        return JsonResponse({'success': False, 'error': 'User ID not provided'})
    if not save_path or not filename:
        return JsonResponse({'success': False, 'error': 'savePath and filename are required', 'user_id': user_id})

    try:
        try:
            cols, sheet_names, used_sheet = infer_columns_from_file(
                save_path=save_path,
                filename=filename,
                active_sheet_name=active_sheet_name
            )
        except Exception as infer_err:
            try:
                df = pd.read_excel(save_path, nrows=200)
                cols = list(df.columns)
                sheet_names = []
                used_sheet = None
            except Exception:
                return JsonResponse({
                    'success': False,
                    'error': f'Failed to infer columns: {infer_err}',
                    'user_id': user_id
                })

        # Normalize column names by stripping quotes
        def normalize_column(col):
            col_str = str(col)
            # Remove surrounding quotes if present
            if (col_str.startswith('"') and col_str.endswith('"')) or \
               (col_str.startswith("'") and col_str.endswith("'")):
                col_str = col_str[1:-1].strip()
            return col_str
        
        normalized_cols = [canonicalize_column(col) for col in cols]


        return JsonResponse({
            'success': True,
            'user_id': user_id,
            'columns': normalized_cols,  # Send normalized columns to frontend
            'original_columns': cols,    # Keep original for debugging
            'sheet_names': sheet_names,
            'active_sheet_used': used_sheet,
            'filename': filename,
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e), 'user_id': user_id})

import re

def canonicalize_column(col: str) -> str:
    """
    Create a stable, comparable column key:
    - remove newlines
    - collapse multiple spaces
    - strip
    """
    if col is None:
        return col
    col = str(col)
    col = col.replace('\n', ' ').replace('\r', ' ')
    col = re.sub(r'\s+', ' ', col)  # collapse whitespace
    return col.strip()


def normalize_column_name(col: str, df: pd.DataFrame) -> str:
    """
    Normalize column name by removing quotes and matching with actual DataFrame columns.
    
    Args:
        col: Column name as received from frontend (may have quotes)
        df: DataFrame to match against
        
    Returns:
        Normalized column name that exists in df.columns
    """
    if not col:
        return col
    
    # Remove surrounding quotes if present
    col_clean = col.strip()
    if (col_clean.startswith('"') and col_clean.endswith('"')) or \
       (col_clean.startswith("'") and col_clean.endswith("'")):
        col_clean = col_clean[1:-1].strip()
    
    # Try to match with existing columns (exact match first)
    if col_clean in df.columns:
        return col_clean
    
    # Try case-insensitive match
    for actual_col in df.columns:
        if str(actual_col).strip().lower() == col_clean.lower():
            return actual_col
    
    # Try matching without extra whitespace
    for actual_col in df.columns:
        if str(actual_col).strip() == col_clean:
            return actual_col
    
    # If still not found, return the cleaned version
    return col_clean


def validate_and_get_column(col: str, df: pd.DataFrame, test_name: str = "") -> tuple[str, bool, str]:
    """
    Validate column name and return normalized version.
    
    Returns:
        tuple: (normalized_column, is_valid, error_message)
    """
    if not col:
        return col, False, f"No column specified for {test_name}"
    
    normalized_col = normalize_column_name(col, df)
    
    if normalized_col not in df.columns:
        # Try to find the exact match with quotes
        if f'"{normalized_col}"' in df.columns:
            return f'"{normalized_col}"', True, ""
        elif f"'{normalized_col}'" in df.columns:
            return f"'{normalized_col}'", True, ""
        else:
            return normalized_col, False, f"Column '{col}' not found in dataset. Available columns: {list(df.columns)}"
    
    return normalized_col, True, ""


# def get_columns(request):
#     if request.method == 'POST' and request.FILES.get('file'):
#         user_id = request.POST.get('userID')
#         print(f"Received user_id: {user_id}")

#         if not user_id:
#             return JsonResponse({'success': False, 'error': 'User ID not provided'})

#         try:
#             excel_file = request.FILES['file']
#             filename = request.FILES['file'].name
#             print(f"Received file: {filename}")

#             # Create the user's uploads folder: media/ID_<user_id>_uploads/temporary_uploads
#             user_folder = os.path.join(settings.MEDIA_ROOT, f"ID_{user_id}_uploads", "temporary_uploads")
#             os.makedirs(user_folder, exist_ok=True)

#             # Save the ORIGINAL upload unchanged
#             save_path = os.path.join(user_folder, filename)
#             with open(save_path, 'wb+') as dest:
#                 for chunk in excel_file.chunks():
#                     dest.write(chunk)

#             # Infer columns (supports CSV and Excel). We'll reassign df AFTER saving.
#             cols = []
#             sheet_names = []
#             try:
#                 ext = os.path.splitext(filename)[1].lower()
#                 if ext == '.csv':
#                     df = pd.read_csv(save_path, nrows=200)  # peek only
#                     cols = list(df.columns)
#                 else:
#                     xls = pd.ExcelFile(save_path)
#                     sheet_names = xls.sheet_names
#                     df = pd.read_excel(save_path, sheet_name=sheet_names[0], nrows=200)
#                     cols = list(df.columns)
#             except Exception as infer_err:
#                 # Fallback: try generic Excel read (keeps your original behavior path)
#                 df = pd.read_excel(save_path)
#                 cols = list(df.columns)
#                 print(f"Inference fallback due to: {infer_err}")

#             # Only process tests if selected (keeping your placeholders)
#             matched_results = {}
#             input_info = {}

#             return JsonResponse({
#                 'success': True,
#                 'user_id': user_id,
#                 'columns': cols,
#                 'input_info': input_info,
#                 # Clean, browser-usable URL (forward slashes)
#                 'fileURL': os.path.join(
#                     settings.MEDIA_URL, f"ID_{user_id}_uploads", "temporary_uploads", filename
#                 ).replace('\\', '/'),
#                 # Helpful for multi-tab preview in the client (Excel only; CSV -> [])
#                 'sheet_names': sheet_names,
#                 'filename': filename,
#             })

#         except Exception as e:
#             return JsonResponse({
#                 'success': False,
#                 'error': str(e),
#                 'user_id': user_id
#             })

#     return JsonResponse({'success': False, 'error': 'Invalid request method or no file uploaded'})


from django.http import HttpResponse, JsonResponse
import os
import json
import pandas as pd
from django.http import JsonResponse, HttpResponse
from django.conf import settings


def analyze_data_api(request):
    if request.method == 'POST':
        try:
            user_id = request.POST.get('userID')
            filename = request.POST.get('file_name')
            sheet_name= request.POST.get('sheet_name')
            file_url= request.POST.get('Fileurl') 
            
            if not user_id:
                return JsonResponse({'success': False, 'error': 'User ID not provided'})

            # Load the DataFrame
            file_path = os.path.join(settings.MEDIA_ROOT, file_url.replace("/media/", "")) 
            if not os.path.exists(file_path):
                return JsonResponse({'success': False, 'error': 'No uploaded file found for this user'}, status=404)

            lower = filename.lower()
            if lower.endswith(('.xls', '.xlsx', '.xlsm', '.xlsb', '.ods')):
                try:
                    xls = pd.ExcelFile(file_path)
                    sheet_names = xls.sheet_names
                except Exception as e:
                    return JsonResponse({'success': False, 'error': f'Failed to read workbook: {e}'}, status=400)

                if sheet_name:
                    if sheet_name not in sheet_names:
                        return JsonResponse({
                            'success': False,
                            'error': f"Requested sheet '{sheet_name}' not found. Available: {sheet_names}"
                        }, status=400)
                    df = pd.read_excel(xls, sheet_name=sheet_name)
                    active_sheet = sheet_name
                else:
                    df = pd.read_excel(xls, sheet_name=sheet_names[0] if sheet_names else 0)
                    active_sheet = sheet_names[0] if sheet_names else 'Sheet1'
            else:
                df = pd.read_csv(file_path)
                active_sheet = 'Sheet1'
            
            original_columns = df.columns.tolist()
            df.columns = [canonicalize_column(c) for c in df.columns]

            # Optional debug
            print("Original columns:", original_columns)
            print("Normalized columns:", df.columns.tolist())

            # Get test parameters
            if request.content_type == 'application/json':
                body = json.loads(request.body)
                test_type = body.get('test_type', '')
                col1 = body.get('column1', '')
                col2 = body.get('column2', '')
                col3 = body.get('column3', '')
                col_group = body.get('primary_col')
                col_covariate = body.get('secondary_col')
                col_outcome = body.get('dependent_col')
                language = body.get('language', 'en')
                file_format = body.get('format', 'png')
                image_width = body.get('image_width')
                image_height = body.get('image_height') 
            else:
                test_type = request.POST.get('test_type', '')
                col1 = request.POST.get('column1', '')
                col2 = request.POST.get('column2', '')
                col3 = request.POST.get('column3', '')
                col_group = request.POST.get('primary_col')
                col_covariate = request.POST.get('secondary_col')
                col_outcome = request.POST.get('dependent_col')
                language = request.POST.get('language', 'en')
                file_format = request.POST.get('format', 'png')
                image_width = request.POST.get('image_width')
                image_height = request.POST.get('image_height')

            print(f"Received test_type: {test_type}, col1: {col1}, col2: {col2}, col3: {col3}")
            
            # Helper function to normalize column names

            def normalize_column_name(col: str, df: pd.DataFrame) -> str:
                """
                Normalize column name and match it to actual DataFrame columns.
                Handles:
                - quotes
                - newlines
                - extra whitespace
                - case differences
                """
                if not col:
                    return col

                # Step 1: strip quotes first (frontend artifacts)
                col_clean = col.strip()
                if (col_clean.startswith('"') and col_clean.endswith('"')) or \
                (col_clean.startswith("'") and col_clean.endswith("'")):
                    col_clean = col_clean[1:-1]

                # Step 2: canonicalize (NEWLINES + SPACES)
                col_clean = canonicalize_column(col_clean)

                # Step 3: exact match
                if col_clean in df.columns:
                    return col_clean

                # Step 4: case-insensitive canonical match
                for actual_col in df.columns:
                    if canonicalize_column(actual_col).lower() == col_clean.lower():
                        return actual_col

                return col_clean


            # Helper function to validate column
            def validate_column(col: str, df: pd.DataFrame, test_name: str = "") -> tuple[bool, str, str]:
                """
                Validate column name and return normalized version.
                Returns: (is_valid, normalized_column, error_message)
                """
                if not col:
                    return False, col, f"No column specified for {test_name}"
                
                normalized_col = normalize_column_name(col, df)
                
                if normalized_col not in df.columns:
                    # Check if original exists
                    if col in df.columns:
                        return True, col, ""
                    else:
                        return False, normalized_col, f"Column '{col}' not found in dataset. Available columns: {list(df.columns)}"
                
                return True, normalized_col, ""
            
            # Normalize all column names
            col1_valid, col1_norm, col1_error = validate_column(col1, df, test_type)
            if col1 and not col1_valid:
                return JsonResponse({'success': False, 'error': col1_error})
            
            col2_valid, col2_norm, col2_error = validate_column(col2, df, test_type)
            if col2 and not col2_valid:
                return JsonResponse({'success': False, 'error': col2_error})
            
            col3_valid, col3_norm, col3_error = validate_column(col3, df, test_type)
            if col3 and not col3_valid:
                return JsonResponse({'success': False, 'error': col3_error})
            
            col_group_valid, col_group_norm, col_group_error = validate_column(col_group, df, test_type)
            if col_group and not col_group_valid:
                return JsonResponse({'success': False, 'error': col_group_error})
            
            col_covariate_valid, col_covariate_norm, col_covariate_error = validate_column(col_covariate, df, test_type)
            if col_covariate and not col_covariate_valid:
                return JsonResponse({'success': False, 'error': col_covariate_error})
            
            col_outcome_valid, col_outcome_norm, col_outcome_error = validate_column(col_outcome, df, test_type)
            if col_outcome and not col_outcome_valid:
                return JsonResponse({'success': False, 'error': col_outcome_error})
            
            # Process categorical data BEFORE passing to test functions
            print("Converting categorical columns to numerical values...")
            ordinal_mappings = {}
            categorical_cols = df.select_dtypes(include=['object']).columns.tolist()

            # For Mann–Whitney, DO NOT ENCODE col1
            if test_type == 'mannwhitney' and col1_norm in categorical_cols:
                categorical_cols.remove(col1_norm)

            if categorical_cols:
                encoder = OrdinalEncoder()
                df[categorical_cols] = encoder.fit_transform(df[categorical_cols])

                print(f"Converted categorical columns: {list(categorical_cols)}")

                for i, col in enumerate(categorical_cols):
                    categories = encoder.categories_[i]
                    ordinal_mappings[col] = {idx: cat for idx, cat in enumerate(categories)}

                print(ordinal_mappings)
            
            print(f"DataFrame head after conversion:\n{df[[col1_norm, col2_norm]].head() if col2_norm in df.columns else df[[col1_norm]].head()}")
            
            # Route to appropriate test function with NORMALIZED column names
            if test_type == 'kruskal':
                return process_kruskal_test(request, df, col1_norm, col2_norm, user_id, ordinal_mappings)
            
            elif test_type == 'mannwhitney':
                return process_mannwhitney_test(request, df, col1_norm, col2_norm, user_id, ordinal_mappings)
 
            elif test_type == 'pearson':
                selected_columns = []
                for key in request.POST:
                    if key.startswith("column"):
                        value = request.POST[key]
                        if value:
                            valid, norm_value, error = validate_column(value, df, "Pearson")
                            if valid and norm_value in df.columns:
                                selected_columns.append(norm_value)
                print("Selected columns for Pearson:", selected_columns)
                return process_pearson_test(request, df, selected_columns, user_id)

            elif test_type == 'spearman':
                selected_columns = [] 
                for key in request.POST:
                    if key.startswith("column"):
                        value = request.POST[key]
                        if value:
                            valid, norm_value, error = validate_column(value, df, "Spearman")
                            if valid and norm_value in df.columns:
                                selected_columns.append(norm_value)
                print("Selected columns for Spearman:", selected_columns)
                return process_spearman_test(request, df, selected_columns, user_id)
          
            elif test_type == 'wilcoxon':
                return process_wilcoxon_test(request, df, col1_norm, col2_norm, user_id)
      
            elif test_type == 'shapiro':
                return process_shapiro_test(request, df, col1_norm, user_id)
            
            elif test_type == 'linear_regression':
                return process_linear_regression_test(request, df, col1_norm, col2_norm, user_id)
            
            elif test_type == 'anova':
                return process_anova_test(request, df, col1_norm, col2_norm, user_id)
            
            elif test_type == 'ancova':
                return process_ancova_test(request, df, col_group_norm, col_covariate_norm, col_outcome_norm, user_id)

            elif test_type == 'kolmogorov':
                column = request.POST.get('column')
                valid, norm_column, error = validate_column(column, df, "Kolmogorov")
                if not valid:
                    return JsonResponse({'success': False, 'error': error})
                return process_ks_test(request, df, norm_column, user_id)

            elif test_type == 'anderson':
                column = request.POST.get('column')
                valid, norm_column, error = validate_column(column, df, "Anderson")
                if not valid:
                    return JsonResponse({'success': False, 'error': error})
                return process_anderson_darling_test(request, df, norm_column, user_id)            

            elif test_type == 'f_test':                
                return process_f_test(request, df, col1_norm, col2_norm, user_id)
            
            elif test_type == 'z_test':
                return process_z_test(request, df, col1_norm, col2_norm, user_id)  
            
            elif test_type == 't_test':
                return process_t_test(request, df, col1_norm, col2_norm, user_id)
            
            elif test_type == 'fzt_visualization':
                return process_fzt_visualization(request, df, col1_norm, col2_norm, user_id)

            elif test_type == 'cross_tabulation':
                selected_columns = []
                for key in request.POST:
                    if key.startswith("column"):
                        value = request.POST[key]
                        if value:
                            valid, norm_value, error = validate_column(value, df, "Cross-Tabulation")
                            if valid and norm_value in df.columns:
                                selected_columns.append(norm_value)
                print("Selected columns for Cross-Tabulation:", selected_columns)
                return process_cross_tabulation(request, df, selected_columns, user_id)

            elif test_type == 'eda_distribution':
                return process_eda_distribution(request, df, col1_norm, user_id)

            elif test_type == 'eda_swarm':
                return process_eda_swarm_plot(request, df, col1_norm, col2_norm, user_id)
            
            elif test_type == 'eda_pie':
                return process_pie_chart(request, df, col1_norm, user_id, ordinal_mappings)

            elif test_type == 'eda_basics':
                return process_eda_basics(request, df, user_id)

            elif test_type == 'similarity':
                return process_similarity(request, df, user_id)

            elif test_type == 'chi_square':
                selected_columns = []
                for key in request.POST:
                    if key.startswith("column"):
                        value = request.POST[key]
                        if value:
                            valid, norm_value, error = validate_column(value, df, "Chi-Square")
                            if valid and norm_value in df.columns:
                                selected_columns.append(norm_value)
                print("Selected columns for Chi-Square:", selected_columns)
                return process_chi_square(request, df, selected_columns, user_id)

            elif test_type == 'cramers':
                selected_columns = []
                for key in request.POST:
                    if key.startswith("column"):
                        value = request.POST[key]
                        if value:
                            valid, norm_value, error = validate_column(value, df, "Cramer's V")
                            if valid and norm_value in df.columns:
                                selected_columns.append(norm_value)
                print("Selected columns for Cramer's V:", selected_columns)
                return process_cramers_test(request, df, selected_columns, user_id)

            elif test_type == 'network_graph':
                selected_columns = []
                for key in request.POST:
                    if key.startswith("column") and request.POST[key] in df.columns:
                        value = request.POST[key]
                        if value:
                            valid, norm_value, error = validate_column(value, df, "Network Graph")
                            if valid and norm_value in df.columns:
                                selected_columns.append(norm_value)
                return process_network_graph(request, df, selected_columns, user_id)

            elif test_type == 'bar_chart':
                orientation = request.POST.get('orientation', 'vertical')
                return process_bar_chart_test(request, df, col1_norm, user_id, orientation, ordinal_mappings)
            
            else:
                return JsonResponse({
                    'success': False,
                    'error': f'Unknown test type: {test_type}'
                })
                
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"Error in analyze_data_api: {str(e)}\n{error_details}")
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    else:
        form = AnalysisForm(columns=[])
    return render(request, 'analysis/upload.html', {'form': form})

# views.py (or wherever your view lives)

import os
import io
import json
import math
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import seaborn as sns
import pandas as pd
from PIL import Image, ImageDraw, ImageFont
from scipy import stats

from django.conf import settings
from django.http import JsonResponse

# Optional: googletrans for auto-translation of labels to Bengali
try:
    from googletrans import Translator  # pip install googletrans==4.0.0-rc1
except Exception:  # pragma: no cover
    Translator = None


@csrf_exempt
def get_groups(request):
    """Get unique groups from a column for Mann-Whitney test."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'})
    
    user_id = request.POST.get('userID')
    filename = request.POST.get('filename')
    file_url = request.POST.get('Fileurl')
    column_name = request.POST.get('column')
    
    if not all([user_id, filename, file_url, column_name]):
        return JsonResponse({'success': False, 'error': 'Missing required parameters'})
    
    try:
        # Load the file
        file_path = os.path.join(settings.MEDIA_ROOT, file_url.replace("/media/", ""))
        
        # Read the file
        if filename.lower().endswith('.csv'):
            df = pd.read_csv(file_path)
        else:
            df = pd.read_excel(file_path)
        
        # Check if column exists
        if column_name not in df.columns:
            return JsonResponse({'success': False, 'error': f'Column "{column_name}" not found'})
        
        # Get unique groups (remove NaN)
        unique_groups = df[column_name].dropna().astype(str).unique()
        groups_list = [str(g) for g in unique_groups[:20]]  # Limit to first 20
        
        return JsonResponse({
            'success': True,
            'groups': groups_list,
            'total_groups': len(unique_groups)
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@csrf_exempt
def get_column_types(request):
    """
    Analyze columns in a file and categorize them as numeric or categorical.
    Returns separate lists for numeric and categorical columns.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'})
    
    user_id = request.POST.get('userID')
    filename = request.POST.get('filename')
    file_url = request.POST.get('Fileurl')
    
    if not all([user_id, filename, file_url]):
        return JsonResponse({'success': False, 'error': 'Missing required parameters'})
    
    try:
        # Load the file
        file_path = os.path.join(settings.MEDIA_ROOT, file_url.replace("/media/", ""))
        
        # Read the file (first 500 rows only)
        if filename.lower().endswith('.csv'):
            df = pd.read_csv(file_path, nrows=500)
        else:
            df = pd.read_excel(file_path, nrows=500)
        
        # Initialize lists
        numeric_columns = []
        categorical_columns = []

        # --- NEW, FIXED, RELIABLE TYPE DETECTION ---
        for col in df.columns:
            series = df[col]

            # Convert to numeric where possible
            numeric_series = pd.to_numeric(series, errors='coerce')
            numeric_ratio = numeric_series.notna().mean()  # % numeric values

            # Treat as numeric if at least 90% of values are numeric
            if numeric_ratio >= 0.9:
                numeric_columns.append(col)
            else:
                categorical_columns.append(col)
        # --------------------------------------------

        return JsonResponse({
            'success': True,
            'numeric_columns': numeric_columns,
            'categorical_columns': categorical_columns,
            'all_columns': list(df.columns),
            'total_numeric': len(numeric_columns),
            'total_categorical': len(categorical_columns)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error analyzing column types: {str(e)}'
        })


def process_kruskal_test(request, df: pd.DataFrame, col1: str, col2: str, user_id: str, ordinal_mappings):
    """
    Performs Kruskal-Wallis H-test and returns analysis results without generating plots.
    Frontend will handle visualization using the returned data.
    
    Args:
        request: Django request object
        df: DataFrame containing the data
        col1: Column name for grouping variable (categorical)
        col2: Column name for value variable (numeric)
        user_id: User identifier
        
    Returns:
        JsonResponse with test results and data for plotting
    """
    print(f"[Kruskal] cols: {col1}(group) | {col2}(value)")
    
    # Validate column names
    if col1 not in df.columns or col2 not in df.columns:
        return JsonResponse({
            'success': False, 
            'error': 'Invalid column names.'
        })

    # Extract language preference
    try:
        language = request.POST.get('language', 'en').lower()
    except Exception:
        language = 'en'
    
    if language not in ('en', 'bn'):
        language = 'en'

    # Prepare working dataset
    work = df[[col1, col2]].copy()
    work = work.dropna(subset=[col1, col2])
    print(f"[Kruskal] after dropna: {len(work)} rows")

    # Check if there are enough rows
    if len(work) == 0:
        return JsonResponse({
            'success': False, 
            'error': 'No valid data after removing missing values.'
        })

    # Check if all values in col2 are identical
    unique_values = work[col2].nunique()
    if unique_values == 1:
        # All values are identical - Kruskal-Wallis cannot be performed
        warning_message_en = "All values in the numeric column are identical. The Kruskal-Wallis test requires variation in the data to compute ranks. When all values are the same, all ranks will be tied, making the test statistically invalid."
        
        warning_message_bn = "সংখ্যাসূচক কলামের সকল মান অভিন্ন। ক্রুসকাল-ওয়ালিস পরীক্ষার জন্য ডেটার মধ্যে বৈচিত্র্য প্রয়োজন যাতে র্যাঙ্ক নির্ধারণ করা যায়। যখন সকল মান একই থাকে, সকল র্যাঙ্ক সমান হয়ে যায়, ফলে পরীক্ষাটি পরিসংখ্যানগতভাবে অবৈধ হয়ে পড়ে।"
        
        warning_message = warning_message_bn if language == 'bn' else warning_message_en
        
        return JsonResponse({
            'success': True,
            'identical_values': True,
            'warning_message': warning_message,
            'language': language,
            'identical_value': float(work[col2].iloc[0]),  # The single value that repeats
            'n_observations': int(len(work)),
            'column_names': {
                'group': str(col1),
                'value': str(col2)
            },
            'test': 'Kruskal-Wallis H-test' if language == 'en' else 'ক্রুসকাল-ওয়ালিস এইচ-টেস্ট'
        })

    # Ensure grouping column is categorical
    if not pd.api.types.is_categorical_dtype(work[col1]):
        work[col1] = work[col1].astype('category')

    categories = list(work[col1].cat.categories)
    
    if len(categories) < 2:
        return JsonResponse({
            'success': False, 
            'error': 'Need at least 2 groups in the factor column for Kruskal–Wallis.'
        })

    # Ensure value column is numeric
    if not pd.api.types.is_numeric_dtype(work[col2]):
        work[col2] = pd.to_numeric(work[col2], errors='coerce')
        work = work.dropna(subset=[col2])
        
        if not pd.api.types.is_numeric_dtype(work[col2]):
            return JsonResponse({
                'success': False, 
                'error': f'"{col2}" must be numeric for Kruskal–Wallis.'
            })

    # Prepare groups for statistical test
    groups = [work.loc[work[col1] == g, col2].values for g in categories]
    
    if any(len(g) == 0 for g in groups):
        return JsonResponse({
            'success': False, 
            'error': 'Each group must contain at least one observation.'
        })

    # Perform Kruskal-Wallis H-test
    try:
        stat, p_value = stats.kruskal(*groups)
        print(f"[Kruskal] result: H={stat:.6f}, p={p_value:.6g}")
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'Error in Kruskal–Wallis test: {e}'
        })

    # Prepare data for frontend plotting
    plot_data = []
    
    # Get original categories (encoded)
    categories = list(work[col1].cat.categories)

    # Map encoded categories → real labels
    mapped_categories = []
    for cat in categories:
        try:
            mapped_label = ordinal_mappings.get(col1, {}).get(int(cat), str(cat))
        except:
            mapped_label = ordinal_mappings.get(col1, {}).get(str(cat), str(cat))
        mapped_categories.append(mapped_label)

    plot_data = []
    for orig_cat, mapped_label in zip(categories, mapped_categories):
        category_data = work.loc[work[col1] == orig_cat, col2].values

        plot_data.append({
            'category': str(mapped_label),   # fixed
            'values': [float(x) for x in category_data],
            'count': int(len(category_data)),
            'mean': float(np.mean(category_data)),
            'median': float(np.median(category_data)),
            'std': float(np.std(category_data)),
            'min': float(np.min(category_data)),
            'max': float(np.max(category_data)),
            'q25': float(np.percentile(category_data, 25)),
            'q75': float(np.percentile(category_data, 75))
        }) 

    # Prepare response with localized labels
    test_name = 'Kruskal-Wallis H-test' if language == 'en' else 'ক্রুসকাল-ওয়ালিস এইচ-টেস্ট'
    
    response_data = {
        'success': True,
        'identical_values': False,  # This is the normal case
        'test': test_name,
        'language': language,
        'statistic': float(stat),
        'p_value': float(p_value),
        'degrees_of_freedom': int(len(categories) - 1),
        'n_groups': int(len(categories)),
        'total_observations': int(len(work)),
        'column_names': {
            'group': str(col1),
            'value': str(col2)
        },
        'plot_data': plot_data,
        'metadata': {
            'categories': [str(c) for c in categories],
            'significant': bool(p_value < 0.05),
            'alpha': 0.05
        }
    }
    return JsonResponse(response_data)


def process_mannwhitney_test(request, df: pd.DataFrame, col1: str, col2: str, user_id: str, ordinal_mappings):
    """
    Performs Mann-Whitney U test with proper validation.
    Requirements:
    - col1: Categorical column with exactly two unique groups
    - col2: Numerical column
    """
    from scipy.stats import mannwhitneyu, rankdata
    import numpy as np
    import json
    
    try:
        # Get language
        lang = request.POST.get('language', 'en')
        
        print(f"[DEBUG MANNWHITNEY] Starting Mann-Whitney test")
        print(f"[DEBUG MANNWHITNEY] Column 1 (categorical): {col1}")
        print(f"[DEBUG MANNWHITNEY] Column 2 (numerical): {col2}")
        
        # Check if columns exist
        if col1 not in df.columns or col2 not in df.columns:
            error_msg = f'Selected columns not found. Available: {list(df.columns)}'
            print(f"[DEBUG MANNWHITNEY] Error: {error_msg}")
            return JsonResponse({
                'success': False, 
                'error': error_msg
            })
        
        # Get selected groups from request
        selected_groups = []
        selected_groups_json = request.POST.get('selected_groups', '')
        
        print(f"[DEBUG MANNWHITNEY] Raw selected_groups JSON: {selected_groups_json}")
        
        if selected_groups_json:
            try:
                selected_groups = json.loads(selected_groups_json)
                print(f"[DEBUG MANNWHITNEY] Parsed selected groups: {selected_groups}")
            except json.JSONDecodeError as e:
                print(f"[DEBUG MANNWHITNEY] JSON decode error: {e}")
                selected_groups = []
        
        # If no groups selected, use all groups in the column
        if not selected_groups:
            unique_groups = sorted(df[col1].dropna().astype(str).unique())
            print(f"[DEBUG MANNWHITNEY] No groups selected, using all groups: {unique_groups}")
        else:
            # Convert selected groups to strings for comparison
            selected_groups = [str(g) for g in selected_groups]
            df[col1] = df[col1].astype(str)
            df = df[df[col1].isin(selected_groups)]
            unique_groups = selected_groups
            print(f"[DEBUG MANNWHITNEY] Filtered to selected groups: {unique_groups}")
        
        print(f"[DEBUG MANNWHITNEY] DataFrame shape before cleaning: {df.shape}")
        
        # Check if col2 is numeric, try to convert if not
        if not pd.api.types.is_numeric_dtype(df[col2]):
            print(f"[DEBUG MANNWHITNEY] Column 2 is not numeric, attempting conversion...")
            try:
                df[col2] = pd.to_numeric(df[col2], errors='coerce')
                print(f"[DEBUG MANNWHITNEY] Column 2 converted to numeric")
            except Exception as e:
                error_msg = f'Column "{col2}" must be numerical for Mann-Whitney test. Conversion error: {e}'
                print(f"[DEBUG MANNWHITNEY] Error: {error_msg}")
                return JsonResponse({
                    'success': False,
                    'error': error_msg
                })
        
        # Remove NaN values
        df = df.dropna(subset=[col1, col2])
        print(f"[DEBUG MANNWHITNEY] DataFrame shape after cleaning: {df.shape}")
        
        # Check if we have exactly 2 groups
        if len(unique_groups) != 2:
            error_msg = f'Mann-Whitney test requires exactly 2 groups. Found: {len(unique_groups)} groups ({unique_groups})'
            print(f"[DEBUG MANNWHITNEY] Error: {error_msg}")
            return JsonResponse({
                'success': False,
                'error': error_msg
            })
        
        # Check group counts
        group_counts = df[col1].value_counts()
        print(f"[DEBUG MANNWHITNEY] Group counts: {group_counts.to_dict()}")
        
        if any(group_counts < 3):
            explanation = (
                "Mann-Whitney U Test requires:\n"
                "• A categorical column with exactly 2 groups\n"
                "• A numerical column\n"
                "• Each group must have at least 3 observations\n\n"
                f"Current group counts: {group_counts.to_dict()}\n"
                "One or both groups have fewer than 3 observations, so the test cannot be performed."
            )

            print("[DEBUG MANNWHITNEY] Not enough observations. Returning explanation instead of error.")

            return JsonResponse({
                'success': True,        # IMPORTANT: keep true
                'is_explanation': True, # FRONTEND USES THIS FLAG
                'message': explanation,
                'groups': list(group_counts.index),
                'group_sizes': {g: int(group_counts[g]) for g in group_counts.index}
            })
        
        # Get group data
        group1_data = df[df[col1] == unique_groups[0]][col2].values
        group2_data = df[df[col1] == unique_groups[1]][col2].values
        
        print(f"[DEBUG MANNWHITNEY] Group 1 ({unique_groups[0]}) data sample: {group1_data[:5] if len(group1_data) > 0 else 'No data'}")
        print(f"[DEBUG MANNWHITNEY] Group 2 ({unique_groups[1]}) data sample: {group2_data[:5] if len(group2_data) > 0 else 'No data'}")
        
        # ============================================================
        # NEW: CHECK IF ALL VALUES ARE IDENTICAL
        # ============================================================
        all_values = np.concatenate([group1_data, group2_data])
        unique_values = np.unique(all_values)
        
        if len(unique_values) == 1:
            # All values are identical - Mann-Whitney cannot be performed
            print(f"[DEBUG MANNWHITNEY] All values are identical: {unique_values[0]}")
            
            warning_message_en = "All values in the numeric column are identical. The Mann-Whitney test requires variation in the data to compute ranks. When all values are the same, all ranks will be tied, making the test statistically invalid."
            
            warning_message_bn = "সংখ্যাসূচক কলামের সকল মান অভিন্ন। ম্যান-হুইটনি পরীক্ষার জন্য ডেটার মধ্যে বৈচিত্র্য প্রয়োজন যাতে র্যাঙ্ক নির্ধারণ করা যায়। যখন সকল মান একই থাকে, সকল র্যাঙ্ক সমান হয়ে যায়, ফলে পরীক্ষাটি পরিসংখ্যানগতভাবে অবৈধ হয়ে পড়ে।"
            
            warning_message = warning_message_bn if lang == 'bn' else warning_message_en
            
            return JsonResponse({
                'success': True,
                'identical_values': True,
                'warning_message': warning_message,
                'language': lang,
                'identical_value': float(unique_values[0]),  # The single value that repeats
                'n_observations': int(len(df)),
                'column_names': {
                    'group': str(col1),
                    'value': str(col2)
                },
                'groups': unique_groups,
                'group_sizes': {
                    unique_groups[0]: int(len(group1_data)),
                    unique_groups[1]: int(len(group2_data))
                },
                'test': 'Mann-Whitney U Test' if lang == 'en' else 'ম্যান-হুইটনি ইউ টেস্ট'
            })
        # ============================================================
        # END OF NEW CHECK
        # ============================================================
        
        # Perform Mann-Whitney test
        try:
            print(f"[DEBUG MANNWHITNEY] Performing Mann-Whitney U test...")
            u_stat, p_value = mannwhitneyu(
                group1_data,
                group2_data,
                alternative="two-sided"
            )
            print(f"[DEBUG MANNWHITNEY] Test results - U: {u_stat}, p-value: {p_value}")
        except Exception as e:
            error_msg = f'Error performing Mann-Whitney test: {str(e)}'
            print(f"[DEBUG MANNWHITNEY] Error: {error_msg}")
            return JsonResponse({
                'success': False,
                'error': error_msg
            })
        
        # Calculate rank statistics
        all_data = np.concatenate([group1_data, group2_data])
        ranks = rankdata(all_data)
        
        # Split ranks back to groups
        group1_ranks = ranks[:len(group1_data)]
        group2_ranks = ranks[len(group1_data):]
        
        mean_rank1 = np.mean(group1_ranks)
        mean_rank2 = np.mean(group2_ranks)
        
        print(f"[DEBUG MANNWHITNEY] Mean ranks - Group 1: {mean_rank1}, Group 2: {mean_rank2}")
        
        # Prepare plot data
        plot_data = []
        for i, group in enumerate(unique_groups):
            group_data = df[df[col1] == group][col2].values
            group_ranks = group1_ranks if i == 0 else group2_ranks
            
            plot_data.append({
                'category': str(group),
                'values': [float(x) for x in group_data],
                'count': int(len(group_data)),
                'mean': float(np.mean(group_data)) if len(group_data) > 0 else np.nan,
                'median': float(np.median(group_data)) if len(group_data) > 0 else np.nan,
                'std': float(np.std(group_data)) if len(group_data) > 1 else np.nan,
                'min': float(np.min(group_data)) if len(group_data) > 0 else np.nan,
                'max': float(np.max(group_data)) if len(group_data) > 0 else np.nan,
                'q25': float(np.percentile(group_data, 25)) if len(group_data) > 0 else np.nan,
                'q75': float(np.percentile(group_data, 75)) if len(group_data) > 0 else np.nan,
                'mean_rank': float(mean_rank1 if i == 0 else mean_rank2)
            })
        
        # Prepare response
        result = {
            'success': True,
            'identical_values': False,  # Normal case
            'test': 'Mann-Whitney U Test' if lang == 'en' else 'ম্যান-হুইটনি ইউ টেস্ট',
            'statistic': float(u_stat),
            'p_value': float(p_value),
            'conclusion': 'Significant difference' if p_value < 0.05 else 'No significant difference',
            'n_groups': len(unique_groups),
            'total_observations': len(df),
            'column_names': {
                'group': str(col1),
                'value': str(col2)
            },
            'groups': unique_groups,
            'group_sizes': {
                unique_groups[0]: int(len(group1_data)),
                unique_groups[1]: int(len(group2_data))
            },
            'mean_ranks': {
                unique_groups[0]: float(mean_rank1),
                unique_groups[1]: float(mean_rank2)
            },
            'plot_data': plot_data,
            'metadata': {
                'categories': unique_groups,
                'significant': bool(p_value < 0.05),
                'alpha': 0.05
            }
        }
        
        print(f"[DEBUG MANNWHITNEY] Returning successful result")
        return JsonResponse(result)
        
    except Exception as e:
        import traceback
        error_msg = f'Unexpected error: {str(e)}\n{traceback.format_exc()}'
        print(f"[DEBUG MANNWHITNEY] Critical error: {error_msg}")
        return JsonResponse({'success': False, 'error': str(e)})


def process_wilcoxon_test(request, df: pd.DataFrame, col1: str, col2: str, user_id: str):
    """
    Performs Wilcoxon Signed-Rank test and returns analysis results without generating plots.
    Frontend will handle visualization using the returned data.
    """
    print(f"[Wilcoxon] cols: {col1} | {col2}")
    
    # Validate column names
    if col1 not in df.columns or col2 not in df.columns:
        return JsonResponse({
            'success': False, 
            'error': 'Invalid column names.'
        })

    # Extract language preference
    try:
        language = request.POST.get('language', 'en').lower()
    except Exception:
        language = 'en'
    
    if language not in ('en', 'bn'):
        language = 'en'

    # Prepare working dataset
    work = df[[col1, col2]].copy()
    work = work.dropna(subset=[col1, col2])
    print(f"[Wilcoxon] after dropna: {len(work)} rows")

    # Ensure columns are numeric
    for col in [col1, col2]:
        if not pd.api.types.is_numeric_dtype(work[col]):
            work[col] = pd.to_numeric(work[col], errors='coerce')
    
    work = work.dropna(subset=[col1, col2])
    
    if len(work) < 2:
        return JsonResponse({
            'success': False, 
            'error': 'Need at least 2 paired observations for Wilcoxon test.'
        })

    sample1 = work[col1].values
    sample2 = work[col2].values
    differences = sample1 - sample2

    # Perform Wilcoxon Signed-Rank test
    try:
        stat, p_value = stats.wilcoxon(sample1, sample2)
        print(f"[Wilcoxon] result: statistic={stat:.6f}, p={p_value:.6g}")
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'Error in Wilcoxon test: {e}'
        })

    # Calculate regression line data for scatter plot
    try:
        # Linear regression for scatter plot
        slope, intercept, r_value, p_value_reg, std_err = stats.linregress(sample1, sample2)
        regression_data = {
            'slope': float(slope),
            'intercept': float(intercept),
            'r_squared': float(r_value**2),
            'p_value': float(p_value_reg)
        }
    except Exception as e:
        print(f"[Wilcoxon] Regression calculation warning: {e}")
        regression_data = {
            'slope': 0,
            'intercept': 0,
            'r_squared': 0,
            'p_value': 1.0
        }

    # Calculate critical values for reference line display
    try:
        # For Wilcoxon, calculate critical values based on sample statistics
        mean_diff = np.mean(differences)
        std_diff = np.std(differences, ddof=1)
        n = len(differences)
        
        if n > 1:
            # 95% confidence interval for mean difference
            t_critical = stats.t.ppf(0.975, n-1)
            margin_error = t_critical * (std_diff / np.sqrt(n))
            critical_values = {
                'lower': float(mean_diff - margin_error),
                'upper': float(mean_diff + margin_error),
                'mean_difference': float(mean_diff)
            }
        else:
            critical_values = {
                'lower': 0,
                'upper': 0,
                'mean_difference': 0
            }
    except Exception as e:
        print(f"[Wilcoxon] Critical values calculation warning: {e}")
        critical_values = {
            'lower': 0,
            'upper': 0,
            'mean_difference': 0
        }

    # Prepare Q-Q plot data for differences with enhanced data
    try:
        (osm, osr), (slope, intercept, r) = stats.probplot(differences, dist="norm")
        
        # Calculate confidence intervals for Q-Q plot
        n = len(osm)
        if n > 2:
            # 95% confidence interval for Q-Q line
            se_slope = slope * np.sqrt((1 - r**2) / (n - 2))
            t_val = stats.t.ppf(0.975, n - 2)
            
            upper_slope = slope + t_val * se_slope
            lower_slope = slope - t_val * se_slope
            upper_intercept = intercept + t_val * se_slope * np.std(osm)
            lower_intercept = intercept - t_val * se_slope * np.std(osm)
            
            confidence_intervals = {
                'upper_slope': float(upper_slope),
                'lower_slope': float(lower_slope),
                'upper_intercept': float(upper_intercept),
                'lower_intercept': float(lower_intercept)
            }
        else:
            confidence_intervals = {
                'upper_slope': float(slope),
                'lower_slope': float(slope),
                'upper_intercept': float(intercept),
                'lower_intercept': float(intercept)
            }
        
        qq_data = {
            'theoretical_quantiles': [float(x) for x in osm],
            'ordered_values': [float(x) for x in osr],
            'slope': float(slope),
            'intercept': float(intercept),
            'r_squared': float(r**2),
            'confidence_intervals': confidence_intervals,
            'sample_size': n,
            'shapiro_stat': float(stats.shapiro(differences)[0]),
            'shapiro_p': float(stats.shapiro(differences)[1])
        }
    except Exception as e:
        print(f"[Wilcoxon] Q-Q plot data generation warning: {e}")
        qq_data = {
            'theoretical_quantiles': [],
            'ordered_values': [],
            'slope': 0,
            'intercept': 0,
            'r_squared': 0,
            'confidence_intervals': {
                'upper_slope': 0,
                'lower_slope': 0,
                'upper_intercept': 0,
                'lower_intercept': 0
            },
            'sample_size': 0,
            'shapiro_stat': 0,
            'shapiro_p': 1.0
        }

    # Prepare data for frontend plotting - KEEP THE OLD STRUCTURE
    plot_data = {
        'sample1': {
            'name': str(col1),
            'values': [float(x) for x in sample1],
            'count': int(len(sample1)),
            'mean': float(np.mean(sample1)),
            'median': float(np.median(sample1)),
            'std': float(np.std(sample1)),
            'min': float(np.min(sample1)),
            'max': float(np.max(sample1))
        },
        'sample2': {
            'name': str(col2),
            'values': [float(x) for x in sample2],
            'count': int(len(sample2)),
            'mean': float(np.mean(sample2)),
            'median': float(np.median(sample2)),
            'std': float(np.std(sample2)),
            'min': float(np.min(sample2)),
            'max': float(np.max(sample2))
        },
        'differences': {
            'values': [float(x) for x in differences],
            'count': int(len(differences)),
            'mean': float(np.mean(differences)),
            'median': float(np.median(differences)),
            'std': float(np.std(differences)),
            'min': float(np.min(differences)),
            'max': float(np.max(differences)),
            'q25': float(np.percentile(differences, 25)),
            'q75': float(np.percentile(differences, 75))
        },
        # Keep regression data for scatter plot
        'regression': regression_data
    }

    # Prepare response with localized labels
    test_name = 'Wilcoxon Signed-Rank Test' if language == 'en' else 'উইলকক্সন সাইনড র‍্যাঙ্ক টেস্ট'
    
    response_data = {
        'success': True,
        'test': test_name,
        'language': language,
        'statistic': float(stat),
        'p_value': float(p_value),
        'total_pairs': int(len(work)),  # This is required by frontend
        'column_names': {
            'sample1': str(col1),
            'sample2': str(col2)
        },
        'plot_data': plot_data,  # Frontend expects this
        'qq_data': qq_data,      # Frontend expects this
        'critical_values': critical_values,
        'metadata': {
            'significant': bool(p_value < 0.05),
            'alpha': 0.05,
            'test_type': 'paired',
            'normality_test': {
                'statistic': qq_data.get('shapiro_stat', 0),
                'p_value': qq_data.get('shapiro_p', 1.0),
                'normal': qq_data.get('shapiro_p', 1.0) > 0.05
            }
        }
    }
    return JsonResponse(response_data)


def process_anova_test(request, df: pd.DataFrame, col1: str, col2: str, user_id: str):
    """
    Performs One-Way ANOVA test and returns analysis results without generating plots.
    Frontend will handle visualization using the returned data.
    
    Args:
        request: Django request object
        df: DataFrame containing the data
        col1: Column name for grouping variable (categorical)
        col2: Column name for value variable (numeric)
        user_id: User identifier
        
    Returns:
        JsonResponse with test results and data for plotting
    """
    print(f"[ANOVA] cols: {col1}(group) | {col2}(value)")
    
    # Validate column names
    if col1 not in df.columns or col2 not in df.columns:
        return JsonResponse({
            'success': False, 
            'error': 'Invalid column names.'
        })

    # Extract language preference
    try:
        language = request.POST.get('language', 'en').lower()
    except Exception:
        language = 'en'
    
    if language not in ('en', 'bn'):
        language = 'en'

    # Prepare working dataset
    work = df[[col1, col2]].copy()
    work = work.dropna(subset=[col1, col2])
    print(f"[ANOVA] after dropna: {len(work)} rows")

    # Ensure grouping column is categorical
    if not pd.api.types.is_categorical_dtype(work[col1]):
        work[col1] = work[col1].astype('category')

    categories = list(work[col1].cat.categories)
    
    if len(categories) < 2:
        return JsonResponse({
            'success': False, 
            'error': 'Need at least 2 groups in the factor column for ANOVA.'
        })

    # Ensure value column is numeric
    if not pd.api.types.is_numeric_dtype(work[col2]):
        work[col2] = pd.to_numeric(work[col2], errors='coerce')
        work = work.dropna(subset=[col2])
        
        if not pd.api.types.is_numeric_dtype(work[col2]):
            return JsonResponse({
                'success': False, 
                'error': f'"{col2}" must be numeric for ANOVA.'
            })

    # Prepare groups for statistical test
    groups = [work.loc[work[col1] == g, col2].values for g in categories]
    
    if any(len(g) == 0 for g in groups):
        return JsonResponse({
            'success': False, 
            'error': 'Each group must contain at least one observation.'
        })

    # Perform One-Way ANOVA
    try:
        from scipy import stats
        f_statistic, p_value = stats.f_oneway(*groups)
        print(f"[ANOVA] result: F={f_statistic:.6f}, p={p_value:.6g}")
        
        # Calculate additional ANOVA statistics manually
        overall_mean = np.mean(work[col2])
        n_groups = len(categories)
        n_total = len(work)
        
        # Sum of Squares Between (SSB)
        ss_between = 0
        for i, group in enumerate(groups):
            group_mean = np.mean(group)
            group_size = len(group)
            ss_between += group_size * (group_mean - overall_mean) ** 2
        
        # Sum of Squares Within (SSW)
        ss_within = 0
        for i, group in enumerate(groups):
            group_mean = np.mean(group)
            ss_within += np.sum((group - group_mean) ** 2)
        
        # Degrees of freedom
        df_between = n_groups - 1
        df_within = n_total - n_groups
        
        # Mean Squares
        ms_between = ss_between / df_between
        ms_within = ss_within / df_within
        
        # F-statistic (should match scipy's calculation)
        f_statistic_calc = ms_between / ms_within
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'Error in ANOVA test: {e}'
        })

    # Prepare data for frontend plotting
    plot_data = []
    
    for category in categories:
        category_data = work.loc[work[col1] == category, col2].values
        
        # Calculate statistics for each group
        # IMPORTANT: Convert all numpy types to Python native types for JSON serialization
        plot_data.append({
            'category': str(category),
            'values': [float(x) for x in category_data],  # Convert numpy array to list of floats
            'count': int(len(category_data)),
            'mean': float(np.mean(category_data)),
            'median': float(np.median(category_data)),
            'std': float(np.std(category_data)),
            'min': float(np.min(category_data)),
            'max': float(np.max(category_data)),
            'q25': float(np.percentile(category_data, 25)),
            'q75': float(np.percentile(category_data, 75))
        })

    # Prepare response with localized labels
    test_name = 'One-Way ANOVA' if language == 'en' else 'এক-মুখী এনোভা'
    
    response_data = {
        'success': True,
        'test': test_name,
        'language': language,
        'f_statistic': float(f_statistic),
        'p_value': float(p_value),
        'df_between': int(df_between),
        'df_within': int(df_within),
        'sum_squares_between': float(ss_between),
        'sum_squares_within': float(ss_within),
        'mean_square_between': float(ms_between),
        'mean_square_within': float(ms_within),
        'n_groups': int(len(categories)),
        'total_observations': int(len(work)),
        'column_names': {
            'group': str(col1),
            'value': str(col2)
        },
        'plot_data': plot_data,
        'metadata': {
            'categories': [str(c) for c in categories],
            'significant': bool(p_value < 0.05),
            'alpha': 0.05
        }
    }
    
    return JsonResponse(response_data)


def process_ancova_test(request, df: pd.DataFrame, col_group: str, col_covariate: str, col_outcome: str, user_id: str):
    """
    Performs ANCOVA test and returns analysis results without generating plots.
    Frontend will handle visualization using the returned data.
    
    Args:
        request: Django request object
        df: DataFrame containing the data
        col_group: Column name for grouping variable (categorical)
        col_covariate: Column name for covariate variable (numeric)
        col_outcome: Column name for outcome variable (numeric)
        user_id: User identifier
        
    Returns:
        JsonResponse with test results and data for plotting
    """
    import numpy as np
    from sklearn.linear_model import LinearRegression
    from django.http import JsonResponse
    
    print(f"[ANCOVA] cols: {col_group}(group) | {col_covariate}(covariate) | {col_outcome}(outcome)")
    
    # Validate column names
    if col_group not in df.columns or col_covariate not in df.columns or col_outcome not in df.columns:
        return JsonResponse({
            'success': False, 
            'error': 'Invalid column names.'
        })

    # Extract language preference
    try:
        language = request.POST.get('language', 'en').lower()
    except Exception:
        language = 'en'
    
    if language not in ('en', 'bn'):
        language = 'en'

    # Prepare working dataset
    work = df[[col_group, col_covariate, col_outcome]].copy()
    work = work.dropna(subset=[col_group, col_covariate, col_outcome])
    print(f"[ANCOVA] after dropna: {len(work)} rows")

    # Ensure grouping column is categorical
    if not pd.api.types.is_categorical_dtype(work[col_group]):
        work[col_group] = work[col_group].astype('category')

    categories = list(work[col_group].cat.categories)
    
    if len(categories) < 2:
        return JsonResponse({
            'success': False, 
            'error': 'Need at least 2 groups in the factor column for ANCOVA.'
        })

    # Ensure covariate and outcome columns are numeric
    for col, col_name in [(col_covariate, "covariate"), (col_outcome, "outcome")]:
        if not pd.api.types.is_numeric_dtype(work[col]):
            work[col] = pd.to_numeric(work[col], errors='coerce')
            work = work.dropna(subset=[col])
            
            if not pd.api.types.is_numeric_dtype(work[col]):
                return JsonResponse({
                    'success': False, 
                    'error': f'"{col}" must be numeric for ANCOVA.'
                })

    # Check if we have enough data points
    if len(work) < len(categories) * 2:
        return JsonResponse({
            'success': False, 
            'error': 'Not enough data points for ANCOVA analysis.'
        })

    # Perform ANCOVA using statsmodels
    try:
        import statsmodels.api as sm
        import statsmodels.formula.api as smf
        
        # ANCOVA model: outcome ~ covariate + group
        formula = f"{col_outcome} ~ {col_covariate} + C({col_group})"
        ancova_model = smf.ols(formula, data=work).fit()
        ancova_table = sm.stats.anova_lm(ancova_model, typ=2)
        
        print(f"[ANCOVA] ANCOVA table:\n{ancova_table}")
        
        # Extract key statistics with error handling
        try:
            f_statistic_group = ancova_table.loc[f'C({col_group})', 'F']
            p_value_group = ancova_table.loc[f'C({col_group})', 'PR(>F)']
        except KeyError:
            f_statistic_group = float('nan')
            p_value_group = float('nan')
            
        try:
            f_statistic_covariate = ancova_table.loc[col_covariate, 'F']
            p_value_covariate = ancova_table.loc[col_covariate, 'PR(>F)']
        except KeyError:
            f_statistic_covariate = float('nan')
            p_value_covariate = float('nan')
        
        print(f"[ANCOVA] Group effect: F={f_statistic_group:.6f}, p={p_value_group:.6g}")
        print(f"[ANCOVA] Covariate effect: F={f_statistic_covariate:.6f}, p={p_value_covariate:.6g}")
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'Error in ANCOVA analysis: {e}'
        })

    # Prepare data for frontend plotting - KEEP OLD STRUCTURE BUT ADD NEW DATA
    plot_data = []
    data_points = []  # ADD FOR LINEAR REGRESSION STYLE
    residual_data_points = []  # ADD FOR LINEAR REGRESSION STYLE
    
    # Generate separate regression lines for each group
    regression_lines = {}  # ADD FOR LINEAR REGRESSION STYLE
    
    for category in categories:
        category_data = work.loc[work[col_group] == category]
        
        # Get residuals and fitted values for this category
        category_indices = category_data.index
        residuals = [ancova_model.resid[i] for i in category_indices]
        fitted_values = [ancova_model.fittedvalues[i] for i in category_indices]
        
        # Calculate regression line for this group (simple linear regression within group)
        try:
            X_group = category_data[col_covariate].values.reshape(-1, 1)
            y_group = category_data[col_outcome].values
            
            group_model = LinearRegression()
            group_model.fit(X_group, y_group)
            
            # Generate regression line points
            x_min_group = float(np.min(X_group))
            x_max_group = float(np.max(X_group))
            x_range_group = np.linspace(x_min_group, x_max_group, 50)
            y_range_group = group_model.predict(x_range_group.reshape(-1, 1))
            
            regression_lines[str(category)] = {
                'x_range': [float(x) for x in x_range_group],
                'y_range': [float(y) for y in y_range_group],
                'slope': float(group_model.coef_[0]),
                'intercept': float(group_model.intercept_)
            }
        except Exception as e:
            print(f"[ANCOVA] Error calculating regression line for group {category}: {e}")
            regression_lines[str(category)] = None

        # Add data points for scatter plot (NEW FOR LINEAR REGRESSION STYLE)
        for i in range(len(category_data)):
            data_points.append({
                'x': float(category_data[col_covariate].iloc[i]),
                'y': float(category_data[col_outcome].iloc[i]),
                'group': str(category),
                'predicted': float(fitted_values[i]),
                'residual': float(residuals[i])
            })
            
        # Add residual data points (NEW FOR LINEAR REGRESSION STYLE)
        for i in range(len(category_data)):
            residual_data_points.append({
                'x': float(category_data[col_covariate].iloc[i]),
                'y': float(residuals[i]),
                'group': str(category),
                'fitted': float(fitted_values[i])
            })

        # Calculate statistics for each group (KEEP OLD STRUCTURE)
        plot_data.append({
            'category': str(category),
            'values': [float(x) for x in category_data[col_outcome].values],
            'covariate_values': [float(x) for x in category_data[col_covariate].values],
            'count': int(len(category_data)),
            'mean_outcome': float(np.mean(category_data[col_outcome])),
            'mean_covariate': float(np.mean(category_data[col_covariate])),
            'std_outcome': float(np.std(category_data[col_outcome])),
            'std_covariate': float(np.std(category_data[col_covariate])),
            'min_outcome': float(np.min(category_data[col_outcome])),
            'max_outcome': float(np.max(category_data[col_outcome])),
            'min_covariate': float(np.min(category_data[col_covariate])),
            'max_covariate': float(np.max(category_data[col_covariate])),
            'q25_outcome': float(np.percentile(category_data[col_outcome], 25)),
            'q75_outcome': float(np.percentile(category_data[col_outcome], 75)),
            'median_outcome': float(np.median(category_data[col_outcome])),
            'residuals': [float(x) for x in residuals],
            'fitted_values': [float(x) for x in fitted_values]
        })

    # Calculate overall residual statistics (NEW FOR LINEAR REGRESSION STYLE)
    all_residuals = [point['residual'] for point in data_points]
    if all_residuals:
        residual_stats = {
            'mean': float(np.mean(all_residuals)),
            'std': float(np.std(all_residuals)),
            'min': float(np.min(all_residuals)),
            'max': float(np.max(all_residuals)),
            'q25': float(np.percentile(all_residuals, 25)),
            'q75': float(np.percentile(all_residuals, 75))
        }
    else:
        residual_stats = {
            'mean': 0.0, 'std': 0.0, 'min': 0.0, 'max': 0.0, 'q25': 0.0, 'q75': 0.0
        }

    # Prepare response with localized labels
    test_name = 'ANCOVA' if language == 'en' else 'এনকোভা'
    
    response_data = {
        'success': True,
        'test': test_name,
        'language': language,
        'f_statistic_group': float(f_statistic_group),
        'p_value_group': float(p_value_group),
        'f_statistic_covariate': float(f_statistic_covariate),
        'p_value_covariate': float(p_value_covariate),
        'n_groups': int(len(categories)),
        'total_observations': int(len(work)),
        'column_names': {
            'group': str(col_group),
            'covariate': str(col_covariate),
            'outcome': str(col_outcome)
        },
        'plot_data': plot_data,  # KEEP OLD STRUCTURE
        # ADD LINEAR REGRESSION-STYLE DATA STRUCTURES
        'data_points': data_points,
        'regression_lines': regression_lines,
        'residual_data': {
            'values': [point['y'] for point in residual_data_points],
            'fitted': [point['fitted'] for point in residual_data_points],
            'independent': [point['x'] for point in residual_data_points],
            'statistics': residual_stats
        },
        'critical_values': {
            'lower': float(np.percentile(all_residuals, 2.5)) if all_residuals else 0.0,
            'upper': float(np.percentile(all_residuals, 97.5)) if all_residuals else 0.0,
            'mean_difference': float(np.mean(all_residuals)) if all_residuals else 0.0
        },
        'metadata': {
            'categories': [str(c) for c in categories],
            'significant_group': bool(p_value_group < 0.05) if not np.isnan(p_value_group) else False,
            'significant_covariate': bool(p_value_covariate < 0.05) if not np.isnan(p_value_covariate) else False,
            'alpha': 0.05,
            'r_squared': float(ancova_model.rsquared),
            'adj_r_squared': float(ancova_model.rsquared_adj),
            'test_type': 'ancova'
        }
    }
    
    return JsonResponse(response_data)


def process_linear_regression_test(request, df: pd.DataFrame, col1: str, col2: str, user_id: str):
    import numpy as np
    """
    Performs linear regression analysis and returns results without generating plots.
    Frontend will handle visualization using the returned data.
    
    Args:
        request: Django request object
        df: DataFrame containing the data
        col1: Column name for independent variable (predictor)
        col2: Column name for dependent variable (response)
        user_id: User identifier
        
    Returns:
        JsonResponse with regression results and data for plotting
    """
    print(f"[Linear Regression] cols: {col1}(independent) | {col2}(dependent)")
    
    # Validate column names
    if col1 not in df.columns or col2 not in df.columns:
        return JsonResponse({
            'success': False, 
            'error': 'Invalid column names.'
        })

    # Extract language preference
    try:
        language = request.POST.get('language', 'en').lower()
    except Exception:
        language = 'en'
    
    if language not in ('en', 'bn'):
        language = 'en'

    # Prepare working dataset
    work = df[[col1, col2]].copy()
    work = work.dropna(subset=[col1, col2])
    print(f"[Linear Regression] after dropna: {len(work)} rows")

    # Ensure both columns are numeric
    for col in [col1, col2]:
        if not pd.api.types.is_numeric_dtype(work[col]):
            work[col] = pd.to_numeric(work[col], errors='coerce')
            work = work.dropna(subset=[col])
            
            if not pd.api.types.is_numeric_dtype(work[col]):
                return JsonResponse({
                    'success': False, 
                    'error': f'"{col}" must be numeric for linear regression.'
                })

    # Check for sufficient data
    if len(work) < 2:
        return JsonResponse({
            'success': False, 
            'error': 'Need at least 2 observations for linear regression.'
        })

    # Prepare data for linear regression
    X = work[col1].values.reshape(-1, 1)
    y = work[col2].values

    # Perform linear regression
    try:
        from sklearn.linear_model import LinearRegression
        from sklearn.metrics import r2_score, mean_squared_error
        import numpy as np
        
        model = LinearRegression()
        model.fit(X, y)
        y_pred = model.predict(X)
        
        # Calculate statistics
        n = len(y)
        p = 1  # number of predictors
        r_squared = r2_score(y, y_pred)
        adjusted_r_squared = 1 - (1 - r_squared) * (n - 1) / (n - p - 1) if n > p + 1 else r_squared
        mse = mean_squared_error(y, y_pred)
        rmse = np.sqrt(mse)
        
        # Calculate standard errors and p-values
        from scipy import stats as scipy_stats
        X_with_intercept = np.column_stack([np.ones(len(X)), X.reshape(-1)])
        params = np.append(model.intercept_, model.coef_)
        
        # Calculate residuals and standard errors
        residuals = y - y_pred
        residual_sum_of_squares = np.sum(residuals**2)
        degrees_of_freedom = n - X_with_intercept.shape[1]
        mse_residual = residual_sum_of_squares / degrees_of_freedom
        
        # Standard errors of coefficients
        try:
            cov_matrix = mse_residual * np.linalg.inv(X_with_intercept.T @ X_with_intercept)
            standard_errors = np.sqrt(np.diag(cov_matrix))
            
            # t-statistics and p-values
            t_statistics = params / standard_errors
            p_values = [2 * (1 - scipy_stats.t.cdf(np.abs(t), degrees_of_freedom)) for t in t_statistics]
        except:
            # Fallback if matrix inversion fails
            standard_errors = [np.nan, np.nan]
            p_values = [np.nan, np.nan]
        
        print(f"[Linear Regression] result: R²={r_squared:.6f}, coef={model.coef_[0]:.6f}, intercept={model.intercept_:.6f}")
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'Error in linear regression: {e}'
        })

    # Prepare data for frontend plotting - KEEP ORIGINAL STRUCTURE
    data_points = []
    for i in range(len(X)):
        data_points.append({
            'x': float(X[i][0]),
            'y': float(y[i]),
            'predicted': float(y_pred[i]),
            'residual': float(y[i] - y_pred[i])
        })

    # Generate regression line points for plotting
    x_min, x_max = float(np.min(X)), float(np.max(X))
    x_range = np.linspace(x_min, x_max, 100)
    y_range = model.predict(x_range.reshape(-1, 1))
    
    regression_line = []
    for i in range(len(x_range)):
        regression_line.append({
            'x': float(x_range[i]),
            'y': float(y_range[i])
        })

    # Calculate confidence intervals (95%)
    try:
        from scipy.stats import t
        t_val = t.ppf(0.975, degrees_of_freedom)  # 95% confidence
        
        se_fit = np.sqrt(np.sum((X_with_intercept * (cov_matrix @ X_with_intercept.T).T), axis=1))
        confidence_intervals = []
        
        for i in range(len(x_range)):
            x_point = np.array([1, x_range[i]])
            se = np.sqrt(mse_residual * (x_point @ np.linalg.inv(X_with_intercept.T @ X_with_intercept) @ x_point.T))
            margin = t_val * se
            confidence_intervals.append({
                'x': float(x_range[i]),
                'y_lower': float(y_range[i] - margin),
                'y_upper': float(y_range[i] + margin)
            })
    except:
        confidence_intervals = []

    # ADD WILCOXON-STYLE PLOT_DATA STRUCTURE FOR SCATTER PLOT
    plot_data = {
        'sample1': {
            'name': str(col1),
            'values': [float(x[0]) for x in X],
            'count': int(len(X)),
            'mean': float(np.mean(X)),
            'median': float(np.median(X)),
            'std': float(np.std(X)),
            'min': float(np.min(X)),
            'max': float(np.max(X))
        },
        'sample2': {
            'name': str(col2),
            'values': [float(y_val) for y_val in y],
            'count': int(len(y)),
            'mean': float(np.mean(y)),
            'median': float(np.median(y)),
            'std': float(np.std(y)),
            'min': float(np.min(y)),
            'max': float(np.max(y))
        },
        'regression': {
            'slope': float(model.coef_[0]),
            'intercept': float(model.intercept_),
            'r_squared': float(r_squared),
            'p_value': float(p_values[1]) if len(p_values) > 1 else 1.0
        }
    }

    # Prepare response with localized labels
    test_name = 'Linear Regression' if language == 'en' else 'লিনিয়ার রিগ্রেশন'
    
    # Determine significance
    coef_significant = p_values[1] < 0.05 if not np.isnan(p_values[1]) else False
    overall_significant = coef_significant

    response_data = {
        'success': True,
        'test': test_name,
        'language': language,
        'intercept': float(model.intercept_),
        'coefficient': float(model.coef_[0]),
        'r_squared': float(r_squared),
        'adjusted_r_squared': float(adjusted_r_squared),
        'mean_squared_error': float(mse),
        'root_mean_squared_error': float(rmse),
        'standard_errors': {
            'intercept': float(standard_errors[0]) if len(standard_errors) > 0 else float('nan'),
            'coefficient': float(standard_errors[1]) if len(standard_errors) > 1 else float('nan')
        },
        'p_values': {
            'intercept': float(p_values[0]) if len(p_values) > 0 else float('nan'),
            'coefficient': float(p_values[1]) if len(p_values) > 1 else float('nan')
        },
        't_statistics': {
            'intercept': float(t_statistics[0]) if len(t_statistics) > 0 else float('nan'),
            'coefficient': float(t_statistics[1]) if len(t_statistics) > 1 else float('nan')
        },
        'degrees_of_freedom': int(degrees_of_freedom),
        'n_observations': int(n),
        'column_names': {
            'independent': str(col1),
            'dependent': str(col2),
            'sample1': str(col1),  # ADD FOR WILCOXON COMPATIBILITY
            'sample2': str(col2)   # ADD FOR WILCOXON COMPATIBILITY
        },
        'data_points': data_points,
        'regression_line': regression_line,
        'confidence_intervals': confidence_intervals,
        # ADD WILCOXON COMPATIBLE FIELDS
        'plot_data': plot_data,  # ADD THIS FOR SCATTER PLOT
        'statistic': float(t_statistics[1]) if len(t_statistics) > 1 else float('nan'),
        'p_value': float(p_values[1]) if len(p_values) > 1 else float('nan'),
        'total_pairs': int(n),
        'critical_values': {
            'lower': float(np.percentile(residuals, 2.5)),
            'upper': float(np.percentile(residuals, 97.5)),
            'mean_difference': float(np.mean(residuals))
        },
        # In your process_linear_regression_test function, update the residual_data section:
        'residual_data': {
            'values': [float(res) for res in residuals],
            'fitted': [float(pred) for pred in y_pred],
            'independent': [float(x[0]) for x in X],
            'statistics': {
                'mean': float(np.mean(residuals)),
                'std': float(np.std(residuals)),
                'min': float(np.min(residuals)),
                'max': float(np.max(residuals)),
                'q25': float(np.percentile(residuals, 25)),
                'q75': float(np.percentile(residuals, 75))
            }
        },        
        'metadata': {
            'significant': bool(overall_significant),
            'coefficient_significant': bool(coef_significant),
            'alpha': 0.05,
            'test_type': 'regression'
        }
    }
    
    return JsonResponse(response_data)


def process_shapiro_test(request, df: pd.DataFrame, col1: str, user_id: str):
    """
    Performs Shapiro-Wilk normality test and returns analysis results without generating plots.
    Frontend will handle visualization using the returned data.
    
    Args:
        request: Django request object
        df: DataFrame containing the data
        col1: Column name for the numeric variable to test for normality
        user_id: User identifier
        
    Returns:
        JsonResponse with test results and data for plotting
    """
    print(f"[Shapiro] column: {col1}")
    
    # Validate column name
    if col1 not in df.columns:
        return JsonResponse({
            'success': False, 
            'error': 'Invalid column name.'
        })

    # Extract language preference
    try:
        language = request.POST.get('language', 'en').lower()
    except Exception:
        language = 'en'
    
    if language not in ('en', 'bn'):
        language = 'en'

    # Prepare working dataset
    work = df[[col1]].copy()
    work = work.dropna(subset=[col1])
    print(f"[Shapiro] after dropna: {len(work)} rows")

    # Ensure value column is numeric
    if not pd.api.types.is_numeric_dtype(work[col1]):
        work[col1] = pd.to_numeric(work[col1], errors='coerce')
        work = work.dropna(subset=[col1])
        
        if not pd.api.types.is_numeric_dtype(work[col1]):
            return JsonResponse({
                'success': False, 
                'error': f'"{col1}" must be numeric for Shapiro-Wilk test.'
            })

    # Get the data values
    data_values = work[col1].values
    
    if len(data_values) < 3:
        return JsonResponse({
            'success': False, 
            'error': 'Need at least 3 observations for Shapiro-Wilk test.'
        })

    # Perform Shapiro-Wilk test
    try:
        stat, p_value = stats.shapiro(data_values)
        print(f"[Shapiro] result: W={stat:.6f}, p={p_value:.6g}")
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'Error in Shapiro-Wilk test: {e}'
        })

    # Calculate additional statistics for frontend
    mean_val = float(np.mean(data_values))
    std_val = float(np.std(data_values, ddof=1))
    median_val = float(np.median(data_values))
    min_val = float(np.min(data_values))
    max_val = float(np.max(data_values))
    
    # Calculate skewness and kurtosis
    try:
        skewness_val = float(stats.skew(data_values))
        kurtosis_val = float(stats.kurtosis(data_values))
    except:
        skewness_val = 0.0
        kurtosis_val = 0.0

    # Prepare Q-Q plot data using scipy.probplot (same as Wilcoxon)
    try:
        (osm, osr), (slope, intercept, r) = stats.probplot(data_values, dist="norm")
        
        # Calculate confidence intervals for Q-Q plot
        n = len(osm)
        if n > 2:
            # 95% confidence interval for Q-Q line
            se_slope = slope * np.sqrt((1 - r**2) / (n - 2))
            t_val = stats.t.ppf(0.975, n - 2)
            
            upper_slope = slope + t_val * se_slope
            lower_slope = slope - t_val * se_slope
            upper_intercept = intercept + t_val * se_slope * np.std(osm)
            lower_intercept = intercept - t_val * se_slope * np.std(osm)
            
            confidence_intervals = {
                'upper_slope': float(upper_slope),
                'lower_slope': float(lower_slope),
                'upper_intercept': float(upper_intercept),
                'lower_intercept': float(lower_intercept)
            }
        else:
            confidence_intervals = {
                'upper_slope': float(slope),
                'lower_slope': float(slope),
                'upper_intercept': float(intercept),
                'lower_intercept': float(intercept)
            }
        
        qq_data = {
            'theoretical_quantiles': [float(x) for x in osm],
            'ordered_values': [float(x) for x in osr],
            'slope': float(slope),
            'intercept': float(intercept),
            'r_squared': float(r**2),
            'confidence_intervals': confidence_intervals,
            'sample_size': n,
            'shapiro_stat': float(stat),
            'shapiro_p': float(p_value)
        }
    except Exception as e:
        print(f"[Shapiro] Q-Q plot data generation warning: {e}")
        qq_data = {
            'theoretical_quantiles': [],
            'ordered_values': [],
            'slope': 0,
            'intercept': 0,
            'r_squared': 0,
            'confidence_intervals': {
                'upper_slope': 0,
                'lower_slope': 0,
                'upper_intercept': 0,
                'lower_intercept': 0
            },
            'sample_size': 0,
            'shapiro_stat': float(stat),
            'shapiro_p': float(p_value)
        }

    # Prepare data for frontend plotting - FOLLOW WILCOXON STRUCTURE
    plot_data = {
        'sample': {
            'name': str(col1),
            'values': [float(x) for x in data_values],
            'count': int(len(data_values)),
            'mean': float(mean_val),
            'median': float(median_val),
            'std': float(std_val),
            'min': float(min_val),
            'max': float(max_val),
            'q25': float(np.percentile(data_values, 25)),
            'q75': float(np.percentile(data_values, 75))
        },
        # Add normal curve data for histogram overlay
        'normal_curve': {
            'mean': float(mean_val),
            'std': float(std_val)
        }
    }

    # Prepare response with localized labels - FOLLOW WILCOXON STRUCTURE
    test_name = 'Shapiro-Wilk Normality Test' if language == 'en' else 'শ্যাপিরো-উইল্ক স্বাভাবিকতা পরীক্ষা'
    
    response_data = {
        'success': True,
        'test': test_name,
        'language': language,
        'statistic': float(stat),
        'p_value': float(p_value),
        'n_observations': int(len(data_values)),  # This is required by frontend
        'column_names': {
            'sample': str(col1)
        },
        'plot_data': plot_data,  # Frontend expects this (same as Wilcoxon)
        'qq_data': qq_data,      # Frontend expects this (same as Wilcoxon)
        'descriptive_stats': {
            'mean': mean_val,
            'std_dev': std_val,
            'median': median_val,
            'min': min_val,
            'max': max_val,
            'skewness': skewness_val,
            'kurtosis': kurtosis_val
        },
        'metadata': {
            'normal': bool(p_value >= 0.05),
            'alpha': 0.05,
            'test_type': 'normality',
            'interpretation': 'Data follows normal distribution (p ≥ 0.05)' if p_value >= 0.05 else 'Data does not follow normal distribution (p < 0.05)'
        }
    }
    
    return JsonResponse(response_data)


def process_eda_distribution(request, df: pd.DataFrame, col: str, user_id: str):
    """
    Performs EDA distribution analysis and returns data for frontend visualization.
    Frontend will handle histogram, KDE, and distribution plots using the returned data.
    
    Args:
        request: Django request object
        df: DataFrame containing the data
        col: Column name for distribution analysis (numeric)
        user_id: User identifier
        
    Returns:
        JsonResponse with distribution statistics and data for plotting
    """
    import os
    import numpy as np
    import seaborn as sns
    import matplotlib.pyplot as plt
    from scipy import stats
    from django.http import JsonResponse

    try:
        language = request.POST.get('language', 'en').lower()
        if language not in ('en', 'bn'):
            language = 'en'

        # Extract plot parameters
        histogram_bins = request.POST.get('histogram_bins', 'auto')
        kde_bandwidth = request.POST.get('kde_bandwidth', None)
        
        # Parse histogram bins
        if histogram_bins != 'auto' and histogram_bins != '':
            try:
                histogram_bins = int(histogram_bins)
                if histogram_bins <= 0:
                    histogram_bins = 'auto'
            except (ValueError, TypeError):
                histogram_bins = 'auto'
        
        # Parse KDE bandwidth
        if kde_bandwidth and kde_bandwidth not in ['auto', '']:
            try:
                kde_bandwidth = float(kde_bandwidth)
                if kde_bandwidth <= 0:
                    kde_bandwidth = None
            except (ValueError, TypeError):
                kde_bandwidth = None

        # Validate column exists and is numeric
        if col not in df.columns:
            return JsonResponse({
                'success': False, 
                'error': f'Column "{col}" not found in dataset.'
            })

        # Prepare working data
        work = df[[col]].copy()
        work = work.dropna(subset=[col])
        
        # Ensure column is numeric
        if not pd.api.types.is_numeric_dtype(work[col]):
            work[col] = pd.to_numeric(work[col], errors='coerce')
            work = work.dropna(subset=[col])
            
            if not pd.api.types.is_numeric_dtype(work[col]):
                return JsonResponse({
                    'success': False, 
                    'error': f'Column "{col}" must be numeric for distribution analysis.'
                })

        if len(work) == 0:
            return JsonResponse({
                'success': False, 
                'error': 'No valid numeric data available for analysis.'
            })

        print(f"[EDA Distribution] Valid data points: {len(work)}")

        # Calculate distribution statistics
        values = work[col].values
        stats_summary = {
            'count': int(len(values)),
            'mean': float(np.mean(values)),
            'median': float(np.median(values)),
            'std': float(np.std(values)),
            'variance': float(np.var(values)),
            'min': float(np.min(values)),
            'max': float(np.max(values)),
            'q25': float(np.percentile(values, 25)),
            'q75': float(np.percentile(values, 75)),
            'skewness': float(stats.skew(values)),
            'kurtosis': float(stats.kurtosis(values))
        }

        # Generate histogram data (like old backend but return data instead of images)
        hist_counts, bin_edges = np.histogram(values, bins=histogram_bins)
        bin_centers = (bin_edges[:-1] + bin_edges[1:]) / 2
        bin_widths = np.diff(bin_edges)
        
        histogram_bins_data = []
        for i in range(len(hist_counts)):
            histogram_bins_data.append({
                'x': float(bin_centers[i]),
                'y': int(hist_counts[i]),
                'x0': float(bin_edges[i]),
                'x1': float(bin_edges[i + 1]),
                'width': float(bin_widths[i]),
                'count': int(hist_counts[i])
            })

        hist_data = {
            'bins': histogram_bins_data,
            'total_count': int(np.sum(hist_counts)),
            'bin_count': len(hist_counts),
            'range': [float(bin_edges[0]), float(bin_edges[-1])],
            'bins_used': histogram_bins if isinstance(histogram_bins, int) else 'auto'
        }

        # Generate KDE data (like old backend but return data instead of images)
        kde = stats.gaussian_kde(values)
        
        # Set custom bandwidth if provided
        if kde_bandwidth is not None:
            kde.set_bandwidth(kde_bandwidth)
        
        data_min = np.min(values)
        data_max = np.max(values)
        data_range = data_max - data_min
        x_points = np.linspace(data_min - 0.1 * data_range, data_max + 0.1 * data_range, 200)
        y_points = kde(x_points)
        
        # Normalize to make it comparable to histogram (like seaborn does)
        y_points_normalized = y_points * len(values) * (x_points[1] - x_points[0])
        
        kde_curve = []
        for i in range(len(x_points)):
            kde_curve.append({
                'x': float(x_points[i]),
                'y': float(y_points_normalized[i]),
                'density': float(y_points[i])
            })

        kde_data = {
            'curve': kde_curve,
            'bandwidth': float(kde.factor),
            'range': [float(x_points[0]), float(x_points[-1])],
            'peak_density': float(np.max(y_points))
        }

        # Generate combined distribution data (histogram + KDE)
        # This mimics seaborn's displot with kde=True
        distribution_data = {
            'histogram': hist_data,
            'kde': kde_data
        }

        # Prepare response with localized labels
        test_name = 'EDA: Distribution Analysis' if language == 'en' else 'ইডিএ: বন্টন বিশ্লেষণ'
        
        response_data = {
            'success': True,
            'test': test_name,
            'language': language,
            'column_name': str(col),
            'statistics': stats_summary,
            'plot_data': distribution_data,
            'metadata': {
                'data_points': int(len(values)),
                'range': [float(np.min(values)), float(np.max(values))],
                'iqr': float(np.percentile(values, 75) - np.percentile(values, 25)),
                'histogram_bins_used': hist_data['bins_used'],
                'kde_bandwidth_used': kde_data['bandwidth']
            }
        }
        
        return JsonResponse(response_data)

    except Exception as e:
        print(f"[EDA Distribution Error] {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})


def process_eda_swarm_plot(request, df: pd.DataFrame, col_cat: str, col_num: str, user_id: str):
    """
    Processes data for Swarm Plot visualization and returns structured data.
    Frontend will handle the actual plotting using the returned data.
    
    Args:
        request: Django request object
        df: DataFrame containing the data
        col_cat: Column name for categorical variable
        col_num: Column name for numeric variable  
        user_id: User identifier
        
    Returns:
        JsonResponse with swarm plot data and statistics
    """
    print(f"[Swarm Plot] cols: {col_cat}(categorical) | {col_num}(numeric)")
    
    # Validate column names
    if col_cat not in df.columns or col_num not in df.columns:
        return JsonResponse({
            'success': False, 
            'error': 'Invalid column names.'
        })

    # Extract language preference
    try:
        language = request.POST.get('language', 'en').lower()
    except Exception:
        language = 'en'
    
    if language not in ('en', 'bn'):
        language = 'en'

    # Prepare working dataset
    work = df[[col_cat, col_num]].copy()
    work = work.dropna(subset=[col_cat, col_num])
    print(f"[Swarm Plot] after dropna: {len(work)} rows")

    # Ensure categorical column is properly typed
    if not pd.api.types.is_categorical_dtype(work[col_cat]):
        work[col_cat] = work[col_cat].astype('category')

    categories = list(work[col_cat].cat.categories)
    
    if len(categories) == 0:
        return JsonResponse({
            'success': False, 
            'error': 'No valid categories found in the categorical column.'
        })

    # Ensure numeric column is properly typed
    if not pd.api.types.is_numeric_dtype(work[col_num]):
        work[col_num] = pd.to_numeric(work[col_num], errors='coerce')
        work = work.dropna(subset=[col_num])
        
        if not pd.api.types.is_numeric_dtype(work[col_num]):
            return JsonResponse({
                'success': False, 
                'error': f'"{col_num}" must be numeric for Swarm Plot.'
            })

    # Prepare swarm plot data for frontend
    plot_data = []
    
    for category in categories:
        category_data = work.loc[work[col_cat] == category, col_num].values
        
        # Calculate statistics for each category
        plot_data.append({
            'category': str(category),
            'values': [float(x) for x in category_data],  # Convert numpy array to list of floats
            'count': int(len(category_data)),
            'mean': float(np.mean(category_data)),
            'median': float(np.median(category_data)),
            'std': float(np.std(category_data)),
            'min': float(np.min(category_data)),
            'max': float(np.max(category_data)),
            'q25': float(np.percentile(category_data, 25)),
            'q75': float(np.percentile(category_data, 75))
        })

    # Calculate overall statistics
    all_values = work[col_num].values
    overall_stats = {
        'total_count': int(len(work)),
        'global_mean': float(np.mean(all_values)),
        'global_median': float(np.median(all_values)),
        'global_std': float(np.std(all_values)),
        'global_min': float(np.min(all_values)),
        'global_max': float(np.max(all_values))
    }

    # Prepare response with localized labels
    test_name = 'Swarm Plot' if language == 'en' else 'সোয়ার্ম প্লট'
    
    response_data = {
        'success': True,
        'test': test_name,
        'language': language,
        'column_names': {
            'categorical': str(col_cat),
            'numeric': str(col_num)
        },
        'plot_data': plot_data,
        'overall_stats': overall_stats,
        'metadata': {
            'categories': [str(c) for c in categories],
            'n_categories': int(len(categories)),
            'category_counts': {str(cat): int(len(work[work[col_cat] == cat])) for cat in categories}
        }
    }
    
    return JsonResponse(response_data)


def process_bar_chart_test(request, df: pd.DataFrame, col: str, user_id: str, orientation: str, ordinal_mappings: dict):
    """
    Processes bar chart data and returns statistics for frontend visualization.
    Frontend will handle the actual plotting using the returned data.
    
    Args:
        request: Django request object
        df: DataFrame containing the data
        col: Column name for the categorical variable
        user_id: User identifier
        orientation: Chart orientation ('horizontal' or 'vertical')
        ordinal_mappings: Dictionary for ordinal value mappings
        
    Returns:
        JsonResponse with bar chart data and statistics
    """
    print(f"[Bar Chart] Processing column: {col}, orientation: {orientation}")
    
    # Validate column name
    if col not in df.columns:
        return JsonResponse({
            'success': False, 
            'error': f'Column "{col}" not found in dataset.'
        })

    # Extract language preference
    try:
        language = request.POST.get('language', 'en').lower()
    except Exception:
        language = 'en'
    
    if language not in ('en', 'bn'):
        language = 'en'

    # Validate orientation
    orientation = orientation if orientation in ('horizontal', 'vertical') else 'vertical'

    # Prepare working data
    work = df[[col]].copy()
    work = work.dropna(subset=[col])
    print(f"[Bar Chart] after dropna: {len(work)} rows")

    # Calculate counts and prepare categories
    counts = work[col].value_counts(dropna=False).sort_index()
    
    if len(counts) == 0:
        return JsonResponse({
            'success': False, 
            'error': 'No data available after cleaning.'
        })

    # Prepare category labels using ordinal mappings if available
    categories = []
    category_data = []
    
    for category_idx, count in counts.items():
        # Use ordinal mapping if available, otherwise use the raw value
        display_label = ordinal_mappings.get(col, {}).get(int(category_idx), str(category_idx))
        print(display_label)
        
        categories.append(display_label)
        category_data.append({
            'category': str(display_label),
            'count': int(count),
            'percentage': float((count / len(work)) * 100)
        })


    # Sort categories by count (descending) for better visualization
    category_data.sort(key=lambda x: x['count'], reverse=True)

    # Prepare response with localized labels
    test_name = 'Bar Chart' if language == 'en' else 'বার চার্ট'
    orientation_label = 'Horizontal' if orientation == 'horizontal' else 'Vertical'
    orientation_label_bn = 'অনুভূমিক' if orientation == 'horizontal' else 'উল্লম্ব'

    response_data = {
        'success': True,
        'test': test_name,
        'language': language,
        'orientation': orientation,
        'column_name': str(col),
        'total_observations': int(len(work)),
        'unique_categories': int(len(categories)),
        'plot_data': category_data,
        'statistics': {
            'max_count': int(max([item['count'] for item in category_data])),
            'min_count': int(min([item['count'] for item in category_data])),
            'mean_count': float(np.mean([item['count'] for item in category_data])),
            'median_count': float(np.median([item['count'] for item in category_data])),
            'total_count': int(sum([item['count'] for item in category_data]))
        },
        'metadata': {
            'orientation_display': orientation_label if language == 'en' else orientation_label_bn,
            'categories': [item['category'] for item in category_data],
            'has_ordinal_mappings': col in ordinal_mappings
        }
    }
    
    return JsonResponse(response_data)


def process_pie_chart(request, df: pd.DataFrame, col: str, user_id: str, ordinal_mappings: dict):
    """
    Processes pie chart data and returns structured results for frontend visualization.
    Frontend will handle the actual chart rendering.
    
    Args:
        request: Django request object
        df: DataFrame containing the data
        col: Column name for categorical variable
        user_id: User identifier
        ordinal_mappings: Ordinal value mappings
        
    Returns:
        JsonResponse with pie chart data and statistics
    """
    print(f"[Pie Chart] Processing column: {col}")
    
    # Validate column exists
    if col not in df.columns:
        return JsonResponse({
            'success': False, 
            'error': f'Column "{col}" not found in dataset.'
        })

    # Extract language preference
    try:
        language = request.POST.get('language', 'en').lower()
    except Exception:
        language = 'en'
    
    if language not in ('en', 'bn'):
        language = 'en'

    # Prepare working data
    work = df[[col]].copy()
    work[col] = work[col].fillna('NaN')

    if col in ordinal_mappings:
        def map_value(x):
            try:
                if isinstance(x, str) and (x.replace('.', '', 1).isdigit()):
                    num = float(x)
                    if num.is_integer():
                        return ordinal_mappings[col].get(int(num), x)
                return ordinal_mappings[col].get(x, x)
            except Exception:
                return x

        work[col] = work[col].map(map_value)

    work[col] = work[col].astype(str)   
    
    # Calculate value counts
    value_counts = work[col].value_counts()
    categories = value_counts.index.tolist()
    counts = value_counts.values.tolist()
    
    total_observations = sum(counts)
    
    # Check for too many categories
    if len(categories) > 10:
        warning_msg = (
            "শ্রেণি সংখ্যা ১০ এর বেশি। পাই চার্ট উপযুক্ত নয়। বার চার্ট ব্যবহার করুন।"
            if language == 'bn' 
            else "Too many categories (>10). Pie charts are not suitable. Please try a bar chart instead."
        )
        return JsonResponse({
            'success': False, 
            'error': warning_msg
        })

    # Calculate percentages and prepare data for frontend
    plot_data = []
    default_colors = [
        "#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd",
        "#8c564b", "#e377c2", "#7f7f7f", "#bcbd22", "#17becf"
    ]
    
    for i, (category, count) in enumerate(zip(categories, counts)):
        percentage = (count / total_observations) * 100
        
        plot_data.append({
            'category': str(category),
            'count': int(count),
            'percentage': float(percentage),
            'color': default_colors[i % len(default_colors)],
            'is_small_slice': percentage < 5.0  # Flag for frontend small slice handling
        })

    # Sort by count (descending) for better visualization
    plot_data.sort(key=lambda x: x['count'], reverse=True)

    # Prepare response with localized labels
    test_name = 'Pie Chart Analysis' if language == 'en' else 'পাই চার্ট বিশ্লেষণ'

    response_data = {
        'success': True,
        'test': test_name,
        'language': language,
        'column_name': str(col),
        'total_observations': int(total_observations),
        'n_categories': int(len(categories)),
        'plot_data': plot_data,  # This should be accessible as results.plot_data in frontend
        'metadata': {
            'has_small_slices': any(item['is_small_slice'] for item in plot_data),
            'max_categories_reached': len(categories) >= 10,
            'most_common_category': plot_data[0]['category'] if plot_data else None,
            'most_common_percentage': plot_data[0]['percentage'] if plot_data else 0.0
        }
    }
    
    print(f"[Pie Chart] Response data keys: {list(response_data.keys())}")
    print(f"[Pie Chart] Plot data length: {len(plot_data)}")
    
    return JsonResponse(response_data)


''' old backend code of bar chart and pie chart
def process_bar_chart_test(request, df, col, user_id, orientation, ordinal_mappings):
    import os, re, math
    import numpy as np
    import pandas as pd
    import seaborn as sns
    import matplotlib.pyplot as plt
    import matplotlib.font_manager as fm
    from django.http import JsonResponse
    from django.conf import settings
    from PIL import Image, ImageDraw, ImageFont
    try:
        from googletrans import Translator
    except Exception:
        class Translator:
            def translate(self, txt, dest="bn"):
                class _R:
                    def __init__(self, t): self.text = t
                    text = t
                return _R(txt)

    try:
        # ---------------- Parameters ----------------
        language = request.POST.get("language", "en")
        fmt = request.POST.get("format", "png").lower()
        fmt = fmt if fmt in ("png", "jpg", "jpeg", "pdf", "tiff") else "png"
        pil_fmt = {"png": "PNG", "jpg": "JPEG", "jpeg": "JPEG", "pdf": "PDF", "tiff": "TIFF"}[fmt]
        orientation = orientation if orientation in ("horizontal", "vertical") else "vertical"

        use_default = request.POST.get("use_default", "true") == "true"
        if use_default:
            dpi, width, height = 300, 800, 600
            label_font_size, caption_font_size, tick_font_size = 15, 50, 16
            cat_font_size, val_font_size = 44, 10
            bar_color, theme = "steelblue", "darkgrid"
        else:
            dpi = int(request.POST.get("dpi", 300))
            width, height = map(int, request.POST.get("image_size", "800x600").split("x"))
            label_font_size = int(request.POST.get("label_font_size", 15))
            caption_font_size = int(request.POST.get("caption_font_size", 50))
            tick_font_size = int(request.POST.get("tick_font_size", 16))
            cat_font_size = int(request.POST.get("cat_font_size", 44))
            val_font_size = int(request.POST.get("val_font_size", 10))
            bar_color = request.POST.get("bar_color", "steelblue")
            theme = request.POST.get("theme", "darkgrid")

        sns.set_theme(style=theme)

        # ---------------- Fonts & Translator ----------------
        font_path = os.path.join(settings.BASE_DIR, "NotoSansBengali-Regular.ttf")
        fm.fontManager.addfont(font_path)
        bengali_font = fm.FontProperties(fname=font_path).get_name()
        if language == "bn":
            plt.rcParams["font.family"] = bengali_font

        translator = Translator()
        digit_map = str.maketrans("0123456789", "০১২৩৪৫৬৭৮৯")

        def translate(txt):
            try:
                return translator.translate(txt, dest="bn").text if language == "bn" else txt
            except Exception:
                return txt

        def map_digits(txt):
            return txt.translate(digit_map) if language == "bn" else txt

        tick_prop = fm.FontProperties(fname=font_path, size=tick_font_size)
        caption_font = ImageFont.truetype(font_path, size=caption_font_size)
        label_font = ImageFont.truetype(font_path, size=label_font_size)
        cat_font = ImageFont.truetype(font_path, size=cat_font_size)
        val_font = ImageFont.truetype(font_path, size=val_font_size)

        # ---------------- Data Prep ----------------
        if col not in df.columns:
            return JsonResponse({"success": False, "error": f"Column {col} not found"})

        counts = df[col].fillna("NaN").value_counts(dropna=False).sort_index()
        labels_raw = [ordinal_mappings.get(col, {}).get(idx, str(idx)) for idx in counts.index]  
        values = counts.values
        labels_display = [map_digits(translate(str(x))) for x in labels_raw]

        # ---------------- Output Paths ----------------
        safe_col = re.sub(r"[^\w\-_\.]", "_", col)
        plots_dir = os.path.join(settings.MEDIA_ROOT, f"ID_{user_id}_uploads", "temporary_uploads", "plots")
        os.makedirs(plots_dir, exist_ok=True)
        final_path = os.path.join(plots_dir, f"barchart_{safe_col}.{fmt}")

        # ---------------- Plot ----------------
        fig, ax = plt.subplots(figsize=(width / 100, height / 100), dpi=dpi)

        if orientation == "vertical":
            x_idx = np.arange(len(values))
            bars = ax.bar(x_idx, values, color=bar_color, edgecolor="black", linewidth=1.0)
            ax.set_xticks(x_idx)
            ax.set_xticklabels(labels_display, fontproperties=tick_prop, rotation=40, ha="right")
            ax.set_ylabel(map_digits(translate("Count")), fontsize=label_font_size)
            ax.set_xlabel(map_digits(translate(col)), fontsize=label_font_size)

            max_val = max(values) if len(values) > 0 else 1
            ax.set_ylim(0, max_val + max_val * 0.25)  # add headroom for labels

            for bar, val in zip(bars, values):
                ax.text(
                    bar.get_x() + bar.get_width() / 2,
                    bar.get_height() + (0.05 * max_val),
                    map_digits(str(val)),
                    ha="center", va="bottom", fontsize=val_font_size
                )

        else:
            y_idx = np.arange(len(values))
            bars = ax.barh(y_idx, values, color=bar_color, edgecolor="black", linewidth=1.0)
            ax.set_yticks(y_idx)
            ax.set_yticklabels(labels_display, fontproperties=tick_prop)
            ax.set_xlabel(map_digits(translate("Count")), fontsize=label_font_size)
            ax.set_ylabel(map_digits(translate(col)), fontsize=label_font_size)

            max_val = max(values) if len(values) > 0 else 1
            ax.set_xlim(0, max_val + max_val * 0.25)  # add headroom for labels

            for bar, val in zip(bars, values):
                ax.text(
                    bar.get_width() + (0.05 * max_val),
                    bar.get_y() + bar.get_height() / 2,
                    map_digits(str(val)),
                    va="center", ha="left", fontsize=val_font_size
                )

        
        title_txt = map_digits(translate(f"Bar Chart of {col}"))
        

        ax.grid(axis="y" if orientation == "vertical" else "x", linestyle="--", alpha=0.35)
        plt.tight_layout()
        fig.savefig(final_path, dpi=dpi, bbox_inches="tight")
        plt.close(fig)

        # ---------------- Title Overlay (PIL) ----------------
        img = Image.open(final_path).convert("RGB")
        bw, bh = img.size

        # Add extra space above for title (20% of image height, adjust as needed)
        extra_height = int(bh * 0.15)
        new_img = Image.new("RGB", (bw, bh + extra_height), "white")

        # Paste original chart lower (shift down)
        new_img.paste(img, (0, extra_height))

        # Draw title in the new top space
        draw = ImageDraw.Draw(new_img)
        title_txt = map_digits(translate(f"Bar Chart of {col}"))
        tw, th = draw.textbbox((0, 0), title_txt, font=caption_font)[2:]
        draw.text(((bw - tw) // 2, (extra_height - th) // 2), title_txt, font=caption_font, fill="black")

        # Save final
        new_img.save(final_path, format=pil_fmt)

        return JsonResponse({
            "success": True,
            "test": "Bar Chart",
            "orientation": orientation,
            "image_paths": [
                f"{settings.MEDIA_URL}ID_{user_id}_uploads/temporary_uploads/plots/{os.path.basename(final_path)}"
            ]
        })

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})

def process_pie_chart(request, df, col,user_id,ordinal_mappings):
    import re 
    try:
        # --- Params ---
        language = request.POST.get('language', 'en')
        fmt = request.POST.get('format', 'png').lower()
        fmt = fmt if fmt in ('png', 'jpg', 'jpeg', 'pdf', 'tiff') else 'png'

        use_default = request.POST.get('use_default', 'true') == 'true'
        if use_default:
            dpi, width, height = 300, 900, 650
        else:
            dpi = int(request.POST.get('dpi', 300))
            width, height = map(int, request.POST.get('image_size', '900x650').split('x'))

        # --- Fonts / translator ---
        font_path = os.path.join(settings.BASE_DIR, 'NotoSansBengali-Regular.ttf')
        try:
            fm.fontManager.addfont(font_path)
        except Exception:
            pass
        if language == 'bn' and os.path.exists(font_path):
            plt.rcParams['font.family'] = fm.FontProperties(fname=font_path).get_name()

        translator = Translator()
        digit_map_bn = str.maketrans('0123456789', '০১২৩৪৫৬৭৮৯')
        def translate(text): 
            if language == 'bn':
                try:
                    return translator.translate(text, dest='bn').text
                except Exception:
                    return text
            return text
        def map_digits(s): return s.translate(digit_map_bn) if language == 'bn' else s

        # --- Validate col ---
        if col not in df.columns:
            return JsonResponse({'success': False, 'error': f"Column {col} not found"})

        # --- Treat column as categorical always ---
        series = df[col].astype(str).fillna('NaN')
        counts = df[col].value_counts() 
        # labels_data = counts.index.tolist()  
        labels_raw= [ordinal_mappings.get(col, {}).get(idx, str(idx)) for idx in counts.index]  
 
        print(labels_raw) 

        sizes = counts.values

        # --- Too many categories check ---
        if len(labels_raw) > 10:
            msg = ("শ্রেণি সংখ্যা ১০ এর বেশি। পাই চার্ট উপযুক্ত নয়। বার চার্ট ব্যবহার করুন।"
                   if language == 'bn'
                   else "Too many categories (>10). Pie charts are not suitable. Please try a bar chart instead.")
            return JsonResponse({'success': False, 'error': msg})

        # --- Colors & figure ---
        colors = ["#1f77b4","#ff7f0e","#2ca02c","#d62728","#9467bd",
                  "#8c564b","#e377c2","#7f7f7f","#bcbd22","#17becf"]

        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(width/100, height/100), dpi=dpi,
                                       gridspec_kw={'width_ratios': [2, 1]})
        total = sum(sizes)

        # autopct
        def make_autopct(values):
            def my_autopct(pct):
                text = f"{pct:.1f}%"
                return map_digits(text) if language == 'bn' else text
            return my_autopct

        wedges, texts, autotexts = ax1.pie(
            sizes, labels=None, autopct=make_autopct(sizes),
            startangle=90, colors=colors[:len(sizes)],
            wedgeprops=dict(edgecolor='black', linewidth=1.2), radius=1.0
        )
        ax1.axis('equal')

        # --- Improved small slice handling ---
        total = sum(sizes)
        small_slices = []
        for i, (wedge, autotext) in enumerate(zip(wedges, autotexts)):
            try:
                pct = float(autotext.get_text().replace('%', ''))
            except Exception:
                pct = 0.0

            if pct >= 5:
                autotext.set_color("white")
                autotext.set_weight("bold")
            else:
                ang = (wedge.theta2 + wedge.theta1) / 2.0
                small_slices.append((i, wedge, autotext, pct, ang))

        # sort by angle
        small_slices.sort(key=lambda x: x[4])
        distances = [1.3, 1.5, 1.8, 2.0]

        for idx, (i, wedge, autotext, pct, ang) in enumerate(small_slices):
            distance_idx = idx % len(distances)
            label_distance = distances[distance_idx]
            line_end_distance = label_distance - 0.1

            x = label_distance * math.cos(math.radians(ang))
            y = label_distance * math.sin(math.radians(ang))

            percentage_text = f"{pct:.1f}%"
            if language == 'bn':
                percentage_text = map_digits(percentage_text)

            autotext.set_text(percentage_text)
            autotext.set_color("black")
            autotext.set_weight("bold")
            autotext.set_position((x, y))
            autotext.set_ha('center')
            autotext.set_va('center')

            # draw leader line
            x1 = 1.0 * math.cos(math.radians(ang))
            y1 = 1.0 * math.sin(math.radians(ang))
            x2 = line_end_distance * math.cos(math.radians(ang))
            y2 = line_end_distance * math.sin(math.radians(ang))
            ax1.plot([x1, x2], [y1, y2], color="black", linewidth=1)

        # Title
        ax1.set_title(map_digits(translate(f"Pie Chart of {col}")), fontsize=16, pad=20)

        # FIXED: Legend - EXACTLY like the Python code
        # Create labels_display exactly like in Python code
        labels_display = [map_digits(translate(str(lab))) for lab in labels_raw]

        legend_elements = []
        for i, (label, size_val) in enumerate(zip(labels_display, sizes)):
            percentage = size_val/total*100
            percentage_text = f'{percentage:.1f}%'
            if language == 'bn':
                percentage_text = map_digits(percentage_text)

            legend_text = f"{label} - {percentage_text}"
            legend_elements.append(Patch(
                facecolor=colors[i % len(colors)],
                edgecolor='black', 
                label=legend_text
            ))

        ax2.legend(handles=legend_elements, loc='center', frameon=False,
                fontsize=12, handlelength=1, handleheight=1, handletextpad=0.5)
        ax2.axis('off')

        plt.tight_layout()

        # --- Safe filename ---
        def safe_filename(name):
            # Replace invalid characters with underscore and truncate long names
            cleaned = re.sub(r'[<>:"/\\|?*]', '_', name)
            return cleaned[:100]  # max 100 chars to avoid OS limits

        filename = f"pie_{safe_filename(col)}.{fmt}"
        plots_dir =  os.path.join(settings.MEDIA_ROOT, f'ID_{user_id}_uploads', 'temporary_uploads', 'plots')
        os.makedirs(plots_dir, exist_ok=True)
        out_path = os.path.join(plots_dir, filename)

        plt.savefig(out_path, dpi=dpi, bbox_inches="tight")
        plt.close(fig)

        return JsonResponse({
            'success': True,
            'image_paths': [os.path.join(settings.MEDIA_URL, 'plots', filename)],
            'message': "Pie chart generated successfully",
            'columns': [col]
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
'''


def process_ks_test(request, df: pd.DataFrame, col: str, user_id: str):
    """
    Performs Kolmogorov-Smirnov test and returns analysis results without generating plots.
    Frontend will handle visualization using the returned data.
    
    Args:
        request: Django request object
        df: DataFrame containing the data
        col: Column name for the numeric variable to test
        user_id: User identifier
        
    Returns:
        JsonResponse with test results and data for ECDF plotting
    """
    from scipy.stats import kstest, norm
    import numpy as np
    
    print(f"[KS Test] column: {col}")
    
    # Validate column name
    if col not in df.columns:
        return JsonResponse({
            'success': False, 
            'error': 'Invalid column name.'
        })

    # Extract language preference
    try:
        language = request.POST.get('language', 'en').lower()
    except Exception:
        language = 'en'
    
    if language not in ('en', 'bn'):
        language = 'en'

    # Prepare working data
    work = df[[col]].copy()
    work = work.dropna(subset=[col])
    print(f"[KS Test] after dropna: {len(work)} rows")

    # Ensure column is numeric
    if not pd.api.types.is_numeric_dtype(work[col]):
        work[col] = pd.to_numeric(work[col], errors='coerce')
        work = work.dropna(subset=[col])
        
        if not pd.api.types.is_numeric_dtype(work[col]):
            return JsonResponse({
                'success': False, 
                'error': f'"{col}" must be numeric for Kolmogorov-Smirnov test.'
            })

    data_values = work[col].values
    
    if len(data_values) < 2:
        return JsonResponse({
            'success': False, 
            'error': 'Need at least 2 observations for Kolmogorov-Smirnov test.'
        })

    # Calculate statistics for normal distribution comparison
    mean_val = float(np.mean(data_values))
    std_val = float(np.std(data_values, ddof=0))  # population std
    
    # Perform Kolmogorov-Smirnov test against normal distribution
    try:
        stat, p_value = kstest(data_values, 'norm', args=(mean_val, std_val))
        print(f"[KS Test] result: D={stat:.6f}, p={p_value:.6g}")
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'Error in Kolmogorov-Smirnov test: {e}'
        })

    # Prepare ECDF data for frontend plotting
    sorted_data = np.sort(data_values)
    n = len(sorted_data)
    ecdf_y = np.arange(1, n + 1) / n
    
    # Generate theoretical CDF (normal distribution) for plotting
    x_min = float(np.min(data_values))
    x_max = float(np.max(data_values))
    x_range = x_max - x_min
    x_theoretical = np.linspace(x_min - 0.1 * x_range, x_max + 0.1 * x_range, 200)
    cdf_theoretical = norm.cdf(x_theoretical, mean_val, std_val)

    # Prepare plot data
    plot_data = {
        'ecdf': {
            'x': [float(x) for x in sorted_data],  # Convert numpy to native Python types
            'y': [float(y) for y in ecdf_y]
        },
        'cdf': {
            'x': [float(x) for x in x_theoretical],
            'y': [float(y) for y in cdf_theoretical]
        },
        'statistics': {
            'mean': mean_val,
            'std': std_val,
            'min': float(np.min(data_values)),
            'max': float(np.max(data_values)),
            'sample_size': int(n)
        }
    }

    # Prepare response with localized labels
    test_name = 'Kolmogorov-Smirnov Test' if language == 'en' else 'কোলমোগোরভ-স্মিরনভ পরীক্ষা'
    
    # Localized interpretations
    if language == 'bn':
        interpretation = "স্বাভাবিক বন্টনের সাথে সামঞ্জস্যপূর্ণ (p > 0.05)" if p_value > 0.05 else "স্বাভাবিক বন্টন নয় (p ≤ 0.05)"
        conclusion = "ডেটা স্বাভাবিক বন্টন অনুসরণ করে" if p_value > 0.05 else "ডেটা স্বাভাবিক বন্টন অনুসরণ করে না"
    else:
        interpretation = "Consistent with Normal (p > 0.05)" if p_value > 0.05 else "Not Normal (p ≤ 0.05)"
        conclusion = "Data follows normal distribution" if p_value > 0.05 else "Data does not follow normal distribution"

    response_data = {
        'success': True,
        'test': test_name,
        'language': language,
        'statistic': float(stat),  # KS statistic (D)
        'p_value': float(p_value),
        'interpretation': interpretation,
        'conclusion': conclusion,
        'column_names': {
            'variable': str(col)
        },
        'plot_data': plot_data,
        'distribution_parameters': {
            'mean': mean_val,
            'std_dev': std_val,
            'tested_distribution': 'normal'
        },
        'metadata': {
            'sample_size': int(n),
            'significant': bool(p_value < 0.05),
            'alpha': 0.05,
            'test_type': 'goodness_of_fit'
        }
    }
    
    return JsonResponse(response_data)


def process_anderson_darling_test(request, df: pd.DataFrame, col: str, user_id: str):
    """
    Performs Anderson-Darling test for normality and returns analysis results without generating plots.
    Frontend will handle visualization using the returned data.
    
    Args:
        request: Django request object
        df: DataFrame containing the data
        col: Column name for the variable to test (numeric)
        user_id: User identifier
        
    Returns:
        JsonResponse with test results and data for Q-Q plot
    """

    from scipy.stats import anderson, norm
    import numpy as np

    print(f"[Anderson-Darling] Testing column: {col}")
    
    # Validate column name
    if col not in df.columns:
        return JsonResponse({
            'success': False, 
            'error': 'Invalid column name.'
        })

    # Extract language preference
    try:
        language = request.POST.get('language', 'en').lower()
    except Exception:
        language = 'en'
    
    if language not in ('en', 'bn'):
        language = 'en'

    # Prepare working data
    work = df[[col]].copy()
    work = work.dropna(subset=[col])
    print(f"[Anderson-Darling] after dropna: {len(work)} rows")

    # Ensure column is numeric
    if not pd.api.types.is_numeric_dtype(work[col]):
        work[col] = pd.to_numeric(work[col], errors='coerce')
        work = work.dropna(subset=[col])
        
        if not pd.api.types.is_numeric_dtype(work[col]):
            return JsonResponse({
                'success': False, 
                'error': f'"{col}" must be numeric for Anderson-Darling test.'
            })

    data = work[col].values
    
    if len(data) < 3:
        return JsonResponse({
            'success': False, 
            'error': 'Need at least 3 observations for Anderson-Darling test.'
        })

    # Perform Anderson-Darling test
    try:
        result = anderson(data, dist='norm')
        stat = result.statistic
        critical_values = result.critical_values
        significance_levels = result.significance_level
        
        print(f"[Anderson-Darling] result: A²={stat:.6f}")
        print(f"[Anderson-Darling] critical values: {critical_values}")
        print(f"[Anderson-Darling] significance levels: {significance_levels}")
        
        # Find the critical value at 5% significance level
        try:
            crit_idx = list(significance_levels).index(5.0)
            crit_value = critical_values[crit_idx]
            sig_level = significance_levels[crit_idx]
        except ValueError:
            # Use the first available critical value if 5% not found
            crit_value = critical_values[0]
            sig_level = significance_levels[0]
            
        is_normal = stat < crit_value
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'Error in Anderson-Darling test: {e}'
        })

    # Prepare Q-Q plot data for frontend
    n = len(data)
    prob = (np.arange(1, n + 1) - 0.5) / n
    sample_quantiles = np.sort(data)
    
    # Theoretical quantiles from normal distribution
    data_mean = np.mean(data)
    data_std = np.std(data, ddof=0)
    theoretical_quantiles = norm.ppf(prob, loc=data_mean, scale=data_std)
    
    # Reference line (y = x line for perfect normality)
    ref_min = min(theoretical_quantiles.min(), sample_quantiles.min())
    ref_max = max(theoretical_quantiles.max(), sample_quantiles.max())
    reference_line = [ref_min, ref_max]

    # Prepare response with localized labels
    test_name = 'Anderson-Darling Test' if language == 'en' else 'অ্যান্ডারসন-ডার্লিং পরীক্ষা'
    
    if language == 'bn':
        interpretation = (
            f"সম্ভাব্য স্বাভাবিক বণ্টন (A² = {stat:.3f} < {crit_value:.3f} {sig_level}% স্তরে)"
            if is_normal else
            f"স্বাভাবিক বণ্টন নয় (A² = {stat:.3f} ≥ {crit_value:.3f} {sig_level}% স্তরে)"
        )
    else:
        interpretation = (
            f"Likely Normal (A² = {stat:.3f} < {crit_value:.3f} at {sig_level}% level)"
            if is_normal else
            f"Not Normal (A² = {stat:.3f} ≥ {crit_value:.3f} at {sig_level}% level)"
        )

    response_data = {
        'success': True,
        'test': test_name,
        'language': language,
        'statistic': float(stat),
        'critical_value': float(crit_value),
        'significance_level': float(sig_level),
        'is_normal': bool(is_normal),
        'interpretation': interpretation,
        'sample_size': int(n),
        'column_name': str(col),
        'descriptive_stats': {
            'mean': float(data_mean),
            'std': float(data_std),
            'min': float(np.min(data)),
            'max': float(np.max(data)),
            'median': float(np.median(data))
        },
        'qq_plot_data': {
            'theoretical_quantiles': [float(x) for x in theoretical_quantiles],
            'sample_quantiles': [float(x) for x in sample_quantiles],
            'reference_line': [float(x) for x in reference_line]
        },
        'critical_values': {
            'values': [float(x) for x in critical_values],
            'significance_levels': [float(x) for x in significance_levels]
        },
        'metadata': {
            'test_type': 'normality_test',
            'distribution': 'normal',
            'alpha': 0.05
        }
    }
    
    return JsonResponse(response_data)


def process_f_test(request, df: pd.DataFrame, col_group: str, col_value: str, user_id: str):
    """
    Performs F-test for equality of variances - FIXED DATA STRUCTURE
    """
    print(f"[F-Test] cols: {col_group}(group) | {col_value}(value)")
    
    # Validate column names
    if col_group not in df.columns or col_value not in df.columns:
        return JsonResponse({
            'success': False, 
            'error': 'Invalid column names.'
        })

    # Extract language preference
    try:
        language = request.POST.get('language', 'en').lower()
    except Exception:
        language = 'en'
    
    if language not in ('en', 'bn'):
        language = 'en'

    # Prepare working dataset
    work = df[[col_group, col_value]].copy()
    work = work.dropna(subset=[col_group, col_value])
    print(f"[F-Test] after dropna: {len(work)} rows")

    # Ensure grouping column is categorical
    if not pd.api.types.is_categorical_dtype(work[col_group]):
        work[col_group] = work[col_group].astype('category')

    categories = list(work[col_group].cat.categories)
    
    if len(categories) != 2:
        return JsonResponse({
            'success': False, 
            'error': 'F-test requires exactly 2 groups.'
        })

    # Ensure value column is numeric
    if not pd.api.types.is_numeric_dtype(work[col_value]):
        work[col_value] = pd.to_numeric(work[col_value], errors='coerce')
        work = work.dropna(subset=[col_value])
        
        if not pd.api.types.is_numeric_dtype(work[col_value]):
            return JsonResponse({
                'success': False, 
                'error': f'"{col_value}" must be numeric for F-test.'
            })

    # Prepare groups for statistical test
    group1_data = work.loc[work[col_group] == categories[0], col_value].values
    group2_data = work.loc[work[col_group] == categories[1], col_value].values
    
    if len(group1_data) < 2 or len(group2_data) < 2:
        return JsonResponse({
            'success': False, 
            'error': 'Each group must contain at least 2 observations for F-test.'
        })

    # Perform F-test for equality of variances
    try:
        # Calculate variances
        var1 = np.var(group1_data, ddof=1)
        var2 = np.var(group2_data, ddof=1)
        
        # F-statistic (larger variance / smaller variance)
        if var1 >= var2:
            F_stat = var1 / var2
            dfn = len(group1_data) - 1
            dfd = len(group2_data) - 1
        else:
            F_stat = var2 / var1
            dfn = len(group2_data) - 1
            dfd = len(group1_data) - 1
        
        # Two-tailed p-value
        p_value = 2 * min(stats.f.cdf(F_stat, dfn, dfd), 1 - stats.f.cdf(F_stat, dfn, dfd))
        
        print(f"[F-Test] result: F={F_stat:.6f}, p={p_value:.6g}, df=({dfn}, {dfd})")
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'Error in F-test: {e}'
        })

    # Prepare plot data - CRITICAL: This is what frontend needs
    plot_data = []
    for i, category in enumerate(categories):
        category_data = group1_data if i == 0 else group2_data
        
        plot_data.append({
            'category': str(category),
            'count': int(len(category_data)),
            'mean': float(np.mean(category_data)),
            'median': float(np.median(category_data)),
            'std': float(np.std(category_data)),
            'min': float(np.min(category_data)),
            'max': float(np.max(category_data)),
            'q25': float(np.percentile(category_data, 25)),
            'q75': float(np.percentile(category_data, 75))
        })

    # Prepare variances for response
    variances = {
        str(categories[0]): float(np.var(group1_data, ddof=1)),
        str(categories[1]): float(np.var(group2_data, ddof=1))
    }

    # Response structure that matches frontend expectations
    response_data = {
        'success': True,
        'test': 'F-Test for Equality of Variances',
        'language': language,
        'statistic': float(F_stat),
        'p_value': float(p_value),
        'degrees_of_freedom': f"{dfn}, {dfd}",
        'n_groups': int(len(categories)),
        'total_observations': int(len(work)),
        'variances': variances,  # Required for table
        'column_names': {
            'group': str(col_group),
            'value': str(col_value)
        },
        'plot_data': plot_data,  # REQUIRED - frontend checks this
        'metadata': {
            'significant': bool(p_value < 0.05),
            'test_type': 'f_test'
        }
    }

    print(f"[F-Test] Returning {len(plot_data)} plot data items")
    return JsonResponse(response_data)


def process_z_test(request, df: pd.DataFrame, col_group: str, col_value: str, user_id: str):
    
    """
    Performs Z-test for equality of means and returns analysis results without generating plots.
    Frontend will handle visualization using the returned data.
    """
    print(f"[Z-Test] cols: {col_group}(group) | {col_value}(value)")
    
    # Validate column names
    if col_group not in df.columns or col_value not in df.columns:
        return JsonResponse({
            'success': False, 
            'error': 'Invalid column names.'
        })

    # Extract language preference
    try:
        language = request.POST.get('language', 'en').lower()
    except Exception:
        language = 'en'
    
    if language not in ('en', 'bn'):
        language = 'en'

    # Prepare working dataset
    work = df[[col_group, col_value]].copy()
    work = work.dropna(subset=[col_group, col_value])
    print(f"[Z-Test] after dropna: {len(work)} rows")

    # Ensure grouping column is categorical
    if not pd.api.types.is_categorical_dtype(work[col_group]):
        work[col_group] = work[col_group].astype('category')

    categories = list(work[col_group].cat.categories)
    
    if len(categories) != 2:
        return JsonResponse({
            'success': False, 
            'error': 'Z-test requires exactly 2 groups.'
        })

    # Ensure value column is numeric
    if not pd.api.types.is_numeric_dtype(work[col_value]):
        work[col_value] = pd.to_numeric(work[col_value], errors='coerce')
        work = work.dropna(subset=[col_value])
        
        if not pd.api.types.is_numeric_dtype(work[col_value]):
            return JsonResponse({
                'success': False, 
                'error': f'"{col_value}" must be numeric for Z-test.'
            })

    # Prepare groups for statistical test
    group1_data = work.loc[work[col_group] == categories[0], col_value].values
    group2_data = work.loc[work[col_group] == categories[1], col_value].values
    
    if len(group1_data) < 2 or len(group2_data) < 2:
        return JsonResponse({
            'success': False, 
            'error': 'Each group must contain at least 2 observations for Z-test.'
        })

    # Perform Z-test for equality of means
    try:
        from statsmodels.stats.weightstats import ztest
        z_stat, p_value = ztest(group1_data, group2_data)
        
        print(f"[Z-Test] result: Z={z_stat:.6f}, p={p_value:.6g}")
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'Error in Z-test: {e}'
        })

    # Prepare data for frontend plotting
    plot_data = []
    
    for i, category in enumerate(categories):
        if i == 0:
            category_data = group1_data
        else:
            category_data = group2_data
            
        plot_data.append({
            'category': str(category),
            'values': [float(x) for x in category_data],
            'count': int(len(category_data)),
            'mean': float(np.mean(category_data)),
            'median': float(np.median(category_data)),
            'std': float(np.std(category_data)),
            'variance': float(np.var(category_data, ddof=1)),
            'min': float(np.min(category_data)),
            'max': float(np.max(category_data)),
            'q25': float(np.percentile(category_data, 25)),
            'q75': float(np.percentile(category_data, 75))
        })

    # Prepare response with localized labels
    test_name = 'Z-Test for Equality of Means' if language == 'en' else 'জেড-টেস্ট (মানে সমতা)'
    
    response_data = {
        'success': True,
        'test': test_name,
        'language': language,
        'statistic': float(z_stat),
        'p_value': float(p_value),
        'n_groups': int(len(categories)),
        'total_observations': int(len(work)),
        'means': {
            str(categories[0]): float(np.mean(group1_data)),
            str(categories[1]): float(np.mean(group2_data))
        },
        'column_names': {
            'group': str(col_group),
            'value': str(col_value)
        },
        'plot_data': plot_data,
        'metadata': {
            'categories': [str(c) for c in categories],
            'significant': bool(p_value < 0.05),
            'alpha': 0.05,
            'test_type': 'z_test'
        }
    }
    return JsonResponse(response_data)


def process_t_test(request, df: pd.DataFrame, col_group: str, col_value: str, user_id: str):
    """
    Performs Welch's T-test for equality of means and returns analysis results without generating plots.
    Frontend will handle visualization using the returned data.
    """
    print(f"[T-Test] cols: {col_group}(group) | {col_value}(value)")
    
    # Validate column names
    if col_group not in df.columns or col_value not in df.columns:
        return JsonResponse({
            'success': False, 
            'error': 'Invalid column names.'
        })

    # Extract language preference
    try:
        language = request.POST.get('language', 'en').lower()
    except Exception:
        language = 'en'
    
    if language not in ('en', 'bn'):
        language = 'en'

    # Prepare working dataset
    work = df[[col_group, col_value]].copy()
    work = work.dropna(subset=[col_group, col_value])
    print(f"[T-Test] after dropna: {len(work)} rows")

    # Ensure grouping column is categorical
    if not pd.api.types.is_categorical_dtype(work[col_group]):
        work[col_group] = work[col_group].astype('category')

    categories = list(work[col_group].cat.categories)
    
    if len(categories) != 2:
        return JsonResponse({
            'success': False, 
            'error': 'T-test requires exactly 2 groups.'
        })

    # Ensure value column is numeric
    if not pd.api.types.is_numeric_dtype(work[col_value]):
        work[col_value] = pd.to_numeric(work[col_value], errors='coerce')
        work = work.dropna(subset=[col_value])
        
        if not pd.api.types.is_numeric_dtype(work[col_value]):
            return JsonResponse({
                'success': False, 
                'error': f'"{col_value}" must be numeric for T-test.'
            })

    # Prepare groups for statistical test
    group1_data = work.loc[work[col_group] == categories[0], col_value].values
    group2_data = work.loc[work[col_group] == categories[1], col_value].values
    
    if len(group1_data) < 2 or len(group2_data) < 2:
        return JsonResponse({
            'success': False, 
            'error': 'Each group must contain at least 2 observations for T-test.'
        })

    # Perform Welch's T-test (does not assume equal variances)
    try:
        t_stat, p_value = stats.ttest_ind(group1_data, group2_data, equal_var=False)
        
        # Calculate Welch-Satterthwaite degrees of freedom
        v1 = np.var(group1_data, ddof=1)
        v2 = np.var(group2_data, ddof=1)
        n1 = len(group1_data)
        n2 = len(group2_data)
        welch_df = (v1/n1 + v2/n2)**2 / ((v1/n1)**2/(n1-1) + (v2/n2)**2/(n2-1))
        
        print(f"[T-Test] result: t={t_stat:.6f}, p={p_value:.6g}, df={welch_df:.2f}")
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'Error in T-test: {e}'
        })

    # Prepare data for frontend plotting
    plot_data = []
    
    for i, category in enumerate(categories):
        if i == 0:
            category_data = group1_data
        else:
            category_data = group2_data
            
        plot_data.append({
            'category': str(category),
            'values': [float(x) for x in category_data],
            'count': int(len(category_data)),
            'mean': float(np.mean(category_data)),
            'median': float(np.median(category_data)),
            'std': float(np.std(category_data)),
            'variance': float(np.var(category_data, ddof=1)),
            'min': float(np.min(category_data)),
            'max': float(np.max(category_data)),
            'q25': float(np.percentile(category_data, 25)),
            'q75': float(np.percentile(category_data, 75))
        })

    # Prepare response with localized labels
    test_name = "Welch's T-Test for Equality of Means" if language == 'en' else 'টি-টেস্ট (মানে সমতা)'
    
    response_data = {
        'success': True,
        'test': test_name,
        'language': language,
        'statistic': float(t_stat),
        'p_value': float(p_value),
        'degrees_of_freedom': float(welch_df),
        'n_groups': int(len(categories)),
        'total_observations': int(len(work)),
        'means': {
            str(categories[0]): float(np.mean(group1_data)),
            str(categories[1]): float(np.mean(group2_data))
        },
        'variances': {
            str(categories[0]): float(np.var(group1_data, ddof=1)),
            str(categories[1]): float(np.var(group2_data, ddof=1))
        },
        'column_names': {
            'group': str(col_group),
            'value': str(col_value)
        },
        'plot_data': plot_data,
        'metadata': {
            'categories': [str(c) for c in categories],
            'significant': bool(p_value < 0.05),
            'alpha': 0.05,
            'test_type': 't_test'
        }
    }
    return JsonResponse(response_data)


def process_fzt_visualization(request, df, col_group, col_value, user_id):
    """
    Returns ALL THREE test results + combined visualization data
    """
    try:
        # Extract language preference
        language = request.POST.get('language', 'en').lower()
        if language not in ('en', 'bn'):
            language = 'en'

        # Extract plot parameters for histogram and KDE
        histogram_bins = request.POST.get('histogram_bins', 'auto')
        kde_bandwidth = request.POST.get('kde_bandwidth', None)
        
        # Parse histogram bins
        if histogram_bins != 'auto' and histogram_bins != '':
            try:
                histogram_bins = int(histogram_bins)
                if histogram_bins <= 0:
                    histogram_bins = 'auto'
            except (ValueError, TypeError):
                histogram_bins = 'auto'
        
        # Parse KDE bandwidth
        if kde_bandwidth and kde_bandwidth not in ['auto', '']:
            try:
                kde_bandwidth = float(kde_bandwidth)
                if kde_bandwidth <= 0:
                    kde_bandwidth = None
            except (ValueError, TypeError):
                kde_bandwidth = None

        # Prepare working dataset
        work = df[[col_group, col_value]].copy()
        work = work.dropna(subset=[col_group, col_value])
        
        # Ensure grouping column is categorical and value is numeric
        if not pd.api.types.is_categorical_dtype(work[col_group]):
            work[col_group] = work[col_group].astype('category')
            
        if not pd.api.types.is_numeric_dtype(work[col_value]):
            work[col_value] = pd.to_numeric(work[col_value], errors='coerce')
            work = work.dropna(subset=[col_value])

        categories = list(work[col_group].cat.categories)
        
        if len(categories) != 2:
            return JsonResponse({
                'success': False, 
                'error': 'F/Z/T analysis requires exactly 2 groups.'
            })

        # Prepare groups
        group1_data = work.loc[work[col_group] == categories[0], col_value].values
        group2_data = work.loc[work[col_group] == categories[1], col_value].values

        # Perform all three tests using existing logic
        # F-Test
        var1 = np.var(group1_data, ddof=1)
        var2 = np.var(group2_data, ddof=1)
        if var1 >= var2:
            F_stat = var1 / var2
            dfn = len(group1_data) - 1
            dfd = len(group2_data) - 1
        else:
            F_stat = var2 / var1
            dfn = len(group2_data) - 1
            dfd = len(group1_data) - 1
        pF = 2 * min(stats.f.cdf(F_stat, dfn, dfd), 1 - stats.f.cdf(F_stat, dfn, dfd))

        # Z-Test
        from statsmodels.stats.weightstats import ztest
        z_stat, pZ = ztest(group1_data, group2_data)

        # T-Test
        t_stat, pT = stats.ttest_ind(group1_data, group2_data, equal_var=False)
        v1 = np.var(group1_data, ddof=1)
        v2 = np.var(group2_data, ddof=1)
        n1, n2 = len(group1_data), len(group2_data)
        welch_df = (v1/n1 + v2/n2)**2 / ((v1/n1)**2/(n1-1) + (v2/n2)**2/(n2-1))

        # Prepare plot data for each group (similar to distribution plot)
        plot_data = []
        for i, category in enumerate(categories):
            category_data = group1_data if i == 0 else group2_data
            plot_data.append({
                'category': str(category),
                'values': [float(x) for x in category_data],
                'count': int(len(category_data)),
                'mean': float(np.mean(category_data)),
                'median': float(np.median(category_data)),
                'std': float(np.std(category_data)),
                'variance': float(np.var(category_data, ddof=1)),
                'min': float(np.min(category_data)),
                'max': float(np.max(category_data)),
                'q25': float(np.percentile(category_data, 25)),
                'q75': float(np.percentile(category_data, 75))
            })

        # Generate distribution plot data for each group (like EDA distribution)
        distribution_data = {}
        
        for i, category in enumerate(categories):
            category_data = group1_data if i == 0 else group2_data
            
            # Generate histogram data
            hist_counts, bin_edges = np.histogram(category_data, bins=histogram_bins)
            bin_centers = (bin_edges[:-1] + bin_edges[1:]) / 2
            bin_widths = np.diff(bin_edges)
            
            histogram_bins_data = []
            for j in range(len(hist_counts)):
                histogram_bins_data.append({
                    'x': float(bin_centers[j]),
                    'y': int(hist_counts[j]),
                    'x0': float(bin_edges[j]),
                    'x1': float(bin_edges[j + 1]),
                    'width': float(bin_widths[j]),
                    'count': int(hist_counts[j])
                })

            hist_data = {
                'bins': histogram_bins_data,
                'total_count': int(np.sum(hist_counts)),
                'bin_count': len(hist_counts),
                'range': [float(bin_edges[0]), float(bin_edges[-1])],
                'bins_used': histogram_bins if isinstance(histogram_bins, int) else 'auto'
            }

            # Generate KDE data
            kde = stats.gaussian_kde(category_data)
            
            # Set custom bandwidth if provided
            if kde_bandwidth is not None:
                kde.set_bandwidth(kde_bandwidth)
            
            data_min = np.min(category_data)
            data_max = np.max(category_data)
            data_range = data_max - data_min
            x_points = np.linspace(data_min - 0.1 * data_range, data_max + 0.1 * data_range, 200)
            y_points = kde(x_points)
            
            # Normalize to make it comparable to histogram (like seaborn does)
            y_points_normalized = y_points * len(category_data) * (x_points[1] - x_points[0])
            
            kde_curve = []
            for j in range(len(x_points)):
                kde_curve.append({
                    'x': float(x_points[j]),
                    'y': float(y_points_normalized[j]),
                    'density': float(y_points[j])
                })

            kde_data = {
                'curve': kde_curve,
                'bandwidth': float(kde.factor),
                'range': [float(x_points[0]), float(x_points[-1])],
                'peak_density': float(np.max(y_points))
            }

            # Store distribution data for this group
            distribution_data[f'group{i+1}'] = {
                'histogram': hist_data,
                'kde': kde_data,
                'category': str(category)
            }

        return JsonResponse({
            'success': True,
            'test': 'F/Z/T Combined Analysis',
            'language': language,
            'f_test': {
                'statistic': float(F_stat),
                'p_value': float(pF),
                'degrees_of_freedom': f"{dfn}, {dfd}"
            },
            'z_test': {
                'statistic': float(z_stat),
                'p_value': float(pZ)
            },
            't_test': {
                'statistic': float(t_stat),
                'p_value': float(pT),
                'degrees_of_freedom': float(welch_df)
            },
            'plot_data': plot_data,
            'distribution_data': distribution_data,  # Changed from combined_histogram_data
            'column_names': {
                'group': str(col_group),
                'value': str(col_value)
            },
            'n_groups': int(len(categories)),
            'total_observations': int(len(work)),
            'metadata': {
                'categories': [str(c) for c in categories],
                'test_type': 'fzt_visualization',
                'has_all_tests': True,
                'histogram_bins_used': histogram_bins if isinstance(histogram_bins, int) else 'auto'
            }
        })

    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'Error in F/Z/T analysis: {str(e)}'
        })



def process_pearson_test(request, df, selected_columns, user_id):
    from scipy.stats import pearsonr
    import numpy as np
    import pandas as pd
    from django.http import JsonResponse

    try:
        # ── 1) Extract inputs ──────────────────────────────────────────────
        lang = (request.POST.get("language", "en") or "en").lower()
        if lang not in ('en', 'bn'):
            lang = 'en'
        
        is_bn = (lang == "bn")

        # Variables to analyze
        vars_list = selected_columns or []
        if len(vars_list) < 2:
            return JsonResponse({
                'success': False, 
                'error': 'Please select at least two numeric variables.'
            })

        print(f"[Pearson] Analyzing {len(vars_list)} variables: {vars_list}")

        # ── 2) Core helper functions ───────────────────────────────────────
        def fnum(x):
            """Convert to float, handling None and NaN"""
            if x is None or (isinstance(x, float) and np.isnan(x)):
                return None
            try:
                return float(x)
            except Exception:
                return None

        def prepare_scatter_plot_data(var1, var2, data):
            """Prepare scatter plot data for a pair of variables"""
            # Ensure numeric data
            s1 = pd.to_numeric(data[var1], errors='coerce')
            s2 = pd.to_numeric(data[var2], errors='coerce')
            
            # Remove pairs where either is NaN
            mask = s1.notna() & s2.notna()
            s1_clean, s2_clean = s1[mask], s2[mask]
            
            if len(s1_clean) < 2:
                return None
            
            # Calculate regression line
            try:
                slope, intercept, r_value, p_value_reg, std_err = stats.linregress(s1_clean, s2_clean)
                regression_data = {
                    'slope': float(slope),
                    'intercept': float(intercept),
                    'r_squared': float(r_value**2),
                    'p_value': float(p_value_reg)
                }
            except Exception as e:
                print(f"[Pearson] Regression calculation warning: {e}")
                regression_data = {
                    'slope': 0,
                    'intercept': 0,
                    'r_squared': 0,
                    'p_value': 1.0
                }
            
            # Calculate basic statistics for each variable
            return {
                'variable1': str(var1),
                'variable2': str(var2),
                'sample1': {
                    'name': str(var1),
                    'values': [float(x) for x in s1_clean],
                    'count': int(len(s1_clean)),
                    'mean': float(np.mean(s1_clean)),
                    'median': float(np.median(s1_clean)),
                    'std': float(np.std(s1_clean)),
                    'min': float(np.min(s1_clean)),
                    'max': float(np.max(s1_clean))
                },
                'sample2': {
                    'name': str(var2),
                    'values': [float(x) for x in s2_clean],
                    'count': int(len(s2_clean)),
                    'mean': float(np.mean(s2_clean)),
                    'median': float(np.median(s2_clean)),
                    'std': float(np.std(s2_clean)),
                    'min': float(np.min(s2_clean)),
                    'max': float(np.max(s2_clean))
                },
                'regression': regression_data
            }

        def pearson_one_pair(var_a, var_b):
            """
            Perform Pearson correlation for one pair of variables.
            Returns dict with test results.
            """
            # Ensure numeric data
            s1 = pd.to_numeric(df[var_a], errors='coerce')
            s2 = pd.to_numeric(df[var_b], errors='coerce')
            
            # Remove pairs where either is NaN
            mask = s1.notna() & s2.notna()
            s1_clean, s2_clean = s1[mask], s2[mask]
            
            n = len(s1_clean)
            
            if n < 2:
                return {
                    'n': 0,
                    'correlation': None,
                    'p_value': None,
                    'ci_lower': None,
                    'ci_upper': None
                }
            
            try:
                # Calculate Pearson correlation
                corr, p_value = pearsonr(s1_clean, s2_clean)
                
                # Calculate 95% confidence interval using Fisher z-transform
                z = np.arctanh(corr)
                se = 1 / np.sqrt(n - 3)
                z_lower = z - 1.96 * se
                z_upper = z + 1.96 * se
                ci_lower = np.tanh(z_lower)
                ci_upper = np.tanh(z_upper)
                
                return {
                    'n': n,
                    'correlation': float(corr),
                    'p_value': float(p_value),
                    'ci_lower': float(ci_lower),
                    'ci_upper': float(ci_upper)
                }
                
            except Exception as e:
                print(f"[Pearson] Error in pearsonr: {e}")
                return {
                    'n': n,
                    'correlation': None,
                    'p_value': None,
                    'ci_lower': None,
                    'ci_upper': None
                }

        def fdr_bh(pvals):
            """Benjamini-Hochberg FDR correction"""
            p = np.asarray(pvals, float)
            n = p.size
            if n == 0:
                return np.array([])
            order = np.argsort(p)
            ranked = p[order]
            adj = np.empty(n, float)
            cm = 1.0
            for i in range(n-1, -1, -1):
                cm = min(cm, ranked[i] * n / (i+1))
                adj[i] = cm
            out = np.empty(n, float)
            out[order] = np.minimum(adj, 1.0)
            return out

        # ── 3) Compute all pairwise tests ──────────────────────────────────
        n_vars = len(vars_list)
        pairwise_results = []
        all_p_values = []
        scatter_plot_data = []  # NEW: Store scatter plot data for each pair
        
        for i in range(n_vars):
            for j in range(i + 1, n_vars):
                var1 = vars_list[i]
                var2 = vars_list[j]
                
                result = pearson_one_pair(var1, var2)
                
                pairwise_results.append({
                    'variable1': str(var1),
                    'variable2': str(var2),
                    'correlation': fnum(result['correlation']),
                    'p_value': fnum(result['p_value']),
                    'ci_lower': fnum(result['ci_lower']),
                    'ci_upper': fnum(result['ci_upper']),
                    'n': result['n']
                })
                
                if result['p_value'] is not None:
                    all_p_values.append(result['p_value'])
                else:
                    all_p_values.append(np.nan)
                
                # NEW: Prepare scatter plot data for this pair
                scatter_data = prepare_scatter_plot_data(var1, var2, df)
                if scatter_data:
                    scatter_plot_data.append(scatter_data)

        print(f"[Pearson] Completed {len(pairwise_results)} pairwise tests")
        print(f"[Pearson] Prepared {len(scatter_plot_data)} scatter plots")

        # ── 4) Apply FDR correction ────────────────────────────────────────
        p_array = np.array(all_p_values, dtype=float)
        mask = ~np.isnan(p_array)
        
        if mask.any():
            adjusted_p = np.full(len(pairwise_results), np.nan, dtype=float)
            adjusted_p[mask] = fdr_bh(p_array[mask])
            
            for i, result in enumerate(pairwise_results):
                result['p_adjusted'] = float(adjusted_p[i]) if not np.isnan(adjusted_p[i]) else None
        else:
            for result in pairwise_results:
                result['p_adjusted'] = None

        # Sort by absolute correlation (descending)
        pairwise_results.sort(
            key=lambda r: (r['correlation'] is None, -abs(r['correlation']) if r['correlation'] is not None else 0)
        )

        # ── 5) Create correlation matrix for heatmap ───────────────────────
        corr_matrix = np.ones((n_vars, n_vars), dtype=float)
        p_matrix = np.ones((n_vars, n_vars), dtype=float)
        p_adjusted_matrix = np.ones((n_vars, n_vars), dtype=float)
        
        for result in pairwise_results:
            i = vars_list.index(result['variable1'])
            j = vars_list.index(result['variable2'])
            
            corr_val = result['correlation'] if result['correlation'] is not None else 0.0
            corr_matrix[i, j] = corr_val
            corr_matrix[j, i] = corr_val
            
            p_val = result['p_value'] if result['p_value'] is not None else 1.0
            p_matrix[i, j] = p_val
            p_matrix[j, i] = p_val
            
            p_adj_val = result['p_adjusted'] if result['p_adjusted'] is not None else 1.0
            p_adjusted_matrix[i, j] = p_adj_val
            p_adjusted_matrix[j, i] = p_adj_val
        
        np.fill_diagonal(corr_matrix, 1.0)
        np.fill_diagonal(p_matrix, 1.0)
        np.fill_diagonal(p_adjusted_matrix, 1.0)

        # ── 6) Prepare variable-level statistics ──────────────────────────
        variable_stats = []
        
        for var in vars_list:
            col_data = pd.to_numeric(df[var], errors='coerce').dropna()
            
            variable_stats.append({
                'variable': str(var),
                'n_observations': int(len(col_data)),
                'n_missing': int(df[var].isna().sum()),
                'mean': float(col_data.mean()) if len(col_data) > 0 else None,
                'std': float(col_data.std()) if len(col_data) > 0 else None,
                'min': float(col_data.min()) if len(col_data) > 0 else None,
                'max': float(col_data.max()) if len(col_data) > 0 else None,
                'median': float(col_data.median()) if len(col_data) > 0 else None
            })

        # ── 7) Create blocks (one anchor variable at a time) ──────────────
        blocks = []
        
        for anchor in vars_list:
            block_results = []
            
            for result in pairwise_results:
                if result['variable1'] == anchor:
                    block_results.append(result)
                elif result['variable2'] == anchor:
                    # Swap variables so anchor is always first
                    swapped_result = {
                        'variable1': result['variable2'],
                        'variable2': result['variable1'],
                        'correlation': result['correlation'],
                        'p_value': result['p_value'],
                        'p_adjusted': result['p_adjusted'],
                        'ci_lower': result['ci_lower'],
                        'ci_upper': result['ci_upper'],
                        'n': result['n']
                    }
                    block_results.append(swapped_result)
            
            blocks.append({
                'anchor': str(anchor),
                'results': block_results
            })

        # ── 8) Prepare table columns for frontend ─────────────────────────
        if is_bn:
            table_columns = [
                {"key": "variable1", "label": "ভেরিয়েবল ১"},
                {"key": "variable2", "label": "ভেরিয়েবল ২"},
                {"key": "correlation", "label": "পিয়ারসন সম্পর্ক"},
                {"key": "p_value", "label": "পি-মান"},
                {"key": "p_adjusted", "label": "সমন্বিত পি-মান"},
                {"key": "ci_lower", "label": "৯৫% সিআই নিম্ন"},
                {"key": "ci_upper", "label": "৯৫% সিআই উচ্চ"},
                {"key": "n", "label": "নমুনা"},
            ]
        else:
            table_columns = [
                {"key": "variable1", "label": "Variable 1"},
                {"key": "variable2", "label": "Variable 2"},
                {"key": "correlation", "label": "Pearson Correlation"},
                {"key": "p_value", "label": "P-value"},
                {"key": "p_adjusted", "label": "Adjusted P-value"},
                {"key": "ci_lower", "label": "95% CI Lower"},
                {"key": "ci_upper", "label": "95% CI Upper"},
                {"key": "n", "label": "N"},
            ]

        # ── 9) Prepare plot data for visualizations ───────────────────────
        plot_data = []
        for i, var1 in enumerate(vars_list):
            for j, var2 in enumerate(vars_list):
                if i != j:
                    # Find the corresponding result
                    result = None
                    for r in pairwise_results:
                        if (r['variable1'] == var1 and r['variable2'] == var2) or \
                           (r['variable1'] == var2 and r['variable2'] == var1):
                            result = r
                            break
                    
                    if result:
                        plot_data.append({
                            'category': f"{var1}",
                            'variable': f"{var2}",
                            'p_value': result['p_value'],
                            'correlation': result['correlation'],
                            'p_adjusted': result['p_adjusted']
                        })

        # ── 10) Prepare final response ─────────────────────────────────────
        test_name = 'Pearson Correlation Test' if lang == 'en' else 'পিয়ারসন সম্পর্ক পরীক্ষা'
        
        response_data = {
            'success': True,
            'test': test_name,
            'language': lang,
            'n_variables': int(n_vars),
            'n_comparisons': int(len(pairwise_results)),
            'variables': [str(v) for v in vars_list],
            'variable_stats': variable_stats,
            'pairwise_results': pairwise_results,
            'blocks': blocks,
            'scatter_plot_data': scatter_plot_data,  # NEW: Add scatter plot data
            'selected_scatter_pair': None,  # NEW: Currently selected scatter plot pair
            'correlation_matrix': [[float(val) for val in row] for row in corr_matrix.tolist()],
            'p_value_matrix': [[float(val) for val in row] for row in p_matrix.tolist()],
            'p_adjusted_matrix': [[float(val) for val in row] for row in p_adjusted_matrix.tolist()],
            'plot_data': plot_data,
            'table_columns': table_columns,
            'metadata': {
                'fdr_correction': 'Benjamini-Hochberg',
                'alpha': 0.05,
                'total_observations': int(len(df)),
                'confidence_level': 0.95
            }
        }

        print(f"[Pearson] Response prepared successfully with {len(pairwise_results)} comparisons")
        return JsonResponse(response_data)

    except Exception as e:
        import traceback
        error_msg = str(e)
        traceback_msg = traceback.format_exc()
        print(f"[Pearson] Error: {error_msg}")
        print(f"[Pearson] Traceback: {traceback_msg}")
        return JsonResponse({
            'success': False, 
            'error': error_msg,
            'traceback': traceback_msg
        })



def process_spearman_test(request, df, selected_columns, user_id):
    from scipy.stats import spearmanr
    import numpy as np
    import pandas as pd
    from django.http import JsonResponse

    try:
        # ── 1) Extract inputs ──────────────────────────────────────────────
        lang = (request.POST.get("language", "en") or "en").lower()
        if lang not in ('en', 'bn'):
            lang = 'en'
        
        is_bn = (lang == "bn")

        # Variables to analyze
        vars_list = selected_columns or []
        if len(vars_list) < 2:
            return JsonResponse({
                'success': False, 
                'error': 'Please select at least two numeric variables.'
            })

        print(f"[Spearman] Analyzing {len(vars_list)} variables: {vars_list}")

        # ── 2) Core helper functions ───────────────────────────────────────
        def fnum(x):
            """Convert to float, handling None and NaN"""
            if x is None or (isinstance(x, float) and np.isnan(x)):
                return None
            try:
                return float(x)
            except Exception:
                return None

        def spearman_one_pair(var_a, var_b):
            """
            Perform Spearman rank correlation for one pair of variables.
            Returns dict with test results.
            """
            # Ensure numeric data
            s1 = pd.to_numeric(df[var_a], errors='coerce')
            s2 = pd.to_numeric(df[var_b], errors='coerce')
            
            # Remove pairs where either is NaN
            mask = s1.notna() & s2.notna()
            s1_clean, s2_clean = s1[mask], s2[mask]
            
            n = len(s1_clean)
            
            if n < 2:
                return {
                    'n': 0,
                    'correlation': None,
                    'p_value': None,
                    'ci_lower': None,
                    'ci_upper': None
                }
            
            try:
                # Calculate Spearman rank correlation
                rho, p_value = spearmanr(s1_clean, s2_clean)
                
                # Calculate 95% confidence interval using Fisher z-transform
                # Note: This is an approximation for Spearman's rho
                z = np.arctanh(rho)
                se = 1 / np.sqrt(n - 3)
                z_lower = z - 1.96 * se
                z_upper = z + 1.96 * se
                ci_lower = np.tanh(z_lower)
                ci_upper = np.tanh(z_upper)
                
                return {
                    'n': n,
                    'correlation': float(rho),
                    'p_value': float(p_value),
                    'ci_lower': float(ci_lower),
                    'ci_upper': float(ci_upper)
                }
                
            except Exception as e:
                print(f"[Spearman] Error in spearmanr: {e}")
                return {
                    'n': n,
                    'correlation': None,
                    'p_value': None,
                    'ci_lower': None,
                    'ci_upper': None
                }

        def fdr_bh(pvals):
            """Benjamini-Hochberg FDR correction"""
            p = np.asarray(pvals, float)
            n = p.size
            if n == 0:
                return np.array([])
            order = np.argsort(p)
            ranked = p[order]
            adj = np.empty(n, float)
            cm = 1.0
            for i in range(n-1, -1, -1):
                cm = min(cm, ranked[i] * n / (i+1))
                adj[i] = cm
            out = np.empty(n, float)
            out[order] = np.minimum(adj, 1.0)
            return out

        # ── 3) Compute all pairwise tests ──────────────────────────────────
        n_vars = len(vars_list)
        pairwise_results = []
        all_p_values = []
        
        for i in range(n_vars):
            for j in range(i + 1, n_vars):
                var1 = vars_list[i]
                var2 = vars_list[j]
                
                result = spearman_one_pair(var1, var2)
                
                pairwise_results.append({
                    'variable1': str(var1),
                    'variable2': str(var2),
                    'correlation': fnum(result['correlation']),
                    'p_value': fnum(result['p_value']),
                    'ci_lower': fnum(result['ci_lower']),
                    'ci_upper': fnum(result['ci_upper']),
                    'n': result['n']
                })
                
                if result['p_value'] is not None:
                    all_p_values.append(result['p_value'])
                else:
                    all_p_values.append(np.nan)

        print(f"[Spearman] Completed {len(pairwise_results)} pairwise tests")

        # ── 4) Apply FDR correction ────────────────────────────────────────
        p_array = np.array(all_p_values, dtype=float)
        mask = ~np.isnan(p_array)
        
        if mask.any():
            adjusted_p = np.full(len(pairwise_results), np.nan, dtype=float)
            adjusted_p[mask] = fdr_bh(p_array[mask])
            
            for i, result in enumerate(pairwise_results):
                result['p_adjusted'] = float(adjusted_p[i]) if not np.isnan(adjusted_p[i]) else None
        else:
            for result in pairwise_results:
                result['p_adjusted'] = None

        # Sort by absolute correlation (descending)
        pairwise_results.sort(
            key=lambda r: (r['correlation'] is None, -abs(r['correlation']) if r['correlation'] is not None else 0)
        )

        # ── 5) Create correlation matrix for heatmap ───────────────────────
        corr_matrix = np.ones((n_vars, n_vars), dtype=float)
        p_matrix = np.ones((n_vars, n_vars), dtype=float)
        p_adjusted_matrix = np.ones((n_vars, n_vars), dtype=float)
        
        for result in pairwise_results:
            i = vars_list.index(result['variable1'])
            j = vars_list.index(result['variable2'])
            
            corr_val = result['correlation'] if result['correlation'] is not None else 0.0
            corr_matrix[i, j] = corr_val
            corr_matrix[j, i] = corr_val
            
            p_val = result['p_value'] if result['p_value'] is not None else 1.0
            p_matrix[i, j] = p_val
            p_matrix[j, i] = p_val
            
            p_adj_val = result['p_adjusted'] if result['p_adjusted'] is not None else 1.0
            p_adjusted_matrix[i, j] = p_adj_val
            p_adjusted_matrix[j, i] = p_adj_val
        
        np.fill_diagonal(corr_matrix, 1.0)
        np.fill_diagonal(p_matrix, 1.0)
        np.fill_diagonal(p_adjusted_matrix, 1.0)

        # ── 6) Prepare variable-level statistics ──────────────────────────
        variable_stats = []
        
        for var in vars_list:
            col_data = pd.to_numeric(df[var], errors='coerce').dropna()
            
            variable_stats.append({
                'variable': str(var),
                'n_observations': int(len(col_data)),
                'n_missing': int(df[var].isna().sum()),
                'mean': float(col_data.mean()) if len(col_data) > 0 else None,
                'std': float(col_data.std()) if len(col_data) > 0 else None,
                'min': float(col_data.min()) if len(col_data) > 0 else None,
                'max': float(col_data.max()) if len(col_data) > 0 else None,
                'median': float(col_data.median()) if len(col_data) > 0 else None,
                'q1': float(col_data.quantile(0.25)) if len(col_data) > 0 else None,
                'q3': float(col_data.quantile(0.75)) if len(col_data) > 0 else None
            })

        # ── 7) Create blocks (one anchor variable at a time) ──────────────
        blocks = []
        
        for anchor in vars_list:
            block_results = []
            
            for result in pairwise_results:
                if result['variable1'] == anchor:
                    block_results.append(result)
                elif result['variable2'] == anchor:
                    # Swap variables so anchor is always first
                    swapped_result = {
                        'variable1': result['variable2'],
                        'variable2': result['variable1'],
                        'correlation': result['correlation'],
                        'p_value': result['p_value'],
                        'p_adjusted': result['p_adjusted'],
                        'ci_lower': result['ci_lower'],
                        'ci_upper': result['ci_upper'],
                        'n': result['n']
                    }
                    block_results.append(swapped_result)
            
            blocks.append({
                'anchor': str(anchor),
                'results': block_results
            })

        # ── 8) Prepare table columns for frontend ─────────────────────────
        if is_bn:
            table_columns = [
                {"key": "variable1", "label": "ভেরিয়েবল ১"},
                {"key": "variable2", "label": "ভেরিয়েবল ২"},
                {"key": "correlation", "label": "স্পিয়ারম্যান সম্পর্ক"},
                {"key": "p_value", "label": "পি-মান"},
                {"key": "p_adjusted", "label": "সমন্বিত পি-মান"},
                {"key": "ci_lower", "label": "৯৫% সিআই নিম্ন"},
                {"key": "ci_upper", "label": "৯৫% সিআই উচ্চ"},
                {"key": "n", "label": "নমুনা"},
            ]
        else:
            table_columns = [
                {"key": "variable1", "label": "Variable 1"},
                {"key": "variable2", "label": "Variable 2"},
                {"key": "correlation", "label": "Spearman Correlation"},
                {"key": "p_value", "label": "P-value"},
                {"key": "p_adjusted", "label": "Adjusted P-value"},
                {"key": "ci_lower", "label": "95% CI Lower"},
                {"key": "ci_upper", "label": "95% CI Upper"},
                {"key": "n", "label": "N"},
            ]

        # ── 9) Prepare plot data for visualizations ───────────────────────
        plot_data = []
        for i, var1 in enumerate(vars_list):
            for j, var2 in enumerate(vars_list):
                if i != j:
                    # Find the corresponding result
                    result = None
                    for r in pairwise_results:
                        if (r['variable1'] == var1 and r['variable2'] == var2) or \
                           (r['variable1'] == var2 and r['variable2'] == var1):
                            result = r
                            break
                    
                    if result:
                        plot_data.append({
                            'category': f"{var1}",
                            'variable': f"{var2}",
                            'p_value': result['p_value'],
                            'correlation': result['correlation'],
                            'p_adjusted': result['p_adjusted']
                        })

        # ── 10) Prepare final response ─────────────────────────────────────
        test_name = 'Spearman Rank Correlation Test' if lang == 'en' else 'স্পিয়ারম্যান র্যাঙ্ক সম্পর্ক পরীক্ষা'
        
        response_data = {
            'success': True,
            'test': test_name,
            'language': lang,
            'n_variables': int(n_vars),
            'n_comparisons': int(len(pairwise_results)),
            'variables': [str(v) for v in vars_list],
            'variable_stats': variable_stats,
            'pairwise_results': pairwise_results,
            'blocks': blocks,
            'correlation_matrix': [[float(val) for val in row] for row in corr_matrix.tolist()],
            'p_value_matrix': [[float(val) for val in row] for row in p_matrix.tolist()],
            'p_adjusted_matrix': [[float(val) for val in row] for row in p_adjusted_matrix.tolist()],
            'plot_data': plot_data,
            'table_columns': table_columns,
            'metadata': {
                'fdr_correction': 'Benjamini-Hochberg',
                'alpha': 0.05,
                'total_observations': int(len(df)),
                'confidence_level': 0.95,
                'correlation_type': 'spearman'
            }
        }

        print(f"[Spearman] Response prepared successfully with {len(pairwise_results)} comparisons")
        return JsonResponse(response_data)

    except Exception as e:
        import traceback
        error_msg = str(e)
        traceback_msg = traceback.format_exc()
        print(f"[Spearman] Error: {error_msg}")
        print(f"[Spearman] Traceback: {traceback_msg}")
        return JsonResponse({
            'success': False, 
            'error': error_msg,
            'traceback': traceback_msg
        })


'''
def process_cramers_heatmap(request, selected_columns, df, user_id):
    

    try:
        # --- 1) Inputs ---
        lang = request.POST.get("language", "en")
        fmt = request.POST.get("format", "png").lower()
        pil_fmt = {"png": "PNG", "jpg": "JPEG", "jpeg": "JPEG", "pdf": "PDF", "tiff": "TIFF"}.get(fmt, "PNG")

        label_font_size = 58
        tick_font_size =36
        img_quality = int(request.POST.get("image_quality", 90))
        palette = request.POST.get("palette", "coolwarm")
        image_size = request.POST.get("image_size", "800x600")

        
        try:
            width, height = map(int, image_size.lower().split("x"))
        except:
            width, height = 800, 600

        # --- 2) Font & Path setup ---
        font_path = os.path.join(settings.BASE_DIR, 'NotoSansBengali-Regular.ttf')
        fm.fontManager.addfont(font_path)
        bengali_font = fm.FontProperties(fname=font_path).get_name()
        mpl.rcParams['font.family'] = bengali_font
        tick_prop = fm.FontProperties(fname=font_path, size=tick_font_size) if lang == 'bn' else fm.FontProperties(size=tick_font_size)
        label_font = ImageFont.truetype(font_path, size=label_font_size)

        translator = Translator()
        def translate(txt): return translator.translate(txt, dest='bn').text if lang == 'bn' else txt
        def map_digits(txt): return txt.translate(str.maketrans('0123456789', '০১২৩৪৫৬৭৮৯')) if lang == 'bn' else txt

        plots_dir = os.path.join(settings.MEDIA_ROOT, f'ID_{user_id}_uploads', 'temporary_uploads', 'plots')
        os.makedirs(plots_dir, exist_ok=True)
        uid = str(uuid.uuid4())[:8]


        if len(selected_columns) < 2:
            return JsonResponse({'success': False, 'error': "Select at least two columns"})
        print(selected_columns) 

        # --- 3) Compute Cramér’s V matrix ---
        def cramers_v(x, y):
            confusion_matrix = pd.crosstab(x, y)
            # print("Confusion Matrix:\n", confusion_matrix)

            chi2, _, _, _ = chi2_contingency(confusion_matrix)  # ✅ use scipy
            n = confusion_matrix.sum().sum()
            phi2 = chi2 / n
            r, k = confusion_matrix.shape
            #bias correction
            phi2corr = max(0,phi2-((k-1)*(r-1))/(n-1))
            rcorr = r - ((r - 1) ** 2) / ( n - 1 )
            kcorr = k - ((k - 1) ** 2) / ( n - 1 )
            return np.sqrt(phi2corr / min(kcorr - 1, rcorr - 1)) 


        n = len(selected_columns)
        mat = pd.DataFrame(np.zeros((n, n)), index=selected_columns, columns=selected_columns)
        for i in range(n):
            for j in range(n):
                if i == j:
                    mat.iloc[i, j] = 1.0
                else:
                    mat.iloc[i, j] = cramers_v(df[selected_columns[i]], df[selected_columns[j]])
        # print(mat) 
        # --- 4) Heatmap ---
        fig, ax = plt.subplots(figsize=(width/100, height/100), dpi=100)
        cmap = sns.color_palette(palette, as_cmap=True)
        hm = sns.heatmap(mat, annot=True, fmt=".2f", cmap=cmap, ax=ax, cbar=True)
        

        
        # Hide all Matplotlib text; PIL will draw it
        ax.set_xlabel(''); ax.set_ylabel(''); ax.set_title('')
        ax.set_xticklabels([]); ax.set_yticklabels([])
        ax.tick_params(axis='both', length=0)

        
        fig.canvas.draw()
        renderer = fig.canvas.get_renderer()

        n_rows, n_cols = mat.shape

        
        x_left_fig,  _          = ax.transData.transform((0,        0))
        x_right_fig, _          = ax.transData.transform((n_cols,   0))
        
        _, y_top_fig            = ax.transData.transform((0,        0))
        _, y_bottom_fig         = ax.transData.transform((0,   n_rows))

        # column centers (x only)
        col_centers_fig = [ax.transData.transform((j + 0.5, 0))[0] for j in range(n_cols)]
        # row centers (y only)
        row_centers_fig = [ax.transData.transform((0, i + 0.5))[1] for i in range(n_rows)]

    
        save_dpi = 300
        scale = save_dpi / fig.dpi

        x_left_px   = int(x_left_fig   * scale)
        x_right_px  = int(x_right_fig  * scale)
        y_top_px    = int(y_top_fig    * scale)
        y_bottom_px = int(y_bottom_fig * scale)
        col_centers_px = [int(x * scale) for x in col_centers_fig]
        row_centers_px = [int(y * scale) for y in row_centers_fig]

        
        base_png_path = os.path.join(plots_dir, f"cramers_heatmap_{uid}_base.png")
        fig.savefig(base_png_path, dpi=save_dpi)   
        plt.close(fig)


        # Prepare localized labels 
        title_txt  = map_digits(translate("Cramér's V Heatmap"))
        col_labels = [map_digits(translate(str(t))) for t in mat.columns]
        row_labels = [map_digits(translate(str(t))) for t in mat.index]

        
        tick_font  = ImageFont.truetype(font_path, size=tick_font_size)

        # Open base and make a canvas with margins
        base = Image.open(base_png_path).convert("RGB")
        bw, bh = base.size

        def _w(font, txt): 
            if not txt: return 0
            a,b,c,d = font.getbbox(txt); return c-a
        def _h(font, txt): 
            if not txt: return 0
            a,b,c,d = font.getbbox(txt); return d-b

        title_h = _h(label_font, title_txt) if title_txt else 0
        pad     = max(8, tick_font.size // 2)

        # estimate needed margins
        max_x_tick_h = max((_w(tick_font, s) for s in col_labels), default=0) // 2 + tick_font.size // 2
        max_y_tick_w = max((_w(tick_font, s) for s in row_labels), default=0)

        top_margin    = (title_h + pad) if title_txt else pad
        bottom_margin = (max_x_tick_h + pad) if col_labels else pad
        left_margin   = (max_y_tick_w + pad) if row_labels else pad
        right_margin  = pad

        W = bw + left_margin + right_margin
        H = bh + top_margin + bottom_margin
        canvas = Image.new("RGB", (W, H), "white")
        canvas.paste(base, (left_margin, top_margin))
        draw = ImageDraw.Draw(canvas)

        # Title (top, centered)
        if title_txt:
            tlen = draw.textlength(title_txt, font=label_font)
            tx = (W - int(tlen)) // 2
            ty = (top_margin - title_h) // 2
            draw.text((tx, ty), title_txt, font=label_font, fill="black")

        # Convert heatmap edge & center pixels into canvas coords
        x_left_c   = left_margin + x_left_px
        x_right_c  = left_margin + x_right_px
        y_top_c    = top_margin  + y_top_px
        y_bottom_c = top_margin  + y_bottom_px
        col_centers_c = [left_margin + x for x in col_centers_px]
        row_centers_c = [top_margin  + y for y in row_centers_px]

        # X labels (bottom, rotated 45°, centered under each column)
        base_y =  y_bottom_c + pad
        # X labels: bottom of the axis, horizontally centered under each column
        for cx, lab in zip(col_centers_c, col_labels):
            tw = int(draw.textlength(lab, font=tick_font))
            th = tick_font.size
            lx = cx - tw               # center text under the column
            ly = y_bottom_c + pad           # place just below the heatmap bottom edge
            
            lab_img = Image.new("RGBA", (tw + 8, th + 8), (255,255,255,0))
            d2 = ImageDraw.Draw(lab_img)
            d2.text((4, 0), lab, font=tick_font, fill="black")
            lab_rot = lab_img.rotate(30, expand=True)
            lx = cx - lab_rot.width
            ly = y_top_c + pad
            canvas.paste(lab_rot, (lx, ly), lab_rot)



        # Y labels (left, horizontal, centered on each row)
        for cy, lab in zip(row_centers_c, row_labels):
            tw = int(draw.textlength(lab, font=tick_font))
            lx = max(0, x_left_c - pad - tw)
            ly = cy - tick_font.size // 2
            draw.text((lx, ly), lab, font=tick_font, fill="black")

        # Final save
        final_path = os.path.join(plots_dir, f"cramers_heatmap_{uid}.{fmt}")
        if pil_fmt == "JPEG":
            canvas.save(final_path, format=pil_fmt, quality=img_quality, subsampling=0, progressive=True)
        elif pil_fmt == "PNG":
            canvas.save(final_path, format=pil_fmt, optimize=True)
        else:
            canvas.save(final_path, format=pil_fmt)



        print(final_path) 

        return JsonResponse({
            'success': True,
            'image_paths': [os.path.join(settings.MEDIA_URL, f'ID_{user_id}_uploads', 'temporary_uploads', 'plots', os.path.basename(final_path))],
            'columns': selected_columns,
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
'''

def process_cramers_test(request, df, selected_columns, user_id):
    import numpy as np
    import pandas as pd
    from scipy.stats import chi2_contingency
    from django.http import JsonResponse

    try:
        # ── 1) Extract inputs ──────────────────────────────────────────────
        lang = (request.POST.get("language", "en") or "en").lower()
        if lang not in ('en', 'bn'):
            lang = 'en'
        
        is_bn = (lang == "bn")

        # Variables to analyze
        vars_list = selected_columns or []
        if len(vars_list) < 2:
            return JsonResponse({
                'success': False, 
                'error': 'Please select at least two categorical variables.'
            })

        print(f"[Cramer's V] Analyzing {len(vars_list)} variables: {vars_list}")

        # ── 2) Core helper functions ───────────────────────────────────────
        def fnum(x):
            """Convert to float, handling None and NaN"""
            if x is None or (isinstance(x, float) and np.isnan(x)):
                return None
            try:
                return float(x)
            except Exception:
                return None

        def inum(x):
            """Convert to int, handling None and NaN"""
            if x is None or (isinstance(x, float) and np.isnan(x)):
                return None
            try:
                return int(x)
            except Exception:
                try:
                    return int(float(x))
                except Exception:
                    return None

        def cramers_v_one_pair(var_a, var_b, dropna=True, include_na_as_category=False, bias_correction=True):
            """
            Calculate Cramer's V for one pair of variables.
            Returns dict with Cramer's V results and contingency table data.
            """
            s1, s2 = df[var_a].copy(), df[var_b].copy()
            
            # Handle missing values
            if include_na_as_category:
                s1 = s1.fillna("__NA__")
                s2 = s2.fillna("__NA__")
            elif dropna:
                mask = s1.notna() & s2.notna()
                s1, s2 = s1[mask], s2[mask]
            
            # Create contingency table
            ct = pd.crosstab(s1, s2, dropna=False)
            
            if ct.shape[0] == 0 or ct.shape[1] == 0:
                return {
                    'n': 0,
                    'cramers_v': None,
                    'chi2': None,
                    'p_value': None,
                    'dof': None,
                    'contingency_table': None,
                    'expected_frequencies': None,
                    'n_categories_var1': 0,
                    'n_categories_var2': 0,
                    'effect_size': None,
                    'categories_var1': [],
                    'categories_var2': [],
                    'var1_categories': [],
                    'var2_categories': []
                }
            
            n_obs = ct.sum().sum()
            
            # Calculate chi-square statistics
            try:
                chi2, p_value, dof, expected = chi2_contingency(ct)
            except Exception as e:
                print(f"[Cramer's V] Error in chi2_contingency: {e}")
                return {
                    'n': n_obs,
                    'cramers_v': None,
                    'chi2': None,
                    'p_value': None,
                    'dof': None,
                    'contingency_table': [[int(val) for val in row] for row in ct.values.tolist()],
                    'expected_frequencies': None,
                    'n_categories_var1': ct.shape[0],
                    'n_categories_var2': ct.shape[1],
                    'effect_size': None,
                    'categories_var1': [str(c) for c in ct.index.tolist()],
                    'categories_var2': [str(c) for c in ct.columns.tolist()],
                    'var1_categories': [str(c) for c in ct.index.tolist()],
                    'var2_categories': [str(c) for c in ct.columns.tolist()]
                }
            
            # Calculate Cramer's V with optional bias correction
            phi2 = chi2 / n_obs
            r, k = ct.shape
            
            if bias_correction:
                # Bias-corrected Cramer's V
                phi2corr = max(0, phi2 - ((k - 1) * (r - 1)) / (n_obs - 1))
                rcorr = r - ((r - 1) ** 2) / (n_obs - 1)
                kcorr = k - ((k - 1) ** 2) / (n_obs - 1)
                denominator = min((kcorr - 1), (rcorr - 1))
                if denominator <= 0:
                    cramers_v_value = 0.0
                else:
                    cramers_v_value = np.sqrt(phi2corr / denominator)
            else:
                # Standard Cramer's V
                denominator = min((k - 1), (r - 1))
                if denominator <= 0:
                    cramers_v_value = 0.0
                else:
                    cramers_v_value = np.sqrt(phi2 / denominator)
            
            # Interpret effect size
            if cramers_v_value is not None:
                if cramers_v_value < 0.1:
                    effect_size = 'Negligible'
                elif cramers_v_value < 0.3:
                    effect_size = 'Small'
                elif cramers_v_value < 0.5:
                    effect_size = 'Medium'
                else:
                    effect_size = 'Large'
            else:
                effect_size = None
            
            # Get category names
            categories_var1 = [str(c) for c in ct.index.tolist()]
            categories_var2 = [str(c) for c in ct.columns.tolist()]
            
            return {
                'n': int(n_obs),
                'cramers_v': float(cramers_v_value) if not np.isnan(cramers_v_value) else None,
                'chi2': float(chi2),
                'p_value': float(p_value),
                'dof': int(dof),
                'contingency_table': [[int(val) for val in row] for row in ct.values.tolist()],
                'expected_frequencies': [[float(val) if not np.isnan(val) else 0.0 for val in row] for row in expected.tolist()],
                'n_categories_var1': r,
                'n_categories_var2': k,
                'effect_size': effect_size,
                'categories_var1': categories_var1,
                'categories_var2': categories_var2,
                'var1_categories': categories_var1,
                'var2_categories': categories_var2
            }

        def fdr_bh(pvals):
            """Benjamini-Hochberg FDR correction"""
            p = np.asarray(pvals, float)
            n = p.size
            if n == 0:
                return np.array([])
            order = np.argsort(p)
            ranked = p[order]
            adj = np.empty(n, float)
            cm = 1.0
            for i in range(n-1, -1, -1):
                cm = min(cm, ranked[i] * n / (i+1))
                adj[i] = cm
            out = np.empty(n, float)
            out[order] = np.minimum(adj, 1.0)
            return out

        # ── 3) Compute all pairwise tests ──────────────────────────────────
        n_vars = len(vars_list)
        pairwise_results = []
        all_p_values = []
        
        for i in range(n_vars):
            for j in range(i + 1, n_vars):
                var1 = vars_list[i]
                var2 = vars_list[j]
                
                print(f"[Cramer's V] Testing {var1} vs {var2}")
                
                result = cramers_v_one_pair(
                    var1, var2,
                    dropna=True,
                    include_na_as_category=False,
                    bias_correction=True
                )
                
                pairwise_results.append({
                    'variable1': str(var1),
                    'variable2': str(var2),
                    'cramers_v': fnum(result['cramers_v']),
                    'chi2': fnum(result['chi2']),
                    'p_value': fnum(result['p_value']),
                    'dof': inum(result['dof']),
                    'n': inum(result['n']),
                    'n_categories_var1': result['n_categories_var1'],
                    'n_categories_var2': result['n_categories_var2'],
                    'contingency_table': result['contingency_table'],
                    'expected_frequencies': result['expected_frequencies'],
                    'effect_size': result['effect_size'],
                    'categories_var1': result['categories_var1'],
                    'categories_var2': result['categories_var2'],
                    'var1_categories': result['var1_categories'],
                    'var2_categories': result['var2_categories']
                })
                
                if result['p_value'] is not None:
                    all_p_values.append(result['p_value'])
                else:
                    all_p_values.append(np.nan)

        print(f"[Cramer's V] Completed {len(pairwise_results)} pairwise tests")

        # ── 4) Apply FDR correction ────────────────────────────────────────
        p_array = np.array(all_p_values, dtype=float)
        mask = ~np.isnan(p_array)
        
        if mask.any():
            adjusted_p = np.full(len(pairwise_results), np.nan, dtype=float)
            adjusted_p[mask] = fdr_bh(p_array[mask])
            
            for i, result in enumerate(pairwise_results):
                result['p_adjusted'] = float(adjusted_p[i]) if not np.isnan(adjusted_p[i]) else None
        else:
            for result in pairwise_results:
                result['p_adjusted'] = None

        # ── 5) Create matrices for heatmap ────────────────────────────────
        cramers_v_matrix = np.ones((n_vars, n_vars), dtype=float)
        p_value_matrix = np.ones((n_vars, n_vars), dtype=float)
        p_adjusted_matrix = np.ones((n_vars, n_vars), dtype=float)
        effect_size_matrix = np.full((n_vars, n_vars), '', dtype=object)
        
        for result in pairwise_results:
            i = vars_list.index(result['variable1'])
            j = vars_list.index(result['variable2'])
            
            cramers_v_val = result['cramers_v'] if result['cramers_v'] is not None else 0.0
            cramers_v_matrix[i, j] = cramers_v_val
            cramers_v_matrix[j, i] = cramers_v_val
            
            p_val = result['p_value'] if result['p_value'] is not None else 1.0
            p_value_matrix[i, j] = p_val
            p_value_matrix[j, i] = p_val
            
            p_adj_val = result['p_adjusted'] if result['p_adjusted'] is not None else 1.0
            p_adjusted_matrix[i, j] = p_adj_val
            p_adjusted_matrix[j, i] = p_adj_val
            
            effect_size_val = result['effect_size'] if result['effect_size'] is not None else 'Negligible'
            effect_size_matrix[i, j] = effect_size_val
            effect_size_matrix[j, i] = effect_size_val
        
        # Set diagonal to 1.0 for Cramer's V (perfect self-association)
        np.fill_diagonal(cramers_v_matrix, 1.0)
        np.fill_diagonal(p_value_matrix, 1.0)
        np.fill_diagonal(p_adjusted_matrix, 1.0)
        np.fill_diagonal(effect_size_matrix, 'Perfect')

        # ── 6) Prepare variable-level statistics ──────────────────────────
        variable_stats = []
        
        for var in vars_list:
            col_data = df[var].dropna()
            value_counts = col_data.value_counts()
            
            # Calculate entropy
            proportions = value_counts / len(col_data)
            entropy = -sum(p * np.log(p) for p in proportions if p > 0) if len(col_data) > 0 else 0.0
            
            variable_stats.append({
                'variable': str(var),
                'n_categories': int(len(value_counts)),
                'n_observations': int(len(col_data)),
                'n_missing': int(df[var].isna().sum()),
                'categories': [str(c) for c in value_counts.index.tolist()],
                'frequencies': [int(f) for f in value_counts.values.tolist()],
                'percentages': [float(p) for p in (value_counts / len(col_data) * 100).values.tolist()],
                'entropy': float(entropy)
            })

        # ── 7) Create blocks (one anchor variable at a time) ──────────────
        blocks = []
        
        for anchor in vars_list:
            block_results = []
            
            for result in pairwise_results:
                if result['variable1'] == anchor:
                    block_results.append(result)
                elif result['variable2'] == anchor:
                    # Swap variables so anchor is always first
                    swapped_result = {
                        'variable1': result['variable2'],
                        'variable2': result['variable1'],
                        'cramers_v': result['cramers_v'],
                        'chi2': result['chi2'],
                        'p_value': result['p_value'],
                        'p_adjusted': result['p_adjusted'],
                        'dof': result['dof'],
                        'n': result['n'],
                        'n_categories_var1': result['n_categories_var2'],
                        'n_categories_var2': result['n_categories_var1'],
                        # Swap contingency table dimensions
                        'contingency_table': [[result['contingency_table'][j][i] for j in range(len(result['contingency_table']))] 
                                             for i in range(len(result['contingency_table'][0]))] if result['contingency_table'] else None,
                        'expected_frequencies': [[result['expected_frequencies'][j][i] for j in range(len(result['expected_frequencies']))] 
                                                for i in range(len(result['expected_frequencies'][0]))] if result['expected_frequencies'] else None,
                        'effect_size': result['effect_size'],
                        # Swap categories
                        'categories_var1': result['categories_var2'],
                        'categories_var2': result['categories_var1'],
                        'var1_categories': result['var2_categories'],
                        'var2_categories': result['var1_categories']
                    }
                    block_results.append(swapped_result)
            
            # Sort by Cramer's V (descending)
            block_results.sort(key=lambda x: x['cramers_v'] or 0, reverse=True)
            
            blocks.append({
                'anchor': str(anchor),
                'results': block_results
            })

        # ── 8) Prepare table columns for frontend ─────────────────────────
        if is_bn:
            table_columns = [
                {"key": "variable1", "label": "ভেরিয়েবল ১"},
                {"key": "variable2", "label": "ভেরিয়েবল ২"},
                {"key": "cramers_v", "label": "ক্রেমার'স ভি"},
                {"key": "effect_size", "label": "সম্পর্কের মাত্রা"},
                {"key": "p_value", "label": "পি-মান"},
                {"key": "p_adjusted", "label": "সমন্বিত পি-মান"},
                {"key": "chi2", "label": "কাই-স্কয়ার"},
                {"key": "dof", "label": "স্বাধীনতার মাত্রা"},
                {"key": "n", "label": "নমুনা"},
            ]
        else:
            table_columns = [
                {"key": "variable1", "label": "Variable 1"},
                {"key": "variable2", "label": "Variable 2"},
                {"key": "cramers_v", "label": "Cramer's V"},
                {"key": "effect_size", "label": "Effect Size"},
                {"key": "p_value", "label": "P-value"},
                {"key": "p_adjusted", "label": "Adjusted P-value"},
                {"key": "chi2", "label": "Chi²"},
                {"key": "dof", "label": "DoF"},
                {"key": "n", "label": "N"},
            ]

        # ── 9) Prepare plot data for visualizations ───────────────────────
        plot_data = []
        for i, var1 in enumerate(vars_list):
            for j, var2 in enumerate(vars_list):
                if i != j:
                    # Find the corresponding result
                    result = None
                    for r in pairwise_results:
                        if (r['variable1'] == var1 and r['variable2'] == var2) or \
                           (r['variable1'] == var2 and r['variable2'] == var1):
                            result = r
                            break
                    
                    if result:
                        plot_data.append({
                            'category': f"{var1}",
                            'variable': f"{var2}",
                            'cramers_v': result['cramers_v'],
                            'p_value': result['p_value'],
                            'effect_size': result['effect_size'],
                            'p_adjusted': result['p_adjusted']
                        })

        # ── 10) Create association strength summary ──────────────────────
        effect_size_counts = {
            'Large': 0,
            'Medium': 0,
            'Small': 0,
            'Negligible': 0
        }
        
        for result in pairwise_results:
            if result['effect_size'] in effect_size_counts:
                effect_size_counts[result['effect_size']] += 1

        # ── 11) Generate insight message ──────────────────────────────────
        def generate_insight_message():
            strong_associations = [r for r in pairwise_results if r['effect_size'] in ['Medium', 'Large']]
            
            if is_bn:
                if len(strong_associations) == 0:
                    return "কোনো শক্তিশালী সম্পর্ক পাওয়া যায়নি (ক্রেমার'স ভি < ০.৩)। সমস্ত ভেরিয়েবল দুর্বলভাবে সম্পর্কিত।"
                elif len(strong_associations) == len(pairwise_results):
                    return f"সব {len(strong_associations)}টি ভেরিয়েবল জোড়ায় শক্তিশালী সম্পর্ক পাওয়া গেছে (ক্রেমার'স ভি ≥ ০.৩)।"
                else:
                    return f"{len(strong_associations)}টি ভেরিয়েবল জোড়ায় শক্তিশালী সম্পর্ক পাওয়া গেছে (মোট {len(pairwise_results)}টির মধ্যে)।"
            else:
                if len(strong_associations) == 0:
                    return "No strong associations found (Cramer's V < 0.3). All variables are weakly related."
                elif len(strong_associations) == len(pairwise_results):
                    return f"All {len(strong_associations)} variable pairs show strong associations (Cramer's V ≥ 0.3)."
                else:
                    return f"{len(strong_associations)} variable pairs show strong associations (out of {len(pairwise_results)} total)."

        # ── 12) Prepare final response ────────────────────────────────────
        test_name = "Cramer's V Association Analysis" if lang == 'en' else "ক্রেমার'স ভি সম্পর্ক বিশ্লেষণ"
        
        response_data = {
            'success': True,
            'test': test_name,
            'language': lang,
            'n_variables': int(n_vars),
            'n_comparisons': int(len(pairwise_results)),
            'variables': [str(v) for v in vars_list],
            'variable_stats': variable_stats,
            'pairwise_results': pairwise_results,
            'blocks': blocks,
            'cramers_v_matrix': [[float(val) for val in row] for row in cramers_v_matrix.tolist()],
            'p_value_matrix': [[float(val) for val in row] for row in p_value_matrix.tolist()],
            'p_adjusted_matrix': [[float(val) for val in row] for row in p_adjusted_matrix.tolist()],
            'effect_size_matrix': [[str(val) for val in row] for row in effect_size_matrix.tolist()],
            'plot_data': plot_data,
            'table_columns': table_columns,
            'effect_size_summary': effect_size_counts,
            'insight_message': generate_insight_message(),
            'metadata': {
                'fdr_correction': 'Benjamini-Hochberg',
                'bias_correction': True,
                'effect_size_interpretation': {
                    'Negligible': 'V < 0.1',
                    'Small': '0.1 ≤ V < 0.3', 
                    'Medium': '0.3 ≤ V < 0.5',
                    'Large': 'V ≥ 0.5'
                },
                'alpha': 0.05,
                'total_observations': int(len(df))
            }
        }

        print(f"[Cramer's V] Response prepared successfully with {len(pairwise_results)} comparisons")
        return JsonResponse(response_data)

    except Exception as e:
        import traceback
        error_msg = str(e)
        traceback_msg = traceback.format_exc()
        print(f"[Cramer's V] Error: {error_msg}")
        print(f"[Cramer's V] Traceback: {traceback_msg}")
        return JsonResponse({
            'success': False, 
            'error': error_msg,
            'traceback': traceback_msg
        })


def process_cross_tabulation(request, df, selected_columns, user_id):
    import pandas as pd
    import numpy as np
    from django.http import JsonResponse
    import os
    from django.conf import settings

    try:
        # ── 1) Extract inputs ──────────────────────────────────────────────
        lang = (request.POST.get("language", "en") or "en").lower()
        if lang not in ('en', 'bn'):
            lang = 'en'
        
        is_bn = (lang == "bn")

        # Variables to analyze
        vars_list = selected_columns or []
        if len(vars_list) < 2:
            return JsonResponse({
                'success': False, 
                'error': 'Please select at least two categorical variables for cross-tabulation.'
            })

        print(f"[Cross-tabulation] Analyzing {len(vars_list)} variables: {vars_list}")

        # ── 2) Core helper functions ───────────────────────────────────────
        def fnum(x):
            """Convert to float, handling None and NaN"""
            if x is None or (isinstance(x, float) and np.isnan(x)):
                return None
            try:
                return float(x)
            except Exception:
                return None

        def inum(x):
            """Convert to int, handling None and NaN"""
            if x is None or (isinstance(x, float) and np.isnan(x)):
                return None
            try:
                return int(x)
            except Exception:
                try:
                    return int(float(x))
                except Exception:
                    return None

        def create_cross_tab(var_a, var_b, dropna=True, include_na_as_category=False, margins=True):
            """
            Create cross-tabulation for two variables.
            Returns dict with cross-tab data and statistics.
            """
            s1, s2 = df[var_a].copy(), df[var_b].copy()
            
            # Handle missing values
            if include_na_as_category:
                s1 = s1.fillna("__NA__")
                s2 = s2.fillna("__NA__")
            elif dropna:
                mask = s1.notna() & s2.notna()
                s1, s2 = s1[mask], s2[mask]
            
            # Create cross-tabulation with margins
            try:
                ct = pd.crosstab(s1, s2, margins=margins, dropna=False)
                
                # Calculate percentages
                row_percentages = (pd.crosstab(s1, s2, normalize='index') * 100).round(2)
                col_percentages = (pd.crosstab(s1, s2, normalize='columns') * 100).round(2)
                total_percentages = (pd.crosstab(s1, s2, normalize='all') * 100).round(2)
                
            except Exception as e:
                print(f"[Cross-tabulation] Error creating crosstab: {e}")
                return {
                    'n': 0,
                    'contingency_table': None,
                    'row_percentages': None,
                    'col_percentages': None,
                    'total_percentages': None,
                    'categories_var1': [],
                    'categories_var2': [],
                    'row_totals': [],
                    'col_totals': [],
                    'margins_row': [],
                    'margins_col': [],
                    'grand_total': 0,
                    'summary_stats': {}
                }
            
            if ct.shape[0] == 0 or ct.shape[1] == 0:
                return {
                    'n': 0,
                    'contingency_table': None,
                    'row_percentages': None,
                    'col_percentages': None,
                    'total_percentages': None,
                    'categories_var1': [],
                    'categories_var2': [],
                    'row_totals': [],
                    'col_totals': [],
                    'margins_row': [],
                    'margins_col': [],
                    'grand_total': 0,
                    'summary_stats': {}
                }
            
            # Extract data without margins for the main table
            ct_no_margins = ct.iloc[:-1, :-1] if margins else ct
            grand_total = ct.iloc[-1, -1] if margins else ct.values.sum()
            
            # Get category names (without margins)
            categories_var1 = [str(c) for c in ct_no_margins.index.tolist()]
            categories_var2 = [str(c) for c in ct_no_margins.columns.tolist()]
            
            # Calculate totals (without margins)
            row_totals = ct_no_margins.sum(axis=1).tolist()
            col_totals = ct_no_margins.sum(axis=0).tolist()
            
            # Extract margins if available
            margins_row = ct.iloc[:-1, -1].tolist() if margins else []
            margins_col = ct.iloc[-1, :-1].tolist() if margins else []
            
            # Calculate summary statistics
            flat_counts = []
            for i, cat1 in enumerate(categories_var1):
                for j, cat2 in enumerate(categories_var2):
                    count = int(ct_no_margins.values[i][j])
                    flat_counts.append({
                        'category1': cat1,
                        'category2': cat2,
                        'count': count,
                        'row_percentage': float(row_percentages.values[i][j]) if not np.isnan(row_percentages.values[i][j]) else 0.0,
                        'col_percentage': float(col_percentages.values[i][j]) if not np.isnan(col_percentages.values[i][j]) else 0.0,
                        'total_percentage': float(total_percentages.values[i][j]) if not np.isnan(total_percentages.values[i][j]) else 0.0
                    })
            
            # Find most and least frequent combinations
            if flat_counts:
                most_frequent = max(flat_counts, key=lambda x: x['count'])
                least_frequent = min(flat_counts, key=lambda x: x['count'])
            else:
                most_frequent = least_frequent = None
            
            # Prepare data for stacked bar chart (similar to chi-square)
            stacked_data = []
            for i, cat1 in enumerate(categories_var1):
                row_data = {
                    'category': str(cat1),
                    'total': int(row_totals[i])
                }
                
                for j, cat2 in enumerate(categories_var2):
                    count = int(ct_no_margins.values[i][j])
                    percentage = float(row_percentages.values[i][j]) if not np.isnan(row_percentages.values[i][j]) else 0.0
                    
                    row_data[str(cat2)] = percentage
                    row_data[f'{cat2}_count'] = count
                
                stacked_data.append(row_data)
            
            return {
                'n': int(grand_total),
                'contingency_table': [[int(val) if not np.isnan(val) else 0 for val in row] for row in ct_no_margins.values.tolist()],
                'row_percentages': [[float(val) if not np.isnan(val) else 0.0 for val in row] for row in row_percentages.values.tolist()],
                'col_percentages': [[float(val) if not np.isnan(val) else 0.0 for val in row] for row in col_percentages.values.tolist()],
                'total_percentages': [[float(val) if not np.isnan(val) else 0.0 for val in row] for row in total_percentages.values.tolist()],
                'categories_var1': categories_var1,
                'categories_var2': categories_var2,
                'row_totals': [int(val) for val in row_totals],
                'col_totals': [int(val) for val in col_totals],
                'margins_row': [int(val) for val in margins_row],
                'margins_col': [int(val) for val in margins_col],
                'grand_total': int(grand_total),
                'stacked_data': stacked_data,
                'summary_stats': {
                    'total_observations': int(grand_total),
                    'most_frequent': most_frequent,
                    'least_frequent': least_frequent,
                    'n_categories_var1': len(categories_var1),
                    'n_categories_var2': len(categories_var2)
                }
            }

        # ── 3) Compute all pairwise cross-tabulations ──────────────────────
        n_vars = len(vars_list)
        pairwise_results = []
        
        for i in range(n_vars):
            for j in range(i + 1, n_vars):
                var1 = vars_list[i]
                var2 = vars_list[j]
                
                print(f"[Cross-tabulation] Creating crosstab for {var1} vs {var2}")
                
                result = create_cross_tab(
                    var1, var2,
                    dropna=True,
                    include_na_as_category=False,
                    margins=True
                )
                
                pairwise_results.append({
                    'variable1': str(var1),
                    'variable2': str(var2),
                    'n': inum(result['n']),
                    'contingency_table': result['contingency_table'],
                    'row_percentages': result['row_percentages'],
                    'col_percentages': result['col_percentages'],
                    'total_percentages': result['total_percentages'],
                    'categories_var1': result['categories_var1'],
                    'categories_var2': result['categories_var2'],
                    'row_totals': result['row_totals'],
                    'col_totals': result['col_totals'],
                    'margins_row': result['margins_row'],
                    'margins_col': result['margins_col'],
                    'grand_total': result['grand_total'],
                    'stacked_data': result['stacked_data'],
                    'summary_stats': result['summary_stats']
                })

        print(f"[Cross-tabulation] Completed {len(pairwise_results)} pairwise cross-tabulations")

        # Sort by total observations (descending)
        pairwise_results.sort(key=lambda r: r['n'] or 0, reverse=True)

        # ── 4) Prepare variable-level statistics ──────────────────────────
        variable_stats = []
        
        for var in vars_list:
            col_data = df[var].dropna()
            value_counts = col_data.value_counts()
            
            variable_stats.append({
                'variable': str(var),
                'n_categories': int(len(value_counts)),
                'n_observations': int(len(col_data)),
                'n_missing': int(df[var].isna().sum()),
                'categories': [str(c) for c in value_counts.index.tolist()],
                'frequencies': [int(f) for f in value_counts.values.tolist()],
                'percentages': [float(p) for p in (value_counts / len(col_data) * 100).values.tolist()]
            })

        # ── 5) Create blocks (one anchor variable at a time) ──────────────
        blocks = []
        
        for anchor in vars_list:
            block_results = []
            
            for result in pairwise_results:
                if result['variable1'] == anchor:
                    block_results.append(result)
                elif result['variable2'] == anchor:
                    # Swap variables so anchor is always first
                    swapped_result = {
                        'variable1': result['variable2'],
                        'variable2': result['variable1'],
                        'n': result['n'],
                        # Transpose the tables
                        'contingency_table': [[result['contingency_table'][j][i] for j in range(len(result['contingency_table']))] 
                                             for i in range(len(result['contingency_table'][0]))] if result['contingency_table'] else None,
                        'row_percentages': [[result['col_percentages'][j][i] for j in range(len(result['col_percentages']))] 
                                           for i in range(len(result['col_percentages'][0]))] if result['col_percentages'] else None,
                        'col_percentages': [[result['row_percentages'][j][i] for j in range(len(result['row_percentages']))] 
                                           for i in range(len(result['row_percentages'][0]))] if result['row_percentages'] else None,
                        'total_percentages': [[result['total_percentages'][j][i] for j in range(len(result['total_percentages']))] 
                                             for i in range(len(result['total_percentages'][0]))] if result['total_percentages'] else None,
                        # Swap categories
                        'categories_var1': result['categories_var2'],
                        'categories_var2': result['categories_var1'],
                        # Swap totals
                        'row_totals': result['col_totals'],
                        'col_totals': result['row_totals'],
                        'margins_row': result['margins_col'],
                        'margins_col': result['margins_row'],
                        'grand_total': result['grand_total'],
                        # Stacked data needs to be regenerated for swapped variables
                        'stacked_data': result['stacked_data'],  # This will need adjustment in frontend
                        'summary_stats': result['summary_stats']
                    }
                    block_results.append(swapped_result)
            
            blocks.append({
                'anchor': str(anchor),
                'results': block_results
            })

        # ── 6) Prepare table columns for frontend ─────────────────────────
        if is_bn:
            table_columns = [
                {"key": "variable1", "label": "ভেরিয়েবল ১"},
                {"key": "variable2", "label": "ভেরিয়েবল ২"},
                {"key": "n", "label": "মোট পর্যবেক্ষণ"},
                {"key": "n_categories_var1", "label": "ভেরিয়েবল ১-এর শ্রেণী"},
                {"key": "n_categories_var2", "label": "ভেরিয়েবল ২-এর শ্রেণী"},
            ]
        else:
            table_columns = [
                {"key": "variable1", "label": "Variable 1"},
                {"key": "variable2", "label": "Variable 2"},
                {"key": "n", "label": "Total Observations"},
                {"key": "n_categories_var1", "label": "Categories in Var1"},
                {"key": "n_categories_var2", "label": "Categories in Var2"},
            ]

        # ── 7) Create matrices for heatmap compatibility ─────────────────
        # Create frequency-based matrices for heatmap visualization
        freq_matrix = np.zeros((n_vars, n_vars), dtype=float)
        obs_matrix = np.zeros((n_vars, n_vars), dtype=float)

        for result in pairwise_results:
            i = vars_list.index(result['variable1'])
            j = vars_list.index(result['variable2'])
            
            # Use actual observation count for visualization
            obs_count = result['n'] or 0
            obs_matrix[i, j] = obs_count
            obs_matrix[j, i] = obs_count
            
            # Normalize frequency by total possible observations
            freq_val = obs_count / (len(df) or 1) if len(df) > 0 else 0
            freq_matrix[i, j] = freq_val
            freq_matrix[j, i] = freq_val

        np.fill_diagonal(obs_matrix, len(df))  # Diagonal represents max possible
        np.fill_diagonal(freq_matrix, 1.0)     # Diagonal represents same variable

        # ── 8) Prepare plot data for visualizations ───────────────────────
        plot_data = []
        for i, var1 in enumerate(vars_list):
            for j, var2 in enumerate(vars_list):
                if i != j:
                    # Find the corresponding result
                    result = None
                    for r in pairwise_results:
                        if (r['variable1'] == var1 and r['variable2'] == var2) or \
                        (r['variable1'] == var2 and r['variable2'] == var1):
                            result = r
                            break
                    
                    if result:
                        plot_data.append({
                            'category': f"{var1}",
                            'variable': f"{var2}",
                            'value': float(freq_matrix[i, j]),
                            'frequency': result['n'] or 0,
                            'total_observations': result['n'] or 0,
                            'normalized_frequency': float(freq_matrix[i, j])
                        })

        # ── 9) Prepare final response ─────────────────────────────────────
        test_name = 'Cross-Tabulation Analysis' if lang == 'en' else 'ক্রস-ট্যাবুলেশন বিশ্লেষণ'
        
        response_data = {
            'success': True,
            'test': test_name,
            'language': lang,
            'n_variables': int(n_vars),
            'n_comparisons': int(len(pairwise_results)),
            'variables': [str(v) for v in vars_list],
            'variable_stats': variable_stats,
            'pairwise_results': pairwise_results,
            'blocks': blocks,
            # Matrices for heatmap compatibility
            'frequency_matrix': [[float(val) for val in row] for row in freq_matrix.tolist()],
            'observation_matrix': [[float(val) for val in row] for row in obs_matrix.tolist()],
            'p_value_matrix': [[float(freq_matrix[i][j]) for j in range(n_vars)] for i in range(n_vars)],  # Use frequency for p-value matrix
            'chi2_matrix': [[float(obs_matrix[i][j]) for j in range(n_vars)] for i in range(n_vars)],  # Use observations for chi2 matrix
            'p_adjusted_matrix': [[float(freq_matrix[i][j]) for j in range(n_vars)] for i in range(n_vars)],  # Use frequency for adjusted p 
            'plot_data': plot_data,
            'table_columns': table_columns,
            'metadata': {
                'total_observations': int(len(df)),
                'analysis_type': 'frequency_distribution',
                'include_margins': True,
                'include_percentages': True,
                'fdr_correction': 'None',
                'yates_correction': False,
                'alpha': 0.05
            }
        }

        print(f"[Cross-tabulation] Response prepared successfully with {len(pairwise_results)} comparisons")
        return JsonResponse(response_data)

    except Exception as e:
        import traceback
        error_msg = str(e)
        traceback_msg = traceback.format_exc()
        print(f"[Cross-tabulation] Error: {error_msg}")
        print(f"[Cross-tabulation] Traceback: {traceback_msg}")
        return JsonResponse({
            'success': False, 
            'error': error_msg,
            'traceback': traceback_msg
        })



def process_eda_basics(request, df, user_id):
    import numpy as np
    import pandas as pd
    from functools import lru_cache
    from googletrans import Translator
    from django.http import JsonResponse

    try:
        lang = request.POST.get('language', 'en')
        if lang not in ['en', 'bn']:
            lang = 'en'

        digit_map_bn = str.maketrans('0123456789', '০১২৩৪৫৬৭৮৯')

        def to_bn_digits(s):
            return s.translate(digit_map_bn) if lang == 'bn' else s

        @lru_cache(None)
        def translate(text):
            return Translator().translate(text, dest='bn').text if lang == 'bn' else text

        def format_series(series):
            return {col: to_bn_digits(f"{val:.4f}") if isinstance(val, (int, float)) else to_bn_digits(str(val))
                    for col, val in series.items()}

        def format_df(df_obj):
            return {
                "columns": list(df_obj.columns),
                "index": list(df_obj.index),
                "data": [[to_bn_digits(f"{val:.4f}") if isinstance(val, (int, float)) else to_bn_digits(str(val)) for val in row]
                         for row in df_obj.values]
            }

        results = {}

        

        # Dataset structure
        results['missing'] = format_series(df.isnull().sum())
        results['info'] = {
            'rows': to_bn_digits(str(df.shape[0])),
            'columns': to_bn_digits(str(df.shape[1])),
            'duplicates': to_bn_digits(str(df.duplicated().sum())),
            'memory': to_bn_digits(f"{df.memory_usage(deep=True).sum() / 1024:.2f}")
        }

        # Select numerical and categorical
        numerical_df = df.select_dtypes(include=['int64', 'float64'])
        categorical_df = df.select_dtypes(include=['object', 'category'])

        # Descriptive statistics
        results['summary'] = format_df(numerical_df.describe().T)
        results['count'] = format_series(numerical_df.count())
        results['mean'] = format_series(numerical_df.mean())
        results['min'] = format_series(numerical_df.min())
        results['max'] = format_series(numerical_df.max())
        results['median'] = format_series(numerical_df.median())
        results['mode'] = format_series(numerical_df.mode().iloc[0])
        results['variance'] = format_series(numerical_df.var())
        results['std'] = format_series(numerical_df.std())

        # MAD
        mad = numerical_df.apply(lambda x: np.median(np.abs(x - np.median(x))))
        results['mad'] = format_series(mad)

        results['skew'] = format_series(numerical_df.skew())
        results['kurt'] = format_series(numerical_df.kurt())
        results['range'] = format_series(numerical_df.max() - numerical_df.min())
        iqr = numerical_df.quantile(0.75) - numerical_df.quantile(0.25)
        results['iqr'] = format_series(iqr)
        cv = numerical_df.std() / numerical_df.mean()
        results['cv'] = format_series(cv)

        outliers = ((numerical_df < (numerical_df.quantile(0.25) - 1.5 * iqr)) |
                    (numerical_df > (numerical_df.quantile(0.75) + 1.5 * iqr)))
        results['outliers'] = format_series(outliers.sum())

        # Entropy
        def compute_entropy(series):
            p = series.value_counts(normalize=True)
            return -np.sum(p * np.log2(p + 1e-9))

        entropy = {}
        for col in categorical_df.columns:
            e_val = compute_entropy(categorical_df[col])
            entropy[col] = to_bn_digits(f"{e_val:.4f}") if lang == 'bn' else f"{e_val:.4f}"
        results['entropy'] = entropy

        return JsonResponse({'success': True, **results})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

def process_similarity(request, df, user_id):
    import os
    import numpy as np
    import pandas as pd
    from numpy.linalg import norm
    from scipy.stats import spearmanr
    from googletrans import Translator
    from django.http import JsonResponse
    from functools import lru_cache
    from django.conf import settings

    try:
        # 1. Read parameters
        col1 = request.POST.get('column1')
        col2 = request.POST.get('column2')
        language = request.POST.get('language', 'en')
        if language not in ('en', 'bn'):
            language = 'en'

        # 2. Translator & digit mapping
        translator = Translator()
        digit_map = str.maketrans('0123456789', '০১২৩৪৫৬৭৮৯')

        @lru_cache(None)
        def translate(txt): return translator.translate(txt, dest='bn').text if language == 'bn' else txt

        def map_digits(s): return s.translate(digit_map) if language == 'bn' else s

        def fmt(val, spec=".4f"):
            s = format(val, spec)
            return map_digits(s) if language == 'bn' else s

        # 3. Validate columns
        if not col1 or not col2:
            return JsonResponse({'success': False, 'error': 'Missing column names'})

        if col1 not in df.columns or col2 not in df.columns:
            return JsonResponse({'success': False, 'error': translate("One or both columns not found in the dataset.")})

        # 4. Filter numeric columns only
        numeric_df = df.select_dtypes(include=[np.number])
        if col1 not in numeric_df.columns or col2 not in numeric_df.columns:
            return JsonResponse({'success': False, 'error': translate("Please select numeric columns only.")})

        # 5. Prepare data vectors
        A = numeric_df[col1].dropna().to_numpy()
        B = numeric_df[col2].dropna().to_numpy()
        min_len = min(len(A), len(B))
        A, B = A[:min_len], B[:min_len]

        # 6. Compute metrics
        cos_sim         = float(np.dot(A, B) / (norm(A) * norm(B)))
        euclid_dist     = float(norm(A - B))
        manhattan_dist  = float(np.sum(np.abs(A - B)))
        chebyshev_dist  = float(np.max(np.abs(A - B)))
        p               = 3
        minkowski_dist  = float(np.sum(np.abs(A - B)**p)**(1/p))
        pearson_corr    = float(np.corrcoef(A, B)[0, 1])
        spearman_corr, _= spearmanr(A, B)

        result = {
            'cosine_similarity': fmt(cos_sim),
            'euclidean_distance': fmt(euclid_dist),
            'manhattan_distance': fmt(manhattan_dist),
            'chebyshev_distance': fmt(chebyshev_dist),
            'minkowski_distance': fmt(minkowski_dist),
            'pearson_correlation': fmt(pearson_corr),
            'spearman_correlation': fmt(spearman_corr),
            'p': map_digits(str(p)) if language == 'bn' else str(p),
        }

        heading = (
            f"Similarity / Distance between '{col1}' and '{col2}'"
            if language == 'en'
            else f"'{col1}' এবং '{col2}' এর মধ্যে সাদৃশ্য/দূরত্ব"
        )

        return JsonResponse({
            'success': True,
            'test': 'Similarity' if language == 'en' else 'সাদৃশ্য',
            'heading': heading,
            'results': result,
            'columns': [col1, col2]
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# def process_chi_square(request, df, selected_columns, user_id):
#     from scipy.stats import chi2_contingency
   

#     try:
#         # ── 1) Inputs ──────────────────────────────────────────────────────
#         lang = (request.POST.get("language", "en") or "en").lower()
#         is_bn = (lang == "bn")
#         palette = request.POST.get("palette", "viridis")
#         image_size = request.POST.get("image_size", "800x600")
#         try:
#             width, height = map(int, image_size.lower().split("x"))
#         except Exception:
#             width, height = 800, 600

#         # Variables to analyze
#         vars_list = selected_columns or []
#         if len(vars_list) < 2:
#             return JsonResponse({'success': False, 'error': 'Please select at least two categorical variables.'})
#         # ── 2) Paths & fonts 
#         plots_dir = os.path.join(settings.MEDIA_ROOT, f'ID_{user_id}_uploads', 'temporary_uploads', 'plots')
#         os.makedirs(plots_dir, exist_ok=True)
#         uid = str(uuid.uuid4())[:8]

#         # Try Bengali font (safe fallback if missing)
#         font_path = os.path.join(getattr(settings, "BASE_DIR", ""), 'NotoSansBengali-Regular.ttf')
#         try:
#             fm.fontManager.addfont(font_path)
#             bn_font_name = fm.FontProperties(fname=font_path).get_name()
#             mpl.rcParams['font.family'] = bn_font_name if is_bn else mpl.rcParams.get('font.family', 'sans-serif')
#         except Exception:
#             pass
#         # PIL font for title
#         try:
#             title_font = ImageFont.truetype(font_path, 50  if width >= 800 else 42)
#         except Exception:
#             title_font = ImageFont.load_default()

#         # Bengali digit mapping for any overlay text
#         digit_map_bn = str.maketrans('0123456789', '০১২৩৪৫৬৭৮৯')
#         def map_digits(s):
#             if not is_bn: return s
#             return str(s).translate(digit_map_bn)

#         # ── 3) Core helpers ────────────────────────────────────────────────
#         def fnum(x):
#             if x is None or (isinstance(x, float) and np.isnan(x)): return None
#             try: return float(x)
#             except Exception: return None

#         def inum(x):
#             if x is None or (isinstance(x, float) and np.isnan(x)): return None
#             try: return int(x)
#             except Exception:
#                 try: return int(float(x))
#                 except Exception: return None

#         def chi2_one(a, b, dropna=True, include_na_as_category=False, yates_for_2x2=True):
#             s1, s2 = df[a], df[b]
#             if include_na_as_category:
#                 s1, s2 = s1.fillna("__NA__"), s2.fillna("__NA__")
#             elif dropna:
#                 m = s1.notna() & s2.notna()
#                 s1, s2 = s1[m], s2[m] 
#             ct = pd.crosstab(s1, s2, dropna=False)
#             if ct.shape[0] == 0 or ct.shape[1] == 0:
#                 return dict(n=0, chi2=np.nan, dof=np.nan, p=np.nan)
#             yates = yates_for_2x2 and (ct.shape[0] == 2 and ct.shape[1] == 2)
#             chi2, p, dof, _ = chi2_contingency(ct, correction=yates)
#             return dict(n=int(ct.values.sum()), chi2=float(chi2), dof=int(dof), p=float(p))

#         def fdr_bh(pvals):
#             p = np.asarray(pvals, float)
#             n = p.size
#             order = np.argsort(p)
#             ranked = p[order]
#             adj = np.empty(n, float)
#             cm = 1.0
#             for i in range(n-1, -1, -1):
#                 cm = min(cm, ranked[i] * n / (i+1))
#                 adj[i] = cm
#             out = np.empty(n, float)
#             out[order] = np.minimum(adj, 1.0)
#             return out

        
#         def one_block(anchor, do_fdr=True):
#             others = [c for c in vars_list if c != anchor]
#             rows = []
#             for o in others:
#                 r = chi2_one(anchor, o, dropna=True, include_na_as_category=False, yates_for_2x2=True)
#                 rows.append({
#                     "variable1": anchor,
#                     "variable2": o,
#                     "chi2": fnum(r["chi2"]),
#                     "p_value": fnum(r["p"]),
#                     "dof": inum(r["dof"]),
#                     "n": inum(r["n"]),
#                 })

#             if do_fdr and any(row["p_value"] is not None for row in rows):
#                 ps = np.array([row["p_value"] if row["p_value"] is not None else np.nan for row in rows], float)
#                 mask = ~np.isnan(ps)
#                 padj = np.full(len(rows), np.nan, float)
#                 if mask.any():
#                     padj[mask] = fdr_bh(ps[mask])
#                 for i, row in enumerate(rows):
#                     row["_p_adj_tmp"] = float(padj[i]) if not np.isnan(padj[i]) else None
#                 rows.sort(key=lambda r: (r["_p_adj_tmp"] is None, r["_p_adj_tmp"] if r["_p_adj_tmp"] is not None else 1.0))
#                 for row in rows:
#                     row.pop("_p_adj_tmp", None)
#             else:
#                 rows.sort(key=lambda r: (r["p_value"] is None, r["p_value"] if r["p_value"] is not None else 1.0))
#             return rows

#         blocks, summary_rows = [], []
#         for anchor in vars_list:
#             block_rows = one_block(anchor, do_fdr=True)
#             blocks.append({"anchor": anchor, "rows": block_rows})
#             summary_rows.extend(block_rows)

        
#         n = len(vars_list)
#         P = np.ones((n, n), dtype=float)
#         for i, a in enumerate(vars_list):
#             for j in range(i+1, n):
#                 b = vars_list[j]
#                 res = chi2_one(a, b, dropna=True, include_na_as_category=False, yates_for_2x2=True)
#                 P[i, j] = P[j, i] = res["p"] if res["p"] is not None else 1.0
#         np.fill_diagonal(P, 1.0)

        

#         fig, ax = plt.subplots(figsize=(width/100, height/100), dpi=100)
#         sns.heatmap(
#             P,
#             vmin=0, vmax=1,
#             cmap="coolwarm",
#             annot=np.round(P, 2),
#             fmt=".2f",
#             square=True,
#             linewidths=0.75,
#             linecolor="white",
#             cbar_kws={"label": "p-value"}
#         )
#         ax.set_xticks(np.arange(n) + 0.5)
#         ax.set_yticks(np.arange(n) + 0.5)
#         ax.set_xticklabels(vars_list, rotation=45, ha="right", fontsize=10)
#         ax.set_yticklabels(vars_list, rotation=0, fontsize=10)
#         ax.set_xlabel("")
#         ax.set_ylabel("")
#         plt.tight_layout()

#         base_name = f"chi2_pairwise_heatmap_base_{uid}.png"
#         final_name = f"chi2_pairwise_heatmap_{uid}.png"
#         base_path = os.path.join(plots_dir, base_name)
#         final_path = os.path.join(plots_dir, final_name)
#         plt.savefig(base_path, dpi=300, bbox_inches="tight")
#         plt.close(fig)

        
#         base_img = Image.open(base_path).convert("RGB")
#         bw, bh = base_img.size
#         pad = 16
#         title_text = ("p-value pairwise chi-square heatmap"
#                     if lang != "bn"
#                     else "p-value জোড়াভিত্তিক কাই-স্কয়ার হিটম্যাপ")
#         try:
#             title_font = ImageFont.truetype(font_path, 50 if width >= 800 else 28)
#         except Exception:
#             title_font = ImageFont.load_default()

        
#         # measure
#         try:
#             tx0, ty0, tx1, ty1 = title_font.getbbox(title_text)
#             th = ty1 - ty0
#             tw = tx1 - tx0
#         except Exception:
#             th, tw = 40, 400

#         top_margin = th + 2*pad
#         canvas = Image.new("RGB", (bw, bh + top_margin), "white")
#         canvas.paste(base_img, (0, top_margin))
#         draw2 = ImageDraw.Draw(canvas)
#         tx = max(0, (bw - tw) // 2)
#         ty = max(0, (top_margin - th) // 2)
#         draw2.text((tx, ty), title_text, fill="black", font=title_font)
#         canvas.save(final_path, format="PNG")
#         try:
#             os.remove(base_path)
#         except Exception:
#             pass

#         # Public URL for React
#         base_url = os.path.join(settings.MEDIA_URL, f'ID_{user_id}_uploads', 'temporary_uploads', 'plots')
#         heatmap_url = os.path.join(base_url, final_name)


#         # ── 6) Table columns for React ─────────────────────────────────────
#         if is_bn:
#             table_columns = [
#                 {"key": "variable1", "label": "ভেরিয়েবল ১"},
#                 {"key": "variable2", "label": "ভেরিয়েবল ২"},
#                 {"key": "chi2",      "label": "কাই-স্কয়ার পরিসংখ্যান"},
#                 {"key": "p_value",   "label": "পি-মান"},
#                 {"key": "dof",       "label": "স্বাধীনতার মাত্রা"},
#                 {"key": "n",         "label": "নমুনা"},
#             ]
#         else:
#             table_columns = [
#                 {"key": "variable1", "label": "Variable 1"},
#                 {"key": "variable2", "label": "Variable 2"},
#                 {"key": "chi2",      "label": "Chi-square statistic"},
#                 {"key": "p_value",   "label": "P-value"},
#                 {"key": "dof",       "label": "DoF"},
#                 {"key": "n",         "label": "N"},
#             ]

#         # ── 7) Response (no heatmap payload; image URL only) ───────────────
#         return JsonResponse({
#             "success": True, 
#             "variables": vars_list,
#             "table_columns": table_columns,
#             "summary_rows": summary_rows,  
#             "blocks": blocks,             
#             "image_path": heatmap_url     
#         })

#     except Exception as e:
#         return JsonResponse({"success": False, "error": str(e)})


def process_chi_square(request, df, selected_columns, user_id):
    from scipy.stats import chi2_contingency
    import numpy as np
    import pandas as pd
    from django.http import JsonResponse

    try:
        # ── 1) Extract inputs ──────────────────────────────────────────────
        lang = (request.POST.get("language", "en") or "en").lower()
        if lang not in ('en', 'bn'):
            lang = 'en'
        
        is_bn = (lang == "bn")

        # Variables to analyze
        vars_list = selected_columns or []
        if len(vars_list) < 2:
            return JsonResponse({
                'success': False, 
                'error': 'Please select at least two categorical variables.'
            })

        print(f"[Chi-square] Analyzing {len(vars_list)} variables: {vars_list}")

        # ── 2) Core helper functions ───────────────────────────────────────
        def fnum(x):
            """Convert to float, handling None and NaN"""
            if x is None or (isinstance(x, float) and np.isnan(x)):
                return None
            try:
                return float(x)
            except Exception:
                return None

        def inum(x):
            """Convert to int, handling None and NaN"""
            if x is None or (isinstance(x, float) and np.isnan(x)):
                return None
            try:
                return int(x)
            except Exception:
                try:
                    return int(float(x))
                except Exception:
                    return None

        def chi2_one_pair(var_a, var_b, dropna=True, include_na_as_category=False, yates_for_2x2=True):
            """
            Perform chi-square test for one pair of variables.
            Returns dict with test results and contingency table data.
            """
            s1, s2 = df[var_a].copy(), df[var_b].copy()
            
            # Handle missing values
            if include_na_as_category:
                s1 = s1.fillna("__NA__")
                s2 = s2.fillna("__NA__")
            elif dropna:
                mask = s1.notna() & s2.notna()
                s1, s2 = s1[mask], s2[mask]
            
            # Create contingency table
            ct = pd.crosstab(s1, s2, dropna=False)
            
            if ct.shape[0] == 0 or ct.shape[1] == 0:
                return {
                    'n': 0,
                    'chi2': None,
                    'dof': None,
                    'p': None,
                    'contingency_table': None,
                    'expected_frequencies': None,
                    'residuals': None,
                    'categories_var1': [],
                    'categories_var2': [],
                    'var1_categories': [],
                    'var2_categories': [],
                    'row_totals': [],
                    'col_totals': [],
                    'row_percentages': [],
                    'col_percentages': [],
                    'stacked_data': []
                }
            
            # Apply Yates' correction for 2x2 tables
            yates = yates_for_2x2 and (ct.shape[0] == 2 and ct.shape[1] == 2)
            
            try:
                chi2, p, dof, expected = chi2_contingency(ct, correction=yates)
            except Exception as e:
                print(f"[Chi-square] Error in chi2_contingency: {e}")
                return {
                    'n': 0,
                    'chi2': None,
                    'dof': None,
                    'p': None,
                    'contingency_table': None,
                    'expected_frequencies': None,
                    'residuals': None,
                    'categories_var1': [],
                    'categories_var2': [],
                    'var1_categories': [],
                    'var2_categories': [],
                    'row_totals': [],
                    'col_totals': [],
                    'row_percentages': [],
                    'col_percentages': [],
                    'stacked_data': []
                }
            
            # Calculate standardized residuals
            residuals = (ct.values - expected) / np.sqrt(expected + 1e-10)  # Add small value to avoid division by zero
            
            # Get category names (convert to string for JSON serialization)
            categories_var1 = [str(c) for c in ct.index.tolist()]
            categories_var2 = [str(c) for c in ct.columns.tolist()]
            
            # Calculate row and column totals
            row_totals = ct.sum(axis=1).tolist()
            col_totals = ct.sum(axis=0).tolist()
            
            # Calculate percentages
            row_percentages = (ct.div(ct.sum(axis=1), axis=0) * 100).values.tolist()
            col_percentages = (ct.div(ct.sum(axis=0), axis=1) * 100).values.tolist()
            
            # Prepare data for stacked bar chart (rows of var1, columns of var2)
            stacked_data = []
            for i, cat1 in enumerate(categories_var1):
                row_data = {
                    'category': str(cat1),  # This will be the x-axis
                    'total': int(row_totals[i])
                }
                
                # Add both count and percentage for each category of var2
                for j, cat2 in enumerate(categories_var2):
                    count = int(ct.values[i][j])
                    percentage = float(row_percentages[i][j]) if not np.isnan(row_percentages[i][j]) else 0.0
                    
                    # Store both count and percentage with consistent naming
                    row_data[str(cat2)] = percentage  # This is what Recharts uses for bar height (percentage)
                    row_data[f'{cat2}_count'] = count  # This is the actual count
                
                stacked_data.append(row_data)
            
            # Convert to Python native types for JSON serialization
            return {
                'n': int(ct.values.sum()),
                'chi2': float(chi2),
                'dof': int(dof),
                'p': float(p),
                'contingency_table': [[int(val) if not np.isnan(val) else 0 for val in row] for row in ct.values.tolist()],
                'expected_frequencies': [[float(val) if not np.isnan(val) else 0.0 for val in row] for row in expected.tolist()],
                'residuals': [[float(val) if not np.isnan(val) else 0.0 for val in row] for row in residuals.tolist()],
                'categories_var1': categories_var1,
                'categories_var2': categories_var2,
                'var1_categories': categories_var1,  # Duplicate for backward compatibility
                'var2_categories': categories_var2,  # Duplicate for backward compatibility
                'row_totals': [int(val) for val in row_totals],
                'col_totals': [int(val) for val in col_totals],
                'row_percentages': [[float(val) if not np.isnan(val) else 0.0 for val in row] for row in row_percentages],
                'col_percentages': [[float(val) if not np.isnan(val) else 0.0 for val in row] for row in col_percentages],
                'stacked_data': stacked_data  # New field for easy Recharts integration
            }

        def fdr_bh(pvals):
            """Benjamini-Hochberg FDR correction"""
            p = np.asarray(pvals, float)
            n = p.size
            if n == 0:
                return np.array([])
            order = np.argsort(p)
            ranked = p[order]
            adj = np.empty(n, float)
            cm = 1.0
            for i in range(n-1, -1, -1):
                cm = min(cm, ranked[i] * n / (i+1))
                adj[i] = cm
            out = np.empty(n, float)
            out[order] = np.minimum(adj, 1.0)
            return out

        # ── 3) Compute all pairwise tests ──────────────────────────────────
        n_vars = len(vars_list)
        pairwise_results = []
        all_p_values = []
        
        for i in range(n_vars):
            for j in range(i + 1, n_vars):
                var1 = vars_list[i]
                var2 = vars_list[j]
                
                # print(f"[Chi-square] Testing {var1} vs {var2}")
                
                result = chi2_one_pair(
                    var1, var2,
                    dropna=True,
                    include_na_as_category=False,
                    yates_for_2x2=True
                )
                
                pairwise_results.append({
                    'variable1': str(var1),
                    'variable2': str(var2),
                    'chi2': fnum(result['chi2']),
                    'p_value': fnum(result['p']),
                    'dof': inum(result['dof']),
                    'n': inum(result['n']),
                    'contingency_table': result['contingency_table'],
                    'expected_frequencies': result['expected_frequencies'],
                    'residuals': result['residuals'],
                    'categories_var1': result['categories_var1'],
                    'categories_var2': result['categories_var2'],
                    'var1_categories': result['var1_categories'],  # For Stacked Bar & Mosaic plots
                    'var2_categories': result['var2_categories'],  # For Stacked Bar & Mosaic plots
                    'row_totals': result['row_totals'],
                    'col_totals': result['col_totals'],
                    'row_percentages': result['row_percentages'],
                    'col_percentages': result['col_percentages']
                })
                
                if result['p'] is not None:
                    all_p_values.append(result['p'])
                else:
                    all_p_values.append(np.nan)

        print(f"[Chi-square] Completed {len(pairwise_results)} pairwise tests")

        # ── 4) Apply FDR correction ────────────────────────────────────────
        p_array = np.array(all_p_values, dtype=float)
        mask = ~np.isnan(p_array)
        
        if mask.any():
            adjusted_p = np.full(len(pairwise_results), np.nan, dtype=float)
            adjusted_p[mask] = fdr_bh(p_array[mask])
            
            for i, result in enumerate(pairwise_results):
                result['p_adjusted'] = float(adjusted_p[i]) if not np.isnan(adjusted_p[i]) else None
        else:
            for result in pairwise_results:
                result['p_adjusted'] = None

        # Sort by p-value (ascending)
        pairwise_results.sort(
            key=lambda r: (r['p_value'] is None, r['p_value'] if r['p_value'] is not None else 1.0)
        )

        # ── 5) Create p-value matrix for heatmap ───────────────────────────
        p_matrix = np.ones((n_vars, n_vars), dtype=float)
        chi2_matrix = np.zeros((n_vars, n_vars), dtype=float)
        p_adjusted_matrix = np.ones((n_vars, n_vars), dtype=float)
        
        for result in pairwise_results:
            i = vars_list.index(result['variable1'])
            j = vars_list.index(result['variable2'])
            
            p_val = result['p_value'] if result['p_value'] is not None else 1.0
            p_matrix[i, j] = p_val
            p_matrix[j, i] = p_val
            
            chi2_val = result['chi2'] if result['chi2'] is not None else 0.0
            chi2_matrix[i, j] = chi2_val
            chi2_matrix[j, i] = chi2_val
            
            p_adj_val = result['p_adjusted'] if result['p_adjusted'] is not None else 1.0
            p_adjusted_matrix[i, j] = p_adj_val
            p_adjusted_matrix[j, i] = p_adj_val
        
        np.fill_diagonal(p_matrix, 1.0)
        np.fill_diagonal(chi2_matrix, 0.0)
        np.fill_diagonal(p_adjusted_matrix, 1.0)

        # ── 6) Prepare variable-level statistics ──────────────────────────
        variable_stats = []
        
        for var in vars_list:
            col_data = df[var].dropna()
            value_counts = col_data.value_counts()
            
            variable_stats.append({
                'variable': str(var),
                'n_categories': int(len(value_counts)),
                'n_observations': int(len(col_data)),
                'n_missing': int(df[var].isna().sum()),
                'categories': [str(c) for c in value_counts.index.tolist()],
                'frequencies': [int(f) for f in value_counts.values.tolist()],
                'percentages': [float(p) for p in (value_counts / len(col_data) * 100).values.tolist()]
            })

        # ── 7) Create blocks (one anchor variable at a time) ──────────────
        # ── 7) Create blocks (one anchor variable at a time) ──────────────
        blocks = []
        
        for anchor in vars_list:
            block_results = []
            
            for result in pairwise_results:
                if result['variable1'] == anchor:
                    block_results.append(result)
                elif result['variable2'] == anchor:
                    # Swap variables so anchor is always first
                    # IMPORTANT: Must transpose contingency table and expected frequencies
                    swapped_result = {
                        'variable1': result['variable2'],
                        'variable2': result['variable1'],
                        'chi2': result['chi2'],
                        'p_value': result['p_value'],
                        'p_adjusted': result['p_adjusted'],
                        'dof': result['dof'],
                        'n': result['n'],
                        # Transpose the tables
                        'contingency_table': [[result['contingency_table'][j][i] for j in range(len(result['contingency_table']))] 
                                             for i in range(len(result['contingency_table'][0]))] if result['contingency_table'] else None,
                        'expected_frequencies': [[result['expected_frequencies'][j][i] for j in range(len(result['expected_frequencies']))] 
                                                for i in range(len(result['expected_frequencies'][0]))] if result['expected_frequencies'] else None,
                        'residuals': [[result['residuals'][j][i] for j in range(len(result['residuals']))] 
                                     for i in range(len(result['residuals'][0]))] if result['residuals'] else None,
                        # Swap categories
                        'categories_var1': result['categories_var2'],
                        'categories_var2': result['categories_var1'],
                        'var1_categories': result['var2_categories'],
                        'var2_categories': result['var1_categories'],
                        # Swap totals
                        'row_totals': result['col_totals'],
                        'col_totals': result['row_totals'],
                        # Transpose percentages
                        'row_percentages': [[result['col_percentages'][j][i] for j in range(len(result['col_percentages']))] 
                                           for i in range(len(result['col_percentages'][0]))] if result['col_percentages'] else None,
                        'col_percentages': [[result['row_percentages'][j][i] for j in range(len(result['row_percentages']))] 
                                           for i in range(len(result['row_percentages'][0]))] if result['row_percentages'] else None,
                    }
                    block_results.append(swapped_result)
            
            blocks.append({
                'anchor': str(anchor),
                'results': block_results
            })


        # ── 8) Prepare table columns for frontend ─────────────────────────
        if is_bn:
            table_columns = [
                {"key": "variable1", "label": "ভেরিয়েবল ১"},
                {"key": "variable2", "label": "ভেরিয়েবল ২"},
                {"key": "chi2", "label": "কাই-স্কয়ার পরিসংখ্যান"},
                {"key": "p_value", "label": "পি-মান"},
                {"key": "p_adjusted", "label": "সমন্বিত পি-মান"},
                {"key": "dof", "label": "স্বাধীনতার মাত্রা"},
                {"key": "n", "label": "নমুনা"},
            ]
        else:
            table_columns = [
                {"key": "variable1", "label": "Variable 1"},
                {"key": "variable2", "label": "Variable 2"},
                {"key": "chi2", "label": "Chi² Statistic"},
                {"key": "p_value", "label": "P-value"},
                {"key": "p_adjusted", "label": "Adjusted P-value"},
                {"key": "dof", "label": "DoF"},
                {"key": "n", "label": "N"},
            ]

        # ── 9) Prepare plot data for visualizations ───────────────────────
        # Prepare data for heatmap and other plots
        plot_data = []
        for i, var1 in enumerate(vars_list):
            for j, var2 in enumerate(vars_list):
                if i != j:
                    # Find the corresponding result
                    result = None
                    for r in pairwise_results:
                        if (r['variable1'] == var1 and r['variable2'] == var2) or \
                           (r['variable1'] == var2 and r['variable2'] == var1):
                            result = r
                            break
                    
                    if result:
                        plot_data.append({
                            'category': f"{var1}",
                            'variable': f"{var2}",
                            'p_value': result['p_value'],
                            'chi2': result['chi2'],
                            'p_adjusted': result['p_adjusted']
                        })

        # ── 10) Prepare final response ─────────────────────────────────────
        test_name = 'Chi-square Test of Independence' if lang == 'en' else 'কাই-স্কয়ার স্বাধীনতা পরীক্ষা'
        
        response_data = {
            'success': True,
            'test': test_name,
            'language': lang,
            'n_variables': int(n_vars),
            'n_comparisons': int(len(pairwise_results)),
            'variables': [str(v) for v in vars_list],
            'variable_stats': variable_stats,
            'pairwise_results': pairwise_results,
            'blocks': blocks,
            'p_value_matrix': [[float(val) for val in row] for row in p_matrix.tolist()],
            'chi2_matrix': [[float(val) for val in row] for row in chi2_matrix.tolist()],
            'p_adjusted_matrix': [[float(val) for val in row] for row in p_adjusted_matrix.tolist()],
            'plot_data': plot_data,
            'table_columns': table_columns,
            'metadata': {
                'fdr_correction': 'Benjamini-Hochberg',
                'yates_correction': True,
                'alpha': 0.05,
                'total_observations': int(len(df))
            }
        }

        print(f"[Chi-square] Response prepared successfully with {len(pairwise_results)} comparisons")
        return JsonResponse(response_data)

    except Exception as e:
        import traceback
        error_msg = str(e)
        traceback_msg = traceback.format_exc()
        print(f"[Chi-square] Error: {error_msg}")
        print(f"[Chi-square] Traceback: {traceback_msg}")
        return JsonResponse({
            'success': False, 
            'error': error_msg,
            'traceback': traceback_msg
        })


# def process_cramers_heatmap(request, df, user_id):
#     from django.http import JsonResponse
#     from django.conf import settings
#     import pandas as pd
#     import seaborn as sns
#     import matplotlib.pyplot as plt
#     import matplotlib as mpl
#     import matplotlib.font_manager as fm
#     from googletrans import Translator
#     from PIL import Image, ImageDraw, ImageFont
#     import os
#     import uuid

#     try:
#         # --- 1) Inputs ---
#         x_col = request.POST.get("column1")
#         y_col = request.POST.get("column2")
#         lang = request.POST.get("language", "en")
#         fmt = request.POST.get("format", "png").lower()
#         pil_fmt = {"png": "PNG", "jpg": "JPEG", "jpeg": "JPEG", "pdf": "PDF", "tiff": "TIFF"}.get(fmt, "PNG")

#         use_default = request.POST.get("use_default", "true") == "true"
#         label_font_size = int(request.POST.get("label_font_size", 36))
#         tick_font_size = int(request.POST.get("tick_font_size", 16))
#         img_quality = int(request.POST.get("image_quality", 90))
#         palette = request.POST.get("palette", "deep")
#         image_size = request.POST.get("image_size", "800x600")
#         try:
#             width, height = map(int, image_size.lower().split("x"))
#         except:
#             width, height = 800, 600

#         # --- 2) Font & Path setup ---
#         font_path = os.path.join(settings.BASE_DIR, 'NotoSansBengali-Regular.ttf')
#         fm.fontManager.addfont(font_path)
#         bengali_font = fm.FontProperties(fname=font_path).get_name()
#         mpl.rcParams['font.family'] = bengali_font
#         tick_prop = fm.FontProperties(fname=font_path, size=tick_font_size) if lang == 'bn' else fm.FontProperties(size=tick_font_size)
#         label_font = ImageFont.truetype(font_path, size=label_font_size)

#         translator = Translator()
#         def translate(txt): return translator.translate(txt, dest='bn').text if lang == 'bn' else txt
#         def map_digits(txt): return txt.translate(str.maketrans('0123456789', '০১২৩৪৫৬৭৮৯')) if lang == 'bn' else txt

#         plots_dir = os.path.join(settings.MEDIA_ROOT, f'ID_{user_id}_uploads', 'temporary_uploads', 'plots')
#         os.makedirs(plots_dir, exist_ok=True)
#         uid = str(uuid.uuid4())[:8]

#         # --- 3) Crosstab ---
#         ct = pd.crosstab(df[x_col], df[y_col])
#         annot_df = ct.applymap(lambda v: map_digits(translate(str(int(v)))))

#         # --- 4) Heatmap ---
#         fig, ax = plt.subplots(figsize=(width/100, height/100), dpi=100)
#         cmap = sns.color_palette(palette, as_cmap=True)
#         hm = sns.heatmap(ct, annot=annot_df, fmt="", cmap=cmap, ax=ax)

#         ax.set_xlabel('')
#         ax.set_ylabel('')
#         ax.set_title('')

#         # Translate axis labels
#         ax.set_xticklabels([map_digits(translate(str(t))) for t in ct.columns], fontproperties=tick_prop)
#         ax.set_yticklabels([map_digits(translate(str(t))) for t in ct.index], fontproperties=tick_prop)

#         # Colorbar
#         cbar = hm.collections[0].colorbar
#         cb_vals = cbar.get_ticks()
#         cbar.set_ticklabels([map_digits(translate(str(int(v)))) for v in cb_vals])
#         cbar.ax.tick_params(labelsize=tick_font_size)

#         # Save with PIL label wrapping
#         base_path = os.path.join(plots_dir, f"cramers_heatmap_base_{uid}.png")
#         final_path = os.path.join(plots_dir, f"cramers_heatmap_{uid}.{fmt}")
#         fig.savefig(base_path, dpi=300, bbox_inches='tight')
#         plt.close(fig)

#         T = translate("Observed Counts Heatmap")
#         X = translate(x_col)
#         Y = translate(y_col)
#         bw, bh = Image.open(base_path).convert("RGB").size
#         tx = label_font.getbbox(T)[3]
#         xx = label_font.getbbox(X)[3]
#         yy = label_font.getbbox(Y)[2]
#         pad = label_font_size // 2

#         W = bw + yy + pad * 2
#         H = bh + tx + xx + pad * 2
#         canvas = Image.new("RGB", (W, H), "white")
#         img = Image.open(base_path).convert("RGB")
#         canvas.paste(img, (yy + pad, tx + pad))
#         draw = ImageDraw.Draw(canvas)
#         draw.text(((W - draw.textlength(T, font=label_font)) // 2, pad), T, font=label_font, fill='black')
#         draw.text(((W - draw.textlength(X, font=label_font)) // 2, tx + bh + pad), X, font=label_font, fill='black')
#         Yimg = Image.new("RGBA", label_font.getbbox(Y)[2:], (255,255,255,0))
#         ImageDraw.Draw(Yimg).text((0, 0), Y, font=label_font, fill='black')
#         Yrot = Yimg.rotate(90, expand=True)
#         canvas.paste(Yrot, (pad, tx + (bh - Yrot.size[1]) // 2), Yrot)
#         canvas.save(final_path, format=pil_fmt, quality=img_quality)

#         return JsonResponse({
#             'success': True,
#             'image_paths': [os.path.join(settings.MEDIA_URL, f'ID_{user_id}_uploads', 'temporary_uploads', 'plots', os.path.basename(final_path))],
#             'columns': [x_col, y_col],
#         })

#     except Exception as e:
#         return JsonResponse({'success': False, 'error': str(e)})

# def process_network_graph(request, df, selected_columns, user_id): 
#     import os
#     import random
#     from django.http import JsonResponse
#     from django.conf import settings
#     import matplotlib.pyplot as plt
#     import matplotlib.font_manager as fm
#     from PIL import Image, ImageDraw, ImageFont
#     import networkx as nx
#     import numpy as np
#     import re

#     try:
#         language = request.POST.get("language", "en").strip().lower()
#         node_color = request.POST.get("node_color", "#AED6F1")
#         node_size = int(request.POST.get("node_size", 800))
#         text_size = int(request.POST.get("text_size", 25))
#         text_color = request.POST.get("text_color", "black")
#         edge_width_factor = float(request.POST.get("edge_width_factor", 0.5))
#         show_edge_weights = request.POST.get("show_edge_weights", "n").lower() == "y"
#         weight_font_size = int(request.POST.get("weight_font_size", 3)) if show_edge_weights else 0
#         weight_color = request.POST.get("weight_color", "red") if show_edge_weights else "red"
#         use_matrix = request.POST.get("use_matrix", "n").lower() == "y"

#         digit_map = str.maketrans("0123456789", "০১২৩৪৫৬৭৮৯")
#         root_bn = "নোড"

#         print("Selected columns for network graph:", selected_columns)



#         def to_bn(lbl):
#             # Extract number part at the end of the label (e.g. Node12 → 12 → ১২)
#             match = re.search(r'(\d+)$', lbl)
#             if match:
#                 number_part = match.group(1).translate(digit_map)
#                 return root_bn + number_part
#             return lbl  # fallback: no change if no digits found



#         def to_label(lbl): return to_bn(lbl) if language == 'bengali' else lbl

#         # Create graph
#         if use_matrix:
#             if 'matrix_file' not in request.FILES:
#                 raise ValueError("Matrix file not uploaded.")
#             matrix_file = request.FILES['matrix_file']
#             df_matrix = pd.read_csv(matrix_file, index_col=0)
#             if df_matrix.shape[0] != df_matrix.shape[1]:
#                 raise ValueError("Adjacency matrix must be square.")
#             mat = df_matrix.values
#             nodes = list(df_matrix.index)
#             G = nx.Graph()
#             G.add_nodes_from(nodes)
#             for i in range(len(nodes)):
#                 for j in range(i+1, len(nodes)):
#                     w = mat[i, j]
#                     if w != 0:
#                         G.add_edge(nodes[i], nodes[j], weight=w * edge_width_factor)
#         else:
#             num_nodes = 30
#             nbr_count = 3
#             nodes = [f"Node{i}" for i in range(1, num_nodes + 1)]
#             G = nx.Graph()
#             G.add_nodes_from(nodes)
#             for node in nodes:
#                 nbrs = random.sample([n for n in nodes if n != node], k=nbr_count)
#                 for nbr in nbrs:
#                     G.add_edge(node, nbr, weight=edge_width_factor)

#         edge_widths = [d['weight'] for _, _, d in G.edges(data=True)]
#         raw_edge_labels = {}
#         for u, v, d in G.edges(data=True):
#             raw = f"{d['weight']:.2f}"
#             if language == 'bengali':
#                 raw = raw.translate(digit_map)
#             raw_edge_labels[(u, v)] = raw

#         # Layout and draw
#         dpi = 300
#         fig, ax = plt.subplots(figsize=(8, 8), dpi=dpi)
#         pos = nx.spring_layout(G, seed=42, k=0.8, iterations=200)
#         nx.draw_networkx_nodes(G, pos, node_size=node_size, node_color=node_color, ax=ax)
#         nx.draw_networkx_edges(G, pos, width=edge_widths, alpha=0.6, edge_color="gray", ax=ax)
#         ax.margins(0.10)
#         ax.axis("off")
#         fig.canvas.draw()

#         renderer = fig.canvas.get_renderer()
#         bbox = fig.get_tightbbox(renderer)
#         crop_x0_px = int(bbox.x0 * dpi)
#         crop_y0_px = int(bbox.y0 * dpi)
#         final_W_px = int(bbox.width * dpi)
#         final_H_px = int(bbox.height * dpi)
#         data_to_px = ax.transData.transform
#         def data_to_image_px(xy):
#             px, py = data_to_px(xy)
#             return (int(px - crop_x0_px), int(final_H_px - (py - crop_y0_px)))

#         # Save base
#         plots_dir = os.path.join(settings.MEDIA_ROOT,f'ID_{user_id}_uploads', 'temporary_uploads', 'plots')
#         os.makedirs(plots_dir, exist_ok=True)
#         base_path = os.path.join(plots_dir, 'network_base.png')
#         fig.savefig(base_path, dpi=dpi, bbox_inches="tight", pad_inches=0.05)
#         plt.close(fig)

#         # Overlay labels
#         img = Image.open(base_path).convert("RGBA")
#         draw = ImageDraw.Draw(img)
#         font_path = os.path.join(settings.BASE_DIR, 'NotoSansBengali-Regular.ttf')
#         try:
#             font_label = ImageFont.truetype(font_path, size=text_size)
#         except:
#             font_label = ImageFont.load_default()

#         for node in G.nodes():
#             label = to_label(node)
#             px, py = data_to_image_px(pos[node])
#             w, h = draw.textbbox((0, 0), label, font=font_label)[2:]
#             draw.text((px - w/2, py - h/2), label, font=font_label, fill=text_color)

#         if show_edge_weights:
#             for (u, v), raw in raw_edge_labels.items():
#                 midx = (pos[u][0] + pos[v][0]) / 2
#                 midy = (pos[u][1] + pos[v][1]) / 2
#                 px, py = data_to_image_px((midx, midy))
#                 draw.text((px, py), raw, font=font_label, fill=weight_color, anchor="mm")

#         labeled_path = os.path.join(plots_dir, 'network_labeled.png')
#         img.save(labeled_path)

#         return JsonResponse({
#             'success': True,
#             'image_path': os.path.join(settings.MEDIA_URL, f'ID_{user_id}_uploads', 'temporary_uploads','plots', 'network_labeled.png'),
#             'message': "Network graph generated successfully"
#         })
#     except Exception as e:
#         return JsonResponse({'success': False, 'error': str(e)})

def process_network_graph(request, df, selected_columns, user_id): 
    import os
    import random
    import pandas as pd
    from django.http import JsonResponse
    from django.conf import settings
    import matplotlib.pyplot as plt
    import matplotlib.font_manager as fm
    from PIL import Image, ImageDraw, ImageFont
    import networkx as nx
    import numpy as np
    import re
    from scipy.stats import pearsonr
    from sklearn.metrics.pairwise import cosine_similarity
    import warnings
    warnings.filterwarnings('ignore')

    try:
        # Get parameters from request
        language = request.POST.get("language", "en").strip().lower()
        node_color = request.POST.get("node_color", "#AED6F1")
        node_size = int(request.POST.get("node_size", 800))
        text_size = int(request.POST.get("text_size", 25))
        text_color = request.POST.get("text_color", "black")
        edge_width_factor = float(request.POST.get("edge_width_factor", 0.5))
        show_edge_weights = request.POST.get("show_edge_weights", "n").lower() == "y"
        weight_font_size = int(request.POST.get("weight_font_size", 10)) if show_edge_weights else 0
        weight_color = request.POST.get("weight_color", "red") if show_edge_weights else "red"
        use_matrix = request.POST.get("use_matrix", "n").lower() == "y"
        
        # Additional parameters for better visualization
        correlation_method = request.POST.get("correlation_method", "pearson")  # pearson, spearman, cosine
        correlation_threshold = float(request.POST.get("correlation_threshold", 0.3))  # Minimum correlation to show edge
        max_nodes = int(request.POST.get("max_nodes", 30))  # Limit number of nodes for performance

        print("Selected columns for network graph:", selected_columns)

        digit_map = str.maketrans("0123456789", "০১২৩৪৫৬৭৮৯")
        root_bn = "নোড"

        def to_bn(lbl):
            # For actual column names, translate digits but keep the original text
            return str(lbl).translate(digit_map) if language == 'bengali' else str(lbl)

        def to_label(lbl): 
            return to_bn(lbl) if language == 'bengali' else str(lbl)

        def calculate_correlation(col1, col2, method='pearson'):
            """Calculate correlation between two columns"""
            # Remove NaN values and ensure numeric data
            try:
                col1_clean = pd.to_numeric(col1, errors='coerce')
                col2_clean = pd.to_numeric(col2, errors='coerce')
                
                mask = ~(pd.isna(col1_clean) | pd.isna(col2_clean))
                if mask.sum() < 3:  # Need at least 3 valid pairs
                    return 0
                
                clean_col1 = col1_clean[mask]
                clean_col2 = col2_clean[mask]
                
                if method == 'pearson':
                    corr, _ = pearsonr(clean_col1, clean_col2)
                elif method == 'spearman':
                    from scipy.stats import spearmanr
                    corr, _ = spearmanr(clean_col1, clean_col2)
                elif method == 'cosine':
                    # Reshape for cosine similarity
                    cos_sim = cosine_similarity([clean_col1], [clean_col2])[0, 0]
                    corr = cos_sim
                else:
                    corr, _ = pearsonr(clean_col1, clean_col2)
                
                return abs(corr) if not np.isnan(corr) else 0
            except:
                return 0

        def clean_column_name(col_name):
            """Clean and shorten column names for better display"""
            name = str(col_name)
            # Replace underscores with spaces
            name = name.replace('_', ' ')
            # Capitalize first letter of each word
            name = ' '.join(word.capitalize() for word in name.split())
            # Truncate if too long
            if len(name) > 20:
                name = name[:17] + "..."
            return name

        # Create graph
        if use_matrix:
            if 'matrix_file' not in request.FILES:
                raise ValueError("Matrix file not uploaded.")
            matrix_file = request.FILES['matrix_file']
            df_matrix = pd.read_csv(matrix_file, index_col=0)
            if df_matrix.shape[0] != df_matrix.shape[1]:
                raise ValueError("Adjacency matrix must be square.")
            mat = df_matrix.values
            nodes = list(df_matrix.index)[:max_nodes]
            G = nx.Graph()
            G.add_nodes_from(nodes)
            for i in range(len(nodes)):
                for j in range(i+1, len(nodes)):
                    w = mat[i, j]
                    if abs(w) >= correlation_threshold:
                        G.add_edge(nodes[i], nodes[j], weight=abs(w) * edge_width_factor)
        else:
            # Use actual data from DataFrame with selected columns
            if df is None or df.empty:
                raise ValueError("No data provided for network analysis.")
            
            # Determine which columns to analyze
            if selected_columns and len(selected_columns) > 0:
                # Use selected columns
                available_columns = [col for col in selected_columns if col in df.columns]
                if not available_columns:
                    raise ValueError("Selected columns not found in the data.")
                df_analysis = df[available_columns]
                print(f"Using selected columns: {available_columns}")
            else:
                # Use all numeric columns if no selection
                numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
                if not numeric_cols:
                    raise ValueError("No numeric columns found in the data for correlation analysis.")
                df_analysis = df[numeric_cols]
                print(f"No columns selected. Using all numeric columns: {numeric_cols}")
            
            # Limit number of columns for performance
            if len(df_analysis.columns) > max_nodes:
                df_analysis = df_analysis.iloc[:, :max_nodes]
                print(f"Limited to first {max_nodes} columns for performance.")
            
            nodes = list(df_analysis.columns)
            
            if len(nodes) < 2:
                raise ValueError("At least 2 columns are needed to create a network graph.")
            
            G = nx.Graph()
            # Add nodes with cleaned names for display
            node_labels = {}
            for node in nodes:
                clean_name = clean_column_name(node)
                G.add_node(node, label=clean_name)
                node_labels[node] = clean_name
            
            print(f"Creating network with {len(nodes)} nodes: {nodes}")
            
            # Calculate correlations between all pairs of columns
            correlations_found = 0
            for i, col1 in enumerate(nodes):
                for j, col2 in enumerate(nodes):
                    if i < j:  # Avoid duplicate edges
                        correlation = calculate_correlation(
                            df_analysis[col1], 
                            df_analysis[col2], 
                            correlation_method
                        )
                        
                        print(f"Correlation between {col1} and {col2}: {correlation:.3f}")
                        
                        # Only add edge if correlation is above threshold
                        if correlation >= correlation_threshold:
                            # Scale edge width (make it more visible)
                            edge_weight = max(correlation * edge_width_factor * 5, 0.5)
                            G.add_edge(col1, col2, weight=edge_weight, correlation=correlation)
                            correlations_found += 1
            
            print(f"Found {correlations_found} correlations above threshold {correlation_threshold}")

        # Check if graph has any edges
        if G.number_of_edges() == 0:
            # If still no correlations found, create edges for the strongest correlations anyway
            print("No correlations found above threshold. Creating network with strongest available correlations.")
            nodes_list = list(G.nodes())
            
            if 'correlation_details' in locals() and correlation_details:
                # Sort by correlation strength and take top connections
                sorted_correlations = sorted(correlation_details, key=lambda x: x['correlation'], reverse=True)
                
                # Add at least the top 3 correlations or number of possible edges, whichever is smaller
                max_edges = min(len(sorted_correlations), max(3, len(nodes_list) - 1))
                
                for i in range(max_edges):
                    detail = sorted_correlations[i]
                    if detail['correlation'] > 0:  # Only positive correlations
                        edge_weight = max(detail['correlation'] * edge_width_factor * 5, 0.5)
                        G.add_edge(detail['col1'], detail['col2'], 
                                 weight=edge_weight, correlation=detail['correlation'])
                        print(f"Added edge: {detail['col1']} - {detail['col2']} (corr: {detail['correlation']:.3f})")
            
            # If still no edges, create a simple connected graph for visualization
            if G.number_of_edges() == 0:
                for i in range(len(nodes_list) - 1):
                    G.add_edge(nodes_list[i], nodes_list[i+1], weight=0.1, correlation=0.1)

        # Get edge properties
        edge_widths = []
        raw_edge_labels = {}
        
        for u, v, d in G.edges(data=True):
            weight = d.get('weight', 0.5)
            edge_widths.append(max(weight, 0.3))  # Minimum edge width for visibility
            
            if show_edge_weights:
                # Show correlation value if available, otherwise weight
                display_value = d.get('correlation', weight)
                raw = f"{display_value:.2f}"
                if language == 'bengali':
                    raw = raw.translate(digit_map)
                raw_edge_labels[(u, v)] = raw

        # Determine figure size based on number of nodes
        num_nodes = G.number_of_nodes()
        if num_nodes <= 5:
            figsize = (10, 8)
            k_value = 3
        elif num_nodes <= 10:
            figsize = (12, 10)
            k_value = 2
        elif num_nodes <= 20:
            figsize = (14, 12)
            k_value = 1.5
        else:
            figsize = (16, 14)
            k_value = 1

        # Layout and draw
        dpi = 300
        fig, ax = plt.subplots(figsize=figsize, dpi=dpi)
        
        # Use better layout algorithm based on graph structure
        if G.number_of_nodes() <= 8:
            pos = nx.spring_layout(G, seed=42, k=k_value, iterations=500)
        else:
            pos = nx.spring_layout(G, seed=42, k=k_value, iterations=300)

        # Draw network with improved styling
        nx.draw_networkx_nodes(G, pos, 
                              node_size=node_size, 
                              node_color=node_color, 
                              ax=ax, 
                              alpha=0.8,
                              edgecolors='black',
                              linewidths=1)
        
        nx.draw_networkx_edges(G, pos, 
                              width=edge_widths, 
                              alpha=0.7, 
                              edge_color="gray", 
                              ax=ax)
        
        ax.margins(0.2)  # More margin for labels
        ax.axis("off")
        fig.canvas.draw()

        # Calculate crop boundaries
        renderer = fig.canvas.get_renderer()
        bbox = fig.get_tightbbox(renderer)
        crop_x0_px = int(bbox.x0 * dpi)
        crop_y0_px = int(bbox.y0 * dpi)
        final_W_px = int(bbox.width * dpi)
        final_H_px = int(bbox.height * dpi)
        data_to_px = ax.transData.transform
        
        def data_to_image_px(xy):
            px, py = data_to_px(xy)
            return (int(px - crop_x0_px), int(final_H_px - (py - crop_y0_px)))

        # Save base image
        plots_dir = os.path.join(settings.MEDIA_ROOT, f'ID_{user_id}_uploads', 'temporary_uploads', 'plots')
        os.makedirs(plots_dir, exist_ok=True)
        base_path = os.path.join(plots_dir, 'network_base.png')
        fig.savefig(base_path, dpi=dpi, bbox_inches="tight", pad_inches=0.15)
        plt.close(fig)

        # Overlay labels with improved styling
        img = Image.open(base_path).convert("RGBA")
        draw = ImageDraw.Draw(img)
        
        # Font setup
        font_path = os.path.join(settings.BASE_DIR, 'NotoSansBengali-Regular.ttf')
        try:
            font_label = ImageFont.truetype(font_path, size=text_size)
            if show_edge_weights:
                font_weight = ImageFont.truetype(font_path, size=max(weight_font_size, 8))
            else:
                font_weight = font_label
        except:
            font_label = ImageFont.load_default()
            font_weight = ImageFont.load_default()

        # Draw node labels (use clean column names)
        for node in G.nodes():
            # Use cleaned label if available, otherwise use original
            if hasattr(G.nodes[node], 'get') and G.nodes[node].get('label'):
                label = to_label(G.nodes[node]['label'])
            else:
                label = to_label(clean_column_name(node))
            
            px, py = data_to_image_px(pos[node])
            
            # Handle multi-line labels for very long names
            if len(label) > 12:
                # Split into multiple lines
                words = label.split(' ')
                if len(words) > 1:
                    mid = len(words) // 2
                    line1 = ' '.join(words[:mid])
                    line2 = ' '.join(words[mid:])
                    label = f"{line1}\n{line2}"
                elif len(label) > 20:
                    label = f"{label[:12]}\n{label[12:]}"
            
            # Calculate text dimensions
            bbox = draw.multiline_textbbox((0, 0), label, font=font_label)
            w, h = bbox[2] - bbox[0], bbox[3] - bbox[1]
            
            # Add semi-transparent background for better readability
            padding = 4
            bg_color = (255, 255, 255, 200)  # White with transparency
            draw.rectangle([px - w/2 - padding, py - h/2 - padding, 
                          px + w/2 + padding, py + h/2 + padding], 
                          fill=bg_color)
            
            # Add subtle border
            draw.rectangle([px - w/2 - padding, py - h/2 - padding, 
                          px + w/2 + padding, py + h/2 + padding], 
                          outline=(0, 0, 0, 100), width=1)
            
            # Draw text
            draw.multiline_text((px - w/2, py - h/2), label, 
                               font=font_label, fill=text_color, 
                               align='center')

        # Draw edge weight labels with improved styling
        if show_edge_weights and raw_edge_labels:
            for (u, v), raw in raw_edge_labels.items():
                midx = (pos[u][0] + pos[v][0]) / 2
                midy = (pos[u][1] + pos[v][1]) / 2
                px, py = data_to_image_px((midx, midy))
                
                # Background for weight labels
                bbox = draw.textbbox((0, 0), raw, font=font_weight)
                w, h = bbox[2] - bbox[0], bbox[3] - bbox[1]
                
                # Semi-transparent background
                bg_padding = 2
                draw.rectangle([px - w/2 - bg_padding, py - h/2 - bg_padding, 
                              px + w/2 + bg_padding, py + h/2 + bg_padding], 
                              fill=(255, 255, 255, 180))
                
                draw.text((px - w/2, py - h/2), raw, font=font_weight, 
                         fill=weight_color)

        # Save final image with maximum quality and metadata
        labeled_path = os.path.join(plots_dir, 'network_labeled.png')
        
        # Add a subtle watermark/signature (optional)
        watermark_text = f"Network Analysis • {len(G.nodes())} Variables • {len(G.edges())} Connections"
        try:
            watermark_font = ImageFont.truetype(font_path, size=12)
        except:
            watermark_font = font_weight
            
        if watermark_font:
            # Add watermark at bottom right
            img_width, img_height = img.size
            watermark_bbox = draw.textbbox((0, 0), watermark_text, font=watermark_font)
            watermark_w = watermark_bbox[2] - watermark_bbox[0]
            
            # Semi-transparent background for watermark
            draw.rectangle([img_width - watermark_w - 20, img_height - 30, 
                          img_width - 5, img_height - 5], 
                          fill=(255, 255, 255, 180))
            
            draw.text((img_width - watermark_w - 15, img_height - 25), watermark_text, 
                     font=watermark_font, fill=(108, 117, 125, 200))
        
        img.save(labeled_path, 
                quality=100, 
                optimize=False, 
                format='PNG',
                pnginfo=None)  # High quality PNG

        # Generate summary statistics
        analyzed_columns = list(df_analysis.columns) if 'df_analysis' in locals() else selected_columns
        stats = {
            'total_nodes': G.number_of_nodes(),
            'total_edges': G.number_of_edges(),
            'correlation_method': correlation_method,
            'correlation_threshold': correlation_threshold,
            'columns_analyzed': analyzed_columns,
            'selected_columns': selected_columns if selected_columns else "All numeric columns"
        }

        return JsonResponse({
            'success': True,
            'image_path': os.path.join(settings.MEDIA_URL, f'ID_{user_id}_uploads', 
                                     'temporary_uploads', 'plots', 'network_labeled.png'),
            'message': "Network graph generated successfully",
            'stats': stats
        })

    except Exception as e:
        print(f"Error in network graph generation: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})


def create_plot(fig, ax, title, xlabel, ylabel, base, final):
            # Format tick labels first
            xt = [format_tick(x) for x in ax.get_xticks()]
            yt = [format_tick]

def save_plot(plt, filename, user_id):
    import os

    from django.conf import settings
    plot_dir = os.path.join(settings.MEDIA_ROOT, f'ID_{user_id}_uploads', 'temporary_uploads', 'plots')
    os.makedirs(plot_dir, exist_ok=True)
    full_path = os.path.join(plot_dir, filename)
    plot_path = os.path.join('plots', filename)

    try:
        plt.savefig(full_path)
        plt.close()
        return plot_path
    except Exception as e:
        print(f"Failed to save plot: {filename}. Error: {e}")
        return None
    

import hashlib
import os
import re

def build_safe_filename(original_name):
     # Clean illegal characters
    original_name = re.sub(r'[<>:"/\\|?*]', '', original_name).strip()
    
    # Hash for uniqueness
    unique_hash = hashlib.md5(original_name.encode()).hexdigest()[:8]  
    
    # Extract extension
    ext = os.path.splitext(original_name)[1]
    
    # Create shorter file name
    safe_name = f"{unique_hash}{ext}"
    return safe_name

@csrf_exempt
def preview_data(request):
    from django.http import JsonResponse
    from django.conf import settings
    import pandas as pd
    import numpy as np
    import os

    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST is allowed'}, status=405)

    user_id = request.headers.get('userID')
    # filename = request.headers.get('filename')
    # requested_sheet = request.headers.get('sheet')  
    # file_url = request.headers.get('Fileurl')
    data = json.loads(request.body)
    filename = build_safe_filename(data.get("filename"))
    requested_sheet = data.get("sheet")
    file_url = data.get("Fileurl")
 
    if not user_id:
        return JsonResponse({'error': 'User ID not provided'}, status=400)
    if not filename:
        return JsonResponse({'error': 'Filename not provided'}, status=400)
    
    if not file_url:
        return JsonResponse({'error': 'File URL not provided'}, status=400)

    folder_name = f"ID_{user_id}_uploads/temporary_uploads/"
    file_path = os.path.join(settings.MEDIA_ROOT, file_url.replace("/media/", "")) 

    try:
        # Read all sheets once; lets us list names and select the active one
        # If the file is CSV, fall back to a single-sheet behavior (optional)
        if filename.lower().endswith(('.xls', '.xlsx', '.xlsm', '.xlsb', '.ods')):
            book = pd.read_excel(file_path, sheet_name=None)  # dict: {sheet_name: df}
            if not book:
                return JsonResponse({'success': False, 'error': 'No sheets found in workbook.'}, status=400)

            sheet_names = list(book.keys())

            # Choose active sheet: requested name if exists, else first
            active_name = requested_sheet if requested_sheet in book else sheet_names[0]
            df = book[active_name]
        else:
            # Non-Excel fallback (CSV, etc.) – treat as single "Sheet1"
            df = pd.read_csv(file_path)
            sheet_names = ["Sheet1"]
            active_name = "Sheet1"

        # ---- build preview for the chosen sheet ----
        preview_limit = 1000
        df_for_stats = df  # keep original types for stats
        preview_df = df.head(preview_limit).copy()

        # Safe JSON: convert NaN/None to empty string
        preview_df = preview_df.where(pd.notnull(preview_df), "")

        # Stats based on the full selected sheet
        missing_values = df_for_stats.isnull().sum().to_dict()
        num_columns = df_for_stats.select_dtypes(include=[np.number]).columns.tolist()

        return JsonResponse({
            'success': True,
            'sheet_names': sheet_names,       # all available sheet names
            'active_sheet': active_name,      # which one was used
            'columns': preview_df.columns.tolist(),
            'rows': preview_df.to_dict(orient='records'),
            'total_rows': int(len(df_for_stats)),
            'preview_rows': int(min(preview_limit, len(df_for_stats))),
            'missing_values': missing_values,
            'num_columns': num_columns,
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Preview failed: {str(e)}'}, status=500) 
 



@csrf_exempt
def delete_columns_api(request):
    try: 
        ##post_method
        ## extract user ID from request headers
        user_id = request.headers.get('userID')
        data = json.loads(request.body)
        filename =data.get("filename")
        sheet_name = data.get("sheet")
        file_url = data.get("Fileurl")
        print(filename) 
        if not user_id:
            return JsonResponse({'success': False, 'error': 'User ID not provided.'}) 
        print(f"Received User ID: {user_id}")

        if not file_url:
            return JsonResponse({'success': False, 'error': 'File URL not provided.'})
        
        folder_name = f"ID_{user_id}_uploads/temporary_uploads/"
        # Load file safely
        file_path = os.path.join(settings.MEDIA_ROOT, file_url.replace("/media/", "")) 
        if not os.path.exists(file_path):
            return JsonResponse({'success': False, 'error': 'No uploaded file found.'})
        preprocess_folder_name= f"ID_{user_id}_uploads/temporary_uploads/preprocessed/" 
    
        # Create preprocess folder if not exists
        os.makedirs(os.path.join(settings.MEDIA_ROOT, folder_name, "preprocessed"), exist_ok=True)
         
        # Define preprocess file path
        preprocess_file_path = os.path.join(settings.MEDIA_ROOT, preprocess_folder_name, f"preprocess_{filename}")

        # Load from preprocess if exists, otherwise from original
        source_path = preprocess_file_path if os.path.exists(preprocess_file_path) else file_path


        try:
            # Check file extension and load data accordingly
            file_extension = os.path.splitext(filename)[1].lower()
            
            if file_extension == '.xlsx':
                df = pd.read_excel(source_path, sheet_name=None)  # Load all sheets as a dictionary
                if sheet_name in df:
                    df = df[sheet_name]
                
                else:
                     df = list(df.values())[0]

                    
                
            elif file_extension == '.csv':
                df = pd.read_csv(source_path)
                
            else:
                return JsonResponse({'success': False, 'error': 'Unsupported file format. Please upload .csv or .xlsx files.'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Failed to read file: {str(e)}'})


        body = json.loads(request.body)
        columns = body.get('columns', [])

        if not columns:
            return JsonResponse({'success': False, 'error': 'No columns specified.'})

        df.drop(columns=columns, inplace=True, errors='ignore')

        # Save updated sheet
        df.to_excel(preprocess_file_path, index=False)
        print(filename)
        
        file_url = os.path.join(settings.MEDIA_URL, preprocess_folder_name, f"preprocess_{filename}").replace("\\", "/")

        return JsonResponse({
            'success': True,
            'message': f'Deleted columns: {columns}',
            'columns': df.columns.tolist(),
            'rows': df.fillna("").astype(str).to_dict(orient='records'), 
            'file_url': file_url
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@csrf_exempt
# def remove_duplicates_api(request):
#     try:
#         # Extract headers
#         user_id = request.headers.get('userID')
#         filename = request.headers.get('filename')
#         if not user_id:
#             return JsonResponse({'success': False, 'error': 'User ID not provided.'})

#         print(f"Received User ID: {user_id}")
#         folder_name = f"ID_{user_id}_uploads/temporary_uploads/"
#         file_path = os.path.join(settings.MEDIA_ROOT, folder_name, filename)

#         if not os.path.exists(file_path):
#             return JsonResponse({'success': False, 'error': 'No uploaded file found.'})

#         # Load Excel
#         df = pd.read_excel(file_path)
#         original_len = len(df)

#         # Get column filter
#         body = json.loads(request.body)
#         columns = body.get('columns', [])
        
#         print(columns) 
#         removed_info = []

#         if columns:
#             # Column-specific duplicate removal
#             for col in columns:
#                 if col not in df.columns:
#                     removed_info.append({'column': col, 'error': 'Column not found'})
#                     continue

#                 dup_mask = df[col].duplicated(keep=False)
#                 dup_indices = df.index[dup_mask].tolist()
#                 # print(dup_indices)  

#                 if dup_indices:
#                     dup_values = df.loc[dup_indices, col].tolist()
#                     df.drop(df[df[col].duplicated(keep='first')].index, inplace=True)
#                     removed = original_len - len(df)
#                     removed_info.append({
#                         'column': col,
#                         'duplicates_found': len(dup_indices),
#                         'removed_rows': removed,
#                         'duplicate_values': list(set(dup_values))
#                     })
#                     original_len = len(df)  # Update
#                 else:
#                     removed_info.append({'column': col, 'message': 'No duplicates found'})
#                 summary_msg = "Column-wise duplicate removal complete."
#         print(removed_info) 
        

#         # Save the modified file
#         df.to_excel(file_path, index=False)

#         return JsonResponse({
#             'success': True,
#             'message': summary_msg,
#             'info': removed_info,
#             'columns': df.columns.tolist(),
#             'rows': df.head(100).fillna("").astype(str).to_dict(orient='records'),
#             'remaining_rows': len(df)
#         })

#     except Exception as e:
#         return JsonResponse({'success': False, 'error': str(e)})

def find_duplicates_api(request):
    try:
        # --- Extract headers ---
        if(request.method == 'POST'):

            user_id = request.headers.get('userID')
            data = json.loads(request.body)
            filename = data.get("filename")
            sheet_name = data.get("sheet") 
            file_url = data.get("Fileurl")
            print(f"Received User ID: {user_id}") 
            if not user_id:
                return JsonResponse({'success': False, 'error': 'User ID not provided.'})
            if not file_url:
                return JsonResponse({'success': False, 'error': 'File URL not provided.'})
            folder_name = f"ID_{user_id}_uploads/temporary_uploads/"


            file_path = os.path.join(settings.MEDIA_ROOT, file_url.replace("/media/", ""))

            if not os.path.exists(file_path):
                return JsonResponse({'success': False, 'error': 'No uploaded file found.'})
            
            preprocess_folder_name= f"ID_{user_id}_uploads/temporary_uploads/preprocessed/" 
        
            # Create preprocess folder if not exists
            os.makedirs(os.path.join(settings.MEDIA_ROOT, folder_name, "preprocessed"), exist_ok=True)
            
            # Define preprocess file path
            preprocess_file_path = os.path.join(settings.MEDIA_ROOT, preprocess_folder_name, 'preprocess_'+ filename)

            # Load from preprocess if exists, otherwise from original
            source_path = preprocess_file_path if os.path.exists(preprocess_file_path) else file_path

            # --- Load Excel ---
            try:
                if filename.lower().endswith(('.xls', '.xlsx', '.xlsm', '.xlsb', '.ods')):
                    df= pd.read_excel(source_path, sheet_name= None)
                    if sheet_name in df:
                        df = df[sheet_name]
                    else:
                        df = list(df.values())[0]
                else:
                    df = pd.read_csv(source_path)
            except Exception as e:
                return JsonResponse({'success': False, 'error': f'Failed to read Excel: {str(e)}'})
            total_rows = len(df)

            # --- Get columns from request body ---
            body = json.loads(request.body)
            columns = body.get('columns', [])
            df.columns = df.columns.str.replace("\n", " ").str.strip()
            print(df.columns.tolist())            
            
            dup_info = []
            summary_msg = "No duplicates detected."
            duplicate_indices = []  

            if columns:
                existing_cols = [c for c in columns if c in df.columns]
                missing_cols = [c for c in columns if c not in df.columns]
                print(columns)
                print(existing_cols) 
                for col in missing_cols:
                    dup_info.append({'column': col, 'error': 'Column not found'})

                if len(existing_cols) == 1:
                    col = existing_cols[0]
                    dup_mask = df[col].duplicated(keep=False)
                    dupped_rows = df[dup_mask]
          
                    if not dupped_rows.empty:
                        duplicate_indices = dupped_rows.index.tolist()
                        dup_info.append({
                            'column': col,
                            'duplicates_found': len(dupped_rows),
                            'duplicate_values': dupped_rows[col].tolist()
                        })
                        print(dupped_rows) 
                        summary_msg = f"Duplicates detected in column '{col}'."
                    else:
                        dup_info.append({'column': col, 'message': 'No duplicates found'})

                elif len(existing_cols) > 1:
                    dup_mask = df.duplicated(subset=existing_cols, keep=False)
                    dupped_rows = df[dup_mask]

                    if not dupped_rows.empty:
                        duplicate_indices = dupped_rows.index.tolist()
                        dup_info.append({
                            'columns': existing_cols,
                            'duplicates_found': len(dupped_rows),
                            'duplicate_combinations': dupped_rows[existing_cols].drop_duplicates().to_dict(orient="records")
                        })
                        summary_msg = f"Duplicates detected based on combination of columns: {existing_cols}"
                    else:
                        dup_info.append({'columns': existing_cols, 'message': 'No duplicates found'})
            print(duplicate_indices)  
            return JsonResponse({
                'success': True,
                'message': summary_msg,
                'info': dup_info,
                'columns': df.columns.tolist(),
                'rows': df.fillna("").astype(str).to_dict(orient='records'),  
                'duplicate_indices': duplicate_indices,  
                'total_rows': total_rows
            })



    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
    
@csrf_exempt
def remove_duplicates(request):
    user_id = request.headers.get('userID') 
    data = json.loads(request.body)
    filename = data.get("filename")
    sheet_name = data.get("sheet")
    file_url = data.get('Fileurl') 
    columns = data.get('columns', [])
    mode = data.get('mode')  # "all" or "selected"
    selected_indices = data.get('selected', []) 
    print(selected_indices)
    print(file_url) 
    try:
        folder_name = f"ID_{user_id}_uploads/temporary_uploads/"
        file_path = os.path.join(settings.MEDIA_ROOT, file_url.replace("/media/", ""))

        if not os.path.exists(file_path):
            return JsonResponse({'success': False, 'error': 'No uploaded file found.'})
        preprocess_folder_name= f"ID_{user_id}_uploads/temporary_uploads/preprocessed/" 
    
        # Create preprocess folder if not exists
        os.makedirs(os.path.join(settings.MEDIA_ROOT, folder_name, "preprocessed"), exist_ok=True)
         
        # Define preprocess file path
        preprocess_file_path = os.path.join(settings.MEDIA_ROOT, preprocess_folder_name, 'preprocess_'+ filename)

        # Load from preprocess if exists, otherwise from original
        source_path = preprocess_file_path if os.path.exists(preprocess_file_path) else file_path
        print(file_path) 
        # --- Load Excel ---
        try:
            if filename.lower().endswith(('.xls', '.xlsx', '.xlsm', '.xlsb', '.ods')):
                df = pd.read_excel(source_path, sheet_name=None)
                if sheet_name in df:
                    df = df[sheet_name]
                else:
                    df = df[list(df.keys())[0]]
            else:
                df = pd.read_csv(source_path)
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Failed to read Excel: {str(e)}'})
        before = len(df)
        # Normalize dataframe column names
        df.columns = (
            df.columns
            .astype(str)
            .str.strip()                 # remove leading/trailing spaces
            .str.replace(r'\s+', ' ', regex=True)  # collapse multiple spaces
        )
        # Normalize input column names
        columns = [
            str(c).strip().replace('_', ' ')
            for c in columns
        ]



        if mode == "all":
            # If columns not given, use all columns
            subset = columns if len(columns) > 0 else list(df.columns)

            # Validate subset exists
            missing = [c for c in subset if c not in df.columns]
            if missing:
                return JsonResponse({"success": False, "error": f"Columns not found: {missing}"}, status=400)

            # Keep first, drop other duplicates across the subset
            df = df.drop_duplicates(subset=subset, keep="first").reset_index(drop=True)

        elif mode == "selected":
            # Ensure ints and in range; drop by POSITION (not labels)
            try:
                positions = [int(i) for i in selected_indices]
            except (TypeError, ValueError):
                return JsonResponse({"success": False, "error": "'selected' must be an array of integers."}, status=400)

            bad = [p for p in positions if p < 0 or p >= len(df)]
            if bad:
                return JsonResponse({"success": False, "error": f"Selected positions out of range: {sorted(bad)}"}, status=400)
            df = df.drop(df.index[positions]).reset_index(drop=True)

        else:
            return JsonResponse({"success": False, "error": "mode must be 'all' or 'selected'."}, status=400)

        removed = before - len(df)
        df.to_excel(preprocess_file_path, index=False)
        print(removed)
        #print rows after deletion
        print(df) 
        file_url = os.path.join(settings.MEDIA_URL, preprocess_folder_name, 'preprocess_' + filename).replace("\\", "/")

        return JsonResponse({
            "success": True,
            "message": f"{'Kept first & dropped duplicates' if mode=='all' else 'Deleted selected'}; removed {removed} row(s).",
            "removed": removed,
            "rows": df.fillna("").astype(str).to_dict(orient="records"),
            "columns": df.columns.tolist(),
            "file_url": file_url
        })
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})

@csrf_exempt
def handle_missing_api(request):
    try:
        # Extract user ID from request headers
        user_id = request.headers.get('userID')
        data = json.loads(request.body)
        filename = data.get("filename")
        sheet_name = data.get("sheet")
        file_url = data.get("Fileurl")
        if not user_id:
            return JsonResponse({'success': False, 'error': 'User ID not provided.'})
        print(f"Received User ID: {user_id}")

        if not file_url:
            return JsonResponse({'success': False, 'error': 'File URL not provided.'})
        
        folder_name = f"ID_{user_id}_uploads/temporary_uploads/"
        file_path = os.path.join(settings.MEDIA_ROOT, file_url.replace("/media/", ""))
        if not os.path.exists(file_path):
            return JsonResponse({'success': False, 'error': 'No uploaded file found.'})
        preprocess_folder_name= f"ID_{user_id}_uploads/temporary_uploads/preprocessed/" 
    
        # Create preprocess folder if not exists
        os.makedirs(os.path.join(settings.MEDIA_ROOT, folder_name, "preprocessed"), exist_ok=True)
         
        # Define preprocess file path
        preprocess_file_path = os.path.join(settings.MEDIA_ROOT, preprocess_folder_name, 'preprocess_'+ filename)

        # Load from preprocess if exists, otherwise from original
        source_path = preprocess_file_path if os.path.exists(preprocess_file_path) else file_path
        try:
            if filename.lower().endswith(('.xls', '.xlsx', '.xlsm', '.xlsb', '.ods')):
                df = pd.read_excel(source_path, sheet_name=None)
                if sheet_name in df:
                    df = df[sheet_name]
                else:
                    df = df[list(df.keys())[0]]
            else:
                df = pd.read_csv(source_path)
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Failed to read Excel: {str(e)}'})

        body = json.loads(request.body)
        col = body.get('column')
        method = body.get('method', '')
        missing_spec = body.get('missing_spec', '').strip()

        if method not in ['drop', 'fill_mean', 'fill_median', 'fill_mode']:
            return JsonResponse({'success': False, 'error': 'Invalid method.'})
        
        if missing_spec:
            df.replace(missing_spec, np.nan, inplace=True)


        

        empty_columns = df.columns[df.isnull().all()].tolist()
        empty_rows = df.index[df.isnull().all(axis=1)].tolist()
        df.drop(columns=empty_columns, inplace=True)
        df.drop(index=empty_rows, inplace=True)

        targets = df.columns.tolist() if col == 'all' else [col]

        if method == 'drop':
            df.dropna(subset=targets, inplace=True)
        else:
            for c in targets:
                if method == 'fill_mean' and pd.api.types.is_numeric_dtype(df[c]):
                    df[c] = df[c].fillna(df[c].mean())
                elif method == 'fill_median' and pd.api.types.is_numeric_dtype(df[c]):
                    df[c] = df[c].fillna(df[c].median())
                elif method == 'fill_mode':
                    mode_val = df[c].mode(dropna=True) 
                    if not mode_val.empty:
                        df[c] = df[c].fillna(mode_val.iloc[0])

        df.to_excel(preprocess_file_path, index=False)
        file_url = os.path.join(settings.MEDIA_URL, preprocess_folder_name, 'preprocess_' + filename).replace("\\", "/")

        return JsonResponse({
            'success': True,
            'message': f'Missing values handled using method: {method}',
            'deleted_columns': empty_columns,
            'deleted_rows': empty_rows,
            'columns': df.columns.tolist(),
            'rows': df.head(100).fillna("").astype(str).to_dict(orient='records'),
            'file_url': file_url
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@csrf_exempt
def outliers_summary_api(request): 
    try:
        print("Request method:", request.method)

        user_id = request.headers.get('userID')
        data = json.loads(request.body)
        filename = data.get("filename")
        sheet = data.get("sheet")
        file_url = data.get("Fileurl")

        if not user_id or not filename:
            return JsonResponse({'success': False, 'error': 'User ID or filename not provided.'})

        if not file_url:
            return JsonResponse({'success': False, 'error': 'File URL not provided.'})
        folder = f"ID_{user_id}_uploads/temporary_uploads/"
        file_path = os.path.join(settings.MEDIA_ROOT, file_url.replace("/media/", ""))

        if not os.path.exists(file_path):
            return JsonResponse({'success': False, 'error': 'File not found.'})

        preprocess_folder_name= f"ID_{user_id}_uploads/temporary_uploads/preprocessed/" 
    
        # Create preprocess folder if not exists
        os.makedirs(os.path.join(settings.MEDIA_ROOT, folder, "preprocessed"), exist_ok=True)
         
        # Define preprocess file path
        preprocess_file_path = os.path.join(settings.MEDIA_ROOT, preprocess_folder_name, 'preprocess_'+ filename)

        # Load from preprocess if exists, otherwise from original
        source_path = preprocess_file_path if os.path.exists(preprocess_file_path) else file_path

        try:
            if filename.lower().endswith(('.xls', '.xlsx', '.xlsm', '.xlsb', '.ods')):
                df = pd.read_excel(source_path, sheet_name=None)
                if sheet in df:
                    df = df[sheet]
                else:
                    df = df[list(df.keys())[0]]
            else:
                df = pd.read_csv(source_path)
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Failed to read file: {str(e)}'})
        
        # Identify numeric columns
        num_cols = df.select_dtypes(include=[np.number]).columns.tolist()

        if not num_cols:
            return JsonResponse({'success': False, 'error': 'No numeric columns found.'})

        # Compute IQR bounds
        Q1 = df[num_cols].quantile(0.25)
        Q3 = df[num_cols].quantile(0.75)
        IQR = Q3 - Q1
        lower = Q1 - 1.5 * IQR
        upper = Q3 + 1.5 * IQR

        # Detect outliers
        outlier_mask = (df[num_cols] < lower) | (df[num_cols] > upper)

        # Summary per column — convert to int
        outliers_summary = {col: int(count) for col, count in outlier_mask.sum().items()}

        # Cell-level details — convert index and value to JSON-safe types
        outlier_cells = []
        for col in num_cols:
            for idx in df.index[outlier_mask[col]]:
                value = df.at[idx, col]
                # Convert value to native Python type
                safe_value = float(value) if isinstance(value, (np.integer, np.floating)) else value
                outlier_cells.append({
                    'row': int(idx),
                    'column': col,
                    'value': safe_value
                })

        # print(num_cols)  
        # print(outliers_summary)
        # print(outlier_cells) 

        return JsonResponse({
            'success': True,
            'numeric_columns': num_cols,
            'outliers_summary': outliers_summary,
            'outlier_cells': outlier_cells
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@csrf_exempt
def handle_outliers_api(request):
    try:
        # Extract user ID from request headers
        user_id = request.headers.get('userID')
        data = json.loads(request.body)
        filename = data.get("filename")
        sheet = data.get("sheet")
        file_url = data.get("Fileurl")

        if not user_id:
            return JsonResponse({'success': False, 'error': 'User ID not provided.'})
        print(f"Received User ID: {user_id}")

        if not file_url:
            return JsonResponse({'success': False, 'error': 'File URL not provided.'})
        
        folder_name = f"ID_{user_id}_uploads/temporary_uploads/"
        file_path = os.path.join(settings.MEDIA_ROOT, file_url.replace("/media/", ""))
        if not os.path.exists(file_path):
            return JsonResponse({'success': False, 'error': 'No uploaded file found.'})


        preprocess_folder_name= f"ID_{user_id}_uploads/temporary_uploads/preprocessed/" 
    
        # Create preprocess folder if not exists
        os.makedirs(os.path.join(settings.MEDIA_ROOT, folder_name, "preprocessed"), exist_ok=True)
         
        # Define preprocess file path
        preprocess_file_path = os.path.join(settings.MEDIA_ROOT, preprocess_folder_name, 'preprocess_'+ filename)

        # Load from preprocess if exists, otherwise from original
        source_path = preprocess_file_path if os.path.exists(preprocess_file_path) else file_path

        try:
            if filename.lower().endswith(('.xls', '.xlsx', '.xlsm', '.xlsb', '.ods')):
                df = pd.read_excel(source_path, sheet_name=None)
                if sheet in df:
                    df = df[sheet]
                else:
                    df = df[list(df.keys())[0]]
            else:
                df = pd.read_csv(source_path)
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Failed to read Excel: {str(e)}'})
        
        body = json.loads(request.body)
        col = body.get('column', '')
        method = body.get('method', '')

        if col not in df.columns or not pd.api.types.is_numeric_dtype(df[col]):
            return JsonResponse({'success': False, 'error': 'Column is missing or not numeric.'})

        Q1, Q3 = df[col].quantile([0.25, 0.75])
        IQR = Q3 - Q1
        lower, upper = Q1 - 1.5 * IQR, Q3 + 1.5 * IQR

        if method == 'remove':
            before = len(df)
            df = df[(df[col] >= lower) & (df[col] <= upper)]
            removed = before - len(df)
            message = f"Removed {removed} outliers."
        elif method == 'cap':
            df[col] = np.where(df[col] < lower, lower, np.where(df[col] > upper, upper, df[col]))
            message = "Outliers capped to bounds."
        else:
            return JsonResponse({'success': False, 'error': 'Invalid method.'})

        df.to_excel(preprocess_file_path, index=False)
        file_url = os.path.join(settings.MEDIA_URL, preprocess_folder_name, 'preprocess_' + filename).replace("\\", "/")

        return JsonResponse({
            'success': True,
            'message': message,
            'columns': df.columns.tolist(),
            'rows': df.head(100).fillna("").astype(str).to_dict(orient='records'),
            'file_url': file_url
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@csrf_exempt
def rank_categorical_column_api(request):
    try:
        # Extract user ID from request headers
        user_id = request.headers.get('userID')
        data = json.loads(request.body)
        filename = data.get("filename")
        sheet = data.get("sheet")
        file_url = data.get("Fileurl")
        if not user_id:
            return JsonResponse({'success': False, 'error': 'User ID not provided.'})
        print(f"Received User ID: {user_id}")

        if not file_url:
            return JsonResponse({'success': False, 'error': 'File URL not provided.'})
        
        folder_name = f"ID_{user_id}_uploads/temporary_uploads/"
        file_path = os.path.join(settings.MEDIA_ROOT, file_url.replace("/media/", ""))
        if not os.path.exists(file_path):
            return JsonResponse({'success': False, 'error': 'No uploaded file found.'})

        preprocess_folder_name= f"ID_{user_id}_uploads/temporary_uploads/preprocessed/" 
    
        # Create preprocess folder if not exists
        os.makedirs(os.path.join(settings.MEDIA_ROOT, folder_name, "preprocessed"), exist_ok=True)
         
        # Define preprocess file path
        preprocess_file_path = os.path.join(settings.MEDIA_ROOT, preprocess_folder_name, 'preprocess_'+ filename)

        # Load from preprocess if exists, otherwise from original
        source_path = preprocess_file_path if os.path.exists(preprocess_file_path) else file_path

        try:
            if filename.lower().endswith(('.xls', '.xlsx', '.xlsm', '.xlsb', '.ods')):
                df = pd.read_excel(source_path, sheet_name=None)
                if sheet in df:
                    df = df[sheet]
                else:
                    df = df[list(df.keys())[0]]
            else:
                df = pd.read_csv(source_path)
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Failed to read Excel: {str(e)}'})
        
        body = json.loads(request.body)

        col = body.get('column', '')
        mapping = body.get('mapping', {})  # {'Yes': 1, 'No': 2, ...}

        if col not in df.columns or not mapping:
            return JsonResponse({'success': False, 'error': 'Invalid column or mapping.'})

        new_col = col + '_ranked'
        # df[new_col] = df[col].map(mapping)
        

        # Get index of original column
        col_index = df.columns.get_loc(col) 

        # Create ranked values
        ranked_series = df[col].map(mapping)

        # Insert ranked column right after original
        df.insert(col_index + 1, new_col, ranked_series)


        df.to_excel(preprocess_file_path, index=False)

        file_url = os.path.join(settings.MEDIA_URL, preprocess_folder_name, 'preprocess_' + filename).replace("\\", "/")

        return JsonResponse({
            'success': True,
            'message': f"Column '{col}' ranked into '{new_col}'",
            'columns': df.columns.tolist(),
            'rows': df.fillna("").astype(str).to_dict(orient='records'),
            'file_url': file_url
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

# @csrf_exempt
# def split_column_api(request):
#     from django.http import JsonResponse
#     from django.conf import settings
#     import pandas as pd
#     import os, json
#     from collections import Counter

#     try:
#         # --- Extract metadata ---
#         user_id = request.headers.get('userID')
#         filename = request.headers.get('filename') 
#         if not user_id or not filename:
#             return JsonResponse({'success': False, 'error': 'User ID or filename not provided.'})

#         folder_name = f"ID_{user_id}_uploads/temporary_uploads/"
#         file_path = os.path.join(settings.MEDIA_ROOT, folder_name, filename)

#         if not os.path.exists(file_path):
#             return JsonResponse({'success': False, 'error': 'Uploaded file not found.'})

#         try:
#             df = pd.read_excel(file_path)
#         except Exception as e:
#             return JsonResponse({'success': False, 'error': f'Failed to read Excel file: {str(e)}'})

#         # --- Parse JSON body ---
#         body = json.loads(request.body)
#         column = body.get('column', '').strip()
#         method = body.get('method', '').strip().lower()
#         phrases = body.get('phrases', [])
#         delete_original = body.get('delete_original', False)

#         if column not in df.columns:
#             return JsonResponse({'success': False, 'error': f"Column '{column}' not found in dataset."})

#         new_cols = []
#         df_split = None
#         original_idx = df.columns.get_loc(column)

#         # --- Split Logic ---
#         if method == 'comma':
#             df_split = df[column].astype(str).str.split(',', expand=True)
#             df_split.columns = [f"{column}_part_{i+1}" for i in range(df_split.shape[1])]
#             new_cols = df_split.columns.tolist()

#         elif method == 'semicolon':
#             df_split = df[column].astype(str).str.split(';', expand=True)
#             df_split.columns = [f"{column}_part_{i+1}" for i in range(df_split.shape[1])]
#             new_cols = df_split.columns.tolist()

#         elif method == 'tags':
#             df_split = df[column].astype(str).str.extractall(r'<(.*?)>').unstack().droplevel(0, axis=1)
#             df_split.columns = [f"{column}_tag_{i+1}" for i in range(df_split.shape[1])]
#             new_cols = df_split.columns.tolist()

#         elif method == 'custom' and phrases:
#             def count_phrases(text, known_phrases):
#                 tokens = str(text).split(", ")
#                 result = []
#                 i = 0
#                 while i < len(tokens):
#                     if i + 1 < len(tokens):
#                         combined = tokens[i] + ", " + tokens[i + 1]
#                         if combined in known_phrases:
#                             result.append(combined)
#                             i += 2
#                             continue
#                     result.append(tokens[i])
#                     i += 1
#                 return Counter(result)

#             df_split = pd.DataFrame()
#             for phrase in phrases:
#                 df_split[phrase] = df[column].apply(lambda x: count_phrases(x, phrases).get(phrase, 0))
#             new_cols = df_split.columns.tolist()

#         else:
#             return JsonResponse({'success': False, 'error': 'Invalid method or missing required input (phrases).'})

#         # --- Remove new columns if already exist ---
#         for col in new_cols:
#             if col in df.columns:
#                 df.drop(columns=[col], inplace=True)

#         # --- Insert new columns after original ---
#         for i, col in enumerate(new_cols):
#             df.insert(original_idx + 1 + i, col, df_split[col])

#         # --- Delete original column if requested ---
#         if delete_original and column in df.columns:
#             df.drop(columns=[column], inplace=True)

#         # --- Save and Respond ---
#         df.to_excel(file_path, index=False)
#         df = df.fillna("").astype(str)

#         return JsonResponse({
#             'success': True,
#             'message': f"Column '{column}' split successfully using '{method}' method.",
#             'columns': df.columns.tolist(),
#             'rows': df.head(100).fillna("").astype(str).to_dict(orient='records'), 
#             'new_columns': new_cols,
#             'original_column_deleted': delete_original
#         })

#     except Exception as e:
#         return JsonResponse({'success': False, 'error': str(e)})

@csrf_exempt
def split_column_api(request):
    from django.http import JsonResponse
    from django.conf import settings
    import pandas as pd
    import os, json
    import re
    from collections import Counter

    try:
        # --- Extract metadata ---
        user_id = request.headers.get('userID')
        data = json.loads(request.body)
        filename = data.get("filename")
        sheet_name = data.get("sheet") 
        file_url = data.get("Fileurl")

        if not user_id or not filename:
            return JsonResponse({'success': False, 'error': 'User ID or filename not provided.'})
        
        if not file_url:
            return JsonResponse({'success': False, 'error': 'File URL not provided.'})

        folder_name = f"ID_{user_id}_uploads/temporary_uploads/"
        file_path = os.path.join(settings.MEDIA_ROOT, file_url.replace("/media/", ""))

        if not os.path.exists(file_path):
            return JsonResponse({'success': False, 'error': 'Uploaded file not found.'})

        preprocess_folder_name= f"ID_{user_id}_uploads/temporary_uploads/preprocessed/" 
    
        # Create preprocess folder if not exists
        os.makedirs(os.path.join(settings.MEDIA_ROOT, folder_name, "preprocessed"), exist_ok=True)
         
        # Define preprocess file path
        preprocess_file_path = os.path.join(settings.MEDIA_ROOT, preprocess_folder_name, 'preprocess_'+ filename)

        # Load from preprocess if exists, otherwise from original
        source_path = preprocess_file_path if os.path.exists(preprocess_file_path) else file_path


        try:
            if filename.lower().endswith(('.xls', '.xlsx', '.xlsm', '.xlsb', '.ods')):
                df = pd.read_excel(source_path, sheet_name=None)
                if sheet_name in df:
                    df = df[sheet_name]
                else:
                    df = list(df.values())[0]
            else:
                df = pd.read_csv(source_path)
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Failed to read file: {str(e)}'})

        # --- Parse JSON body ---
        body = json.loads(request.body)
        column = body.get('column', '').strip()
        method = body.get('method', '').strip().lower()
        phrases = body.get('phrases', [])
        delete_original = body.get('delete_original', False),


        if column not in df.columns:
            return JsonResponse({'success': False, 'error': f"Column '{column}' not found in dataset."})

        new_cols = []
        df_split = None
        original_idx = df.columns.get_loc(column)

        # --- Helper: count user-defined phrases ---
        def count_phrases(input_str, known_phrases):
            tokens = str(input_str).split(", ")
            result = []
            i = 0
            while i < len(tokens):
                if i + 1 < len(tokens):
                    combined = tokens[i] + ", " + tokens[i + 1]
                    if combined in known_phrases:
                        result.append(combined)
                        i += 2
                        continue
                result.append(tokens[i])
                i += 1
            return Counter(result)

       
        # --- Split Logic ---
        def extract_labels(text, method): 
            if pd.isna(text):
                return []

            text = str(text)

            if method == 'comma':
                items = re.split(r',\s*(?![^()]*\))', text)
                return [
                    re.split(r'\(', item)[0].strip().lower()
                    for item in items
                    if item.strip()
                ]

            elif method == 'semicolon':
                return [
                    item.strip().lower()
                    for item in text.split(';')
                    if item.strip()
                ]

            elif method == 'tags':
                return [
                    tag.strip().lower()
                    for tag in re.findall(r'<(.*?)>', text)
                    if tag.strip()
                ]

            elif method == 'custom':
                return []  # handled separately

            return []

        # if method == 'comma':
        #     # Step 1: split and normalize main labels
        #     split_data = df[column].apply(extract_main_label)
        #     all_labels = sorted({label for sublist in split_data for label in sublist})

        #     # Step 2: one-hot encode
        #     df_split = pd.DataFrame()
        #     for label in all_labels:
        #         df_split[f"{column}[{label}]"] = split_data.apply(lambda x: 1 if label in x else 0)

        #     new_cols = df_split.columns.tolist()

        # elif method == 'semicolon': 
        #     df_split = df[column].astype(str).str.split(';', expand=True)
        #     df_split.columns = [f"{column}[{i}]" for i in df_split.columns]
        #     new_cols = df_split.columns.tolist()

        # elif method == 'tags':
        #     df_split = df[column].astype(str).str.extractall(r'<(.*?)>').unstack().droplevel(0, axis=1)
        #     df_split.columns = [f"{column}[{col.strip()}]" for col in df_split.columns]
        #     new_cols = df_split.columns.tolist()


        if method in ['comma', 'semicolon', 'tags']:

            # Step 1: extract labels
            split_data = df[column].apply(lambda x: extract_labels(x, method))

            # Step 2: collect unique labels
            all_labels = sorted(
                {label for sublist in split_data for label in sublist}
            )

            if not all_labels:
                return JsonResponse({
                    'success': False,
                    'error': 'No valid values found to split.'
                })

            # Step 3: one-hot encode 
            df_split = pd.DataFrame()
            for label in all_labels:
                df_split[f"{column}[{label}]"] = split_data.apply(
                    lambda x: 1 if label in x else 0
                )

            new_cols = df_split.columns.tolist()


        elif method == 'custom' and phrases:
            df_split = pd.DataFrame()
            for phrase in phrases:
                df_split[f"{column}[{phrase}]"] = df[column].apply(lambda x: count_phrases(x, phrases).get(phrase, 0))
            new_cols = df_split.columns.tolist()

        else:
            return JsonResponse({'success': False, 'error': 'Invalid method or missing required input (phrases).'})

        # --- Remove existing new columns ---
        for col in new_cols:
            if col in df.columns:
                df.drop(columns=[col], inplace=True)

        # --- Insert new columns after original ---
        for i, col in enumerate(new_cols):
            df.insert(original_idx + 1 + i, col, df_split[col])

        # --- Delete original column if requested ---
        if delete_original and column in df.columns:
            df.drop(columns=[column], inplace=True)

        # --- Save and respond ---
        df.to_excel(preprocess_file_path, index=False)
        df = df.fillna("").astype(str)

        file_url = os.path.join(settings.MEDIA_URL, preprocess_folder_name, 'preprocess_' + filename).replace("\\", "/")

        return JsonResponse({
            'success': True,
            'message': f"Column '{column}' split successfully using '{method}' method.",
            'columns': df.columns.tolist(),
            'rows': df.head(100).to_dict(orient='records'), 
            'new_columns': new_cols,
            'original_column_deleted': delete_original,
            'file_url': file_url
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@csrf_exempt
def group_data_api(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method.'})

    try:
      

        user_id = request.headers.get('userID')
        filename = request.headers.get('filename')
        sheet= request.headers.get('sheet')
        file_url = request.headers.get('Fileurl')
        if not user_id:
            return JsonResponse({'success': False, 'error': 'User ID not provided.'})
        if not file_url:
            return JsonResponse({'success': False, 'error': 'File URL not provided.'})
        

        folder_name = f"ID_{user_id}_uploads/temporary_uploads/"
        file_path = os.path.join(settings.MEDIA_ROOT, file_url.replace("/media/", ""))

        if not os.path.exists(file_path):
            return JsonResponse({'success': False, 'error': 'Uploaded Excel file not found.'})

        body = json.loads(request.body)
        grouping_pairs = body.get('groupingPairs', [])

        if not grouping_pairs:
            return JsonResponse({'success': False, 'error': 'No grouping pairs provided.'})

        preprocess_folder_name= f"ID_{user_id}_uploads/temporary_uploads/preprocessed/" 
    
        # Create preprocess folder if not exists
        os.makedirs(os.path.join(settings.MEDIA_ROOT, folder_name, "preprocessed"), exist_ok=True)
         
        # Define preprocess file path
        preprocess_file_path = os.path.join(settings.MEDIA_ROOT, preprocess_folder_name, 'preprocess_'+ filename)

        # Load from preprocess if exists, otherwise from original
        source_path = preprocess_file_path if os.path.exists(preprocess_file_path) else file_path

        try:
            if filename.lower().endswith(('.xls', '.xlsx', '.xlsm', '.xlsb', '.ods')):
                df= pd.read_excel(source_path,engine='openpyxl', sheet_name=None)
                if sheet in df:
                    df = df[sheet]
                else:
                    df = list(df.values())[0]
            else:
                df = pd.read_csv(source_path)
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Failed to read Excel file: {str(e)}'})

        grouped_dfs = []
        for pair in grouping_pairs:
            group_col = pair.get('group_col')
            value_col = pair.get('value_col')

            if group_col not in df.columns or value_col not in df.columns:
                return JsonResponse({'success': False, 'error': f"Invalid columns: {group_col}, {value_col}"})

            if not pd.api.types.is_numeric_dtype(df[value_col]):
                return JsonResponse({'success': False, 'error': f"'{value_col}' must be numeric."})

            groups = df[group_col].dropna().unique().tolist()

            grouped_df = pd.concat([
                df.loc[df[group_col] == g, [group_col, value_col]].assign(Group=g)
                for g in groups
            ])
            grouped_dfs.append(grouped_df)

        base_filename = os.path.splitext(filename)[0]
        grouped_filename = f"{base_filename}_grouped_splits.xlsx"
        output_path = os.path.join(settings.MEDIA_ROOT, preprocess_folder_name, grouped_filename)

        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for i, gdf in enumerate(grouped_dfs):
                    sheet_name = f"group_{i+1}"
                    gdf.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Failed to save grouped file: {str(e)}'})
        preview_data = grouped_dfs[0].to_dict(orient='records') if grouped_dfs else []
        file_url = os.path.join(settings.MEDIA_URL, preprocess_folder_name, 'preprocess_' + filename).replace("\\", "/")
        

        return JsonResponse({
            'success': True,
            'message': 'Grouped splits saved to Excel file.',
            'download_url': f"/media/{folder_name}{grouped_filename}",
            'preview_data': preview_data,
            'file_url': file_url
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@csrf_exempt
def generate_unique_id_column_api(request):
    try:
        # Extract user ID from request headers
        user_id = request.headers.get('userID')
        data = json.loads(request.body) 
        filename = data.get("filename")
        sheet = data.get("sheet")
        file_url = data.get("Fileurl")
        prefix = data.get("prefix")

        if not user_id:
            return JsonResponse({'success': False, 'error': 'User ID not provided.'})
        print(f"Received User ID: {user_id}")
        if not file_url:
            return JsonResponse({'success': False, 'error': 'File URL not provided.'})
        

        folder_name = f"ID_{user_id}_uploads/temporary_uploads/"
        file_path = os.path.join(settings.MEDIA_ROOT, file_url.replace("/media/", ""))
        if not os.path.exists(file_path):
            return JsonResponse({'success': False, 'error': 'No uploaded Excel file found.'})

        preprocess_folder_name= f"ID_{user_id}_uploads/temporary_uploads/preprocessed/" 
    
        # Create preprocess folder if not exists
        os.makedirs(os.path.join(settings.MEDIA_ROOT, folder_name, "preprocessed"), exist_ok=True)
         
        # Define preprocess file path
        preprocess_file_path = os.path.join(settings.MEDIA_ROOT, preprocess_folder_name, 'preprocess_'+ filename)

        # Load from preprocess if exists, otherwise from original
        source_path = preprocess_file_path if os.path.exists(preprocess_file_path) else file_path


        try:
            if filename.lower().endswith(('.xls', '.xlsx', '.xlsm', '.xlsb', '.ods')):
                df = pd.read_excel(source_path, sheet_name=None)
                if sheet in df:
                    df = df[sheet]
                else:
                    df = list(df.values())[0]
            else:
                df = pd.read_csv(source_path)
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Failed to read Excel: {str(e)}'})

        col_name = 'row_id'
        if prefix:
            col_name = f"{prefix}_row_id" 
        row_id_column = pd.DataFrame({col_name: np.arange(1, len(df) + 1)})
        df = pd.concat([row_id_column, df], axis=1)

        df.to_excel(preprocess_file_path, index=False)
        file_url = os.path.join(settings.MEDIA_URL, preprocess_folder_name, 'preprocess_' + filename).replace("\\", "/")

        return JsonResponse({
            'success': True,
            'message': f"Column '{col_name}' added with unique IDs.",
            'columns': df.columns.tolist(),
            'rows': df.fillna("").astype(str).to_dict(orient='records'),
            'file_url': file_url,
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
    
@csrf_exempt
def save_preprocessed_file_api(request):
    print("in") 
    if request.method == 'POST':
        try:
            user_id = request.headers.get('userID')
  
            if not user_id:
                return JsonResponse({'success': False, 'error': 'User ID not provided.'})
            print(f"Received User ID: {user_id}")

            uploaded_file = request.FILES.get('file')
            # file_type = request.POST.get('file_type')  # 'survey' or 'preprocessed'

            if not uploaded_file:
                return JsonResponse({'success': False, 'error': 'No file uploaded.'})
            # if file_type not in ['survey', 'preprocessed']:
            #     return JsonResponse({'success': False, 'error': 'Invalid or missing file_type.'})

            # Define folder path
            folder_name = 'survey/' 
            folder_path = os.path.join(settings.MEDIA_ROOT, f'ID_{user_id}_uploads/temporary_uploads/', folder_name)
            os.makedirs(folder_path, exist_ok=True)

            # Handle CSV for survey type
            original_ext = os.path.splitext(uploaded_file.name)[1].lower()
            if original_ext == '.csv':
                df = pd.read_csv(uploaded_file)
                # Sanitize filename and change extension to .xlsx
                base_name = os.path.splitext(uploaded_file.name)[0]
                safe_base_name = "".join(c for c in base_name if c.isalnum() or c in (' ', '_', '-')).rstrip()
                file_name = f"{safe_base_name}.xlsx" 

                file_path = os.path.join(folder_path, file_name)
                df.to_excel(file_path, index=False)
            else:
                # Save original file
                file_name = uploaded_file.name
                file_path = os.path.join(folder_path, file_name)
                with open(file_path, 'wb+') as destination:
                    for chunk in uploaded_file.chunks():
                        destination.write(chunk)
            print(f"File saved to: {file_path}")

            file_url = os.path.join('/media', f'ID_{user_id}_uploads/temporary_uploads/', folder_name, file_name).replace('\\', '/')

            return JsonResponse({
                'success': True,
                'message': f"File saved as {file_name}",
                'file_url': file_url,
            })

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

import os
import shutil
import json
from django.conf import settings
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

# @csrf_exempt
# def save_results_api(request):
#     from django.http import JsonResponse
# from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def save_results_api(request): 
    if request.method == 'POST':
        print("✅ Save Results API hit")

        try:
            data = json.loads(request.body)
            image_paths = data.get('image_paths', [])
            test_name = data.get('test_name', 'general_test')
            user_id = data.get('user_id', 'anonymous')  # optional
            Excel_filename = os.path.splitext(os.path.basename(build_safe_filename(data.get('filename', ''))))[0]

 
            # print(test_name)
            # print(user_id) 
            # print(Excel_filename)  


            save_dir=os.path.join(settings.MEDIA_ROOT,f'ID_{user_id}_uploads/saved_files/', Excel_filename, test_name)
            os.makedirs(save_dir, exist_ok=True)
           

            saved_files = []
            for path in image_paths:
                print(path) 
                full_path = os.path.join(settings.BASE_DIR, path.lstrip('/'))  # Ensure path is absolute
                if os.path.exists(full_path):
                    filename = os.path.basename(full_path)
                    dest_path = os.path.join(save_dir, filename)
                    shutil.copy(full_path, dest_path)
                    saved_files.append(dest_path)
                else:
                    return JsonResponse({
                        'status': 'error',
                        'message': f'File not found: {path}'
                    }, status=404)

            return JsonResponse({
                'status': 'success',
                'saved_to': save_dir,
                'saved_files': saved_files
            })

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
                
# save edited file

from django.core.files.storage import default_storage
@csrf_exempt
def save_edited_excel(request):
    if request.method == "POST" and request.FILES.get("file"):
        excel_file = request.FILES["file"]
        user_id = request.POST.get("user_id")
        original_path = request.POST.get("original_path")
        original_path=os.path.join(settings.MEDIA_ROOT, original_path.replace("/media/", ""))
        replace= request.POST.get("replace_original")
        # print(replace)
        # print(excel_file.name) 

        try:
            if replace== "true": 
                # print("innn")
                if original_path and default_storage.exists(original_path):
                    save_path = original_path
            else:
                # fallback to saving in media/edited_excels/
                save_path = os.path.join(settings.MEDIA_ROOT,  f'ID_{user_id}_uploads/saved_files/', excel_file.name)
            print(save_path) 
            os.makedirs(os.path.dirname(save_path), exist_ok=True) 

            # overwrite or create new file
            with default_storage.open(save_path, "wb+") as destination:
                for chunk in excel_file.chunks():
                    destination.write(chunk)
            
            print(save_path) 

            return JsonResponse({
                "success": True,
                "message": "Excel file replaced successfully.",
                "saved_path": save_path
            })

        except Exception as e:
            return JsonResponse({
                "success": False,
                "error": str(e) 
            }, status=500)
        

    return JsonResponse({"success": False, "error": "Invalid request."}, status=400)

from django.http import FileResponse
from django.views.decorators.http import require_GET
from urllib.parse import unquote

@require_GET
def list_user_files(request):
    user_id = request.GET.get("user_id")
    if not user_id:
        return JsonResponse({"error": "Missing user_id"}, status=400)

    # Optional subfolder navigation
    subpath = request.GET.get("path", "")
    subpath = unquote(subpath)

    user_dir = os.path.join(settings.MEDIA_ROOT, f"ID_{user_id}_uploads", "saved_files")
    target_dir = os.path.join(user_dir, subpath)

    if not os.path.exists(target_dir):
        return JsonResponse({"error": "Folder not found"}, status=404)

    file_list = []
    for name in os.listdir(target_dir):
        full_path = os.path.join(target_dir, name)
        file_type = "folder" if os.path.isdir(full_path) else os.path.splitext(name)[1][1:]
        file_list.append({
            "name": name,
            "type": file_type,
            "size": os.path.getsize(full_path),
        })

    return JsonResponse(file_list, safe=False)

from django.http import FileResponse, JsonResponse, HttpResponse
from django.views.decorators.http import require_GET
from urllib.parse import unquote
import mimetypes
import os
from django.conf import settings

@require_GET
def serve_user_file(request, filename):
    user_id = request.GET.get("user_id")
    if not user_id:
        return JsonResponse({"error": "Missing user_id"}, status=400)

    subpath = unquote(request.GET.get("path", ""))

    # Sanitize subpath
    if ".." in subpath or subpath.startswith("/") or "\\" in subpath:
        return JsonResponse({"error": "Invalid path"}, status=400)

    user_dir = os.path.join(settings.MEDIA_ROOT, f"ID_{user_id}_uploads", "saved_files")
    file_path = os.path.normpath(os.path.join(user_dir, subpath, filename))

    if not file_path.startswith(user_dir):
        return JsonResponse({"error": "Access denied"}, status=403)

    if not os.path.exists(file_path):
        return JsonResponse({"error": "File not found"}, status=404)

    mime_type, _ = mimetypes.guess_type(file_path)
    if not mime_type:
        mime_type = "application/octet-stream"

  
    with open(file_path, "rb") as f:
        response = HttpResponse(f.read(), content_type=mime_type)
        response["Content-Disposition"] = f'attachment; filename="{os.path.basename(file_path)}"'
        response["Content-Length"] = os.path.getsize(file_path)
        response["X-Content-Type-Options"] = "nosniff"
        return response

import shutil
@csrf_exempt 



def delete_temp_folder(request):    

    if(request.method == "POST"):
        user_id = request.POST.get("user_id")
        print(user_id) 

    
        temp_folder = os.path.join(settings.MEDIA_ROOT,f'ID_{user_id}_uploads/', 'temporary_uploads')

        try:
            if os.path.exists(temp_folder):
                shutil.rmtree(temp_folder)  
            os.makedirs(temp_folder, exist_ok=True)  
            return JsonResponse({"status": "success", "message": "Temporary folder deleted"})
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=500)